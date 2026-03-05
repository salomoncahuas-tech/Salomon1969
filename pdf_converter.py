"""
IN Piura - Modulo Conversor PDF -> Excel
Extrae tablas de reportes PDF del proyecto y genera archivos Excel
con formato institucional ANIN y coordenadas UTM WGS84 Zona 17S.

Tipos de reporte soportados:
  - centros_poblados   : listas de CCPP con coordenadas
  - bloques_intervencion: tabla de bloques y areas
  - areas_conservacion : ACR / ACP con intersecciones
  - catastro_minero    : bloques catastro vs comunidades
  - areas_degradadas   : superficie degradada por ecosistema
  - meteorologico      : estaciones con datos hidro/meteo
  - generico           : cualquier tabla (fallback)

Dependencias: pdfplumber, openpyxl, pyproj, pandas
"""

import io
import re
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import pdfplumber
    PDFPLUMBER_OK = True
except ImportError:
    PDFPLUMBER_OK = False

try:
    from pyproj import Transformer
    PYPROJ_OK = True
except ImportError:
    PYPROJ_OK = False


# ── Paleta ANIN ──────────────────────────────────────────────────────────
COLOR_HEADER     = "1B4D2E"   # Verde oscuro institucional
COLOR_SUBHEADER  = "2E7D4F"   # Verde medio
COLOR_ALT_ROW    = "EAF4EE"   # Verde muy claro filas alternas
COLOR_UTM_BG     = "D4EDDA"   # Fondo columnas UTM
COLOR_TOTAL_BG   = "A8D5B5"   # Fondo fila de totales
COLOR_WHITE      = "FFFFFF"
COLOR_CAPTION    = "888888"


def _thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(cell, text, bg=COLOR_HEADER, fg=COLOR_WHITE, size=10):
    cell.value = text
    cell.font = Font(name="Arial", size=size, bold=True, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin()


def _data_cell(cell, value, alt=False, utm=False):
    cell.value = value
    if utm:
        bg = COLOR_UTM_BG
    elif alt:
        bg = COLOR_ALT_ROW
    else:
        bg = COLOR_WHITE
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Arial", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell.border = _thin()


# ── Conversion de coordenadas ─────────────────────────────────────────────

def latlon_a_utm17s(lat, lon):
    """
    Convierte coordenadas geograficas WGS84 a UTM Zona 17S.
    Retorna (este, norte) o (None, None) si falla la validacion.
    """
    if not PYPROJ_OK:
        return None, None
    try:
        t = Transformer.from_crs("EPSG:4326", "EPSG:32717", always_xy=True)
        este, norte = t.transform(float(lon), float(lat))
        # Validar rangos de la cuenca del Piura
        if 450_000 <= este <= 750_000 and 9_300_000 <= norte <= 9_600_000:
            return round(este, 2), round(norte, 2)
        return None, None
    except Exception:
        return None, None


def _parse_num(val):
    """Convierte un valor a float, ignorando errores."""
    if val is None:
        return None
    try:
        return float(str(val).replace(",", ".").replace(" ", "").strip())
    except (ValueError, TypeError):
        return None


# ── Deteccion automatica del tipo de reporte ──────────────────────────────

TIPOS_REPORTE = {
    "centros_poblados":     ["nomcp", "nom_cp", "centro poblado", "cpinei", "ubigeo", "nomcp"],
    "bloques_intervencion": ["bloque", "microcuenca", "area_ha", "intervencion", "m10", "m1b"],
    "areas_conservacion":   ["acr", "acp", "conservacion", "denominacion", "resolucion"],
    "catastro_minero":      ["minero", "catastro", "concesion", "titular", "derecho minero"],
    "areas_degradadas":     ["degradad", "ecosistema", "categoria degradacion", "superficie"],
    "meteorologico":        ["estacion", "precipitacion", "temperatura", "altitud", "hidro"],
}


def detectar_tipo(texto: str, nombre_archivo: str = "") -> str:
    """Detecta el tipo de reporte a partir del texto extraido."""
    fuente = (texto + " " + nombre_archivo).lower()
    scores = {tipo: sum(1 for kw in kws if kw in fuente)
              for tipo, kws in TIPOS_REPORTE.items()}
    mejor = max(scores, key=scores.get)
    return mejor if scores[mejor] >= 2 else "generico"


# ── Extraccion de tablas desde PDF ────────────────────────────────────────

def extraer_pdf(pdf_bytes: bytes) -> tuple:
    """
    Extrae todas las tablas y el texto de un PDF.
    Retorna (lista_dataframes, texto_completo).
    """
    if not PDFPLUMBER_OK:
        raise ImportError("pdfplumber no esta instalado. Agrega 'pdfplumber' a requirements.txt.")

    tablas, texto_paginas = [], []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for pagina in pdf.pages:
            txt = pagina.extract_text() or ""
            texto_paginas.append(txt)

            for tabla in pagina.extract_tables():
                if not tabla or len(tabla) < 2:
                    continue
                # Normalizar cabecera
                raw_headers = [str(h).strip().replace("\n", " ") if h else f"COL_{i}"
                               for i, h in enumerate(tabla[0])]
                # Resolver duplicados en cabecera
                seen, headers = {}, []
                for h in raw_headers:
                    h_clean = h if h else "COL"
                    if h_clean in seen:
                        seen[h_clean] += 1
                        headers.append(f"{h_clean}_{seen[h_clean]}")
                    else:
                        seen[h_clean] = 0
                        headers.append(h_clean)

                filas = [[str(c).strip() if c is not None else "" for c in fila]
                         for fila in tabla[1:]]
                df = pd.DataFrame(filas, columns=headers)
                # Eliminar filas completamente vacias
                df = df[df.apply(lambda r: any(v.strip() for v in r.astype(str)), axis=1)]
                if not df.empty:
                    tablas.append(df)

    return tablas, "\n".join(texto_paginas)


# ── Enriquecimiento con coordenadas UTM ───────────────────────────────────

def _agregar_utm(df: pd.DataFrame) -> pd.DataFrame:
    """
    Detecta columnas de latitud/longitud y agrega UTM Zona 17S.
    Opera sobre una copia del DataFrame.
    """
    if not PYPROJ_OK:
        return df

    cols_upper = {c.upper(): c for c in df.columns}

    lat_keys = ["LAT", "LATITUD", "Y", "COORD_Y", "CENTROI_2", "LAT_DD"]
    lon_keys = ["LON", "LONG", "LONGITUD", "X", "COORD_X", "CENTROI", "LON_DD"]

    lat_col = next((cols_upper[k] for k in lat_keys if k in cols_upper), None)
    lon_col = next((cols_upper[k] for k in lon_keys if k in cols_upper), None)

    if not lat_col or not lon_col:
        return df

    estes, nortes = [], []
    for _, row in df.iterrows():
        lat = _parse_num(row.get(lat_col))
        lon = _parse_num(row.get(lon_col))
        if lat is not None and lon is not None:
            e, n = latlon_a_utm17s(lat, lon)
            estes.append(e)
            nortes.append(n)
        else:
            estes.append(None)
            nortes.append(None)

    df = df.copy()
    df["UTM_ESTE_17S"]  = estes
    df["UTM_NORTE_17S"] = nortes
    df["ZONA_UTM"]      = "17S"
    return df


# ── Procesadores especializados ───────────────────────────────────────────

def _proc_centros_poblados(tablas, texto):
    if not tablas:
        return pd.DataFrame()
    df = pd.concat(tablas, ignore_index=True)
    df.columns = [c.upper().strip().replace(" ", "_").replace("\n", "_")
                  for c in df.columns]
    df = _agregar_utm(df)
    # Deduplicar por nombre + bloque si existen
    nom  = next((c for c in df.columns if "NOM" in c), None)
    blq  = next((c for c in df.columns if "BLOQUE" in c or "BLOQ" in c), None)
    if nom:
        subset = [c for c in [nom, blq] if c]
        df = df.drop_duplicates(subset=subset)
    return df.reset_index(drop=True)


def _proc_generico(tablas, texto):
    if not tablas:
        lineas = [l for l in texto.split("\n") if l.strip()]
        return pd.DataFrame({"CONTENIDO": lineas})
    df = pd.concat(tablas, ignore_index=True)
    df.columns = [c.upper().strip().replace(" ", "_").replace("\n", "_")
                  for c in df.columns]
    df = _agregar_utm(df)
    return df.reset_index(drop=True)


PROCESADORES = {
    "centros_poblados":     _proc_centros_poblados,
    "bloques_intervencion": _proc_generico,
    "areas_conservacion":   _proc_generico,
    "catastro_minero":      _proc_generico,
    "areas_degradadas":     _proc_generico,
    "meteorologico":        _proc_generico,
    "generico":             _proc_generico,
}


# ── Punto de entrada principal ────────────────────────────────────────────

def procesar_pdf(pdf_bytes: bytes, nombre_archivo: str = "",
                 forzar_tipo: str = "") -> dict:
    """
    Procesa un PDF completo y retorna un diccionario con:
      - tipo        : tipo de reporte detectado
      - df          : DataFrame con los datos limpios
      - n_registros : cantidad de filas
      - n_tablas    : tablas encontradas en el PDF
      - utm_ok      : True si se agregaron columnas UTM
      - texto_prev  : primeros 500 chars del texto extraido
    """
    tablas, texto = extraer_pdf(pdf_bytes)
    tipo = forzar_tipo if forzar_tipo else detectar_tipo(texto, nombre_archivo)
    proc = PROCESADORES.get(tipo, _proc_generico)
    df   = proc(tablas, texto)

    utm_cols = [c for c in df.columns if "UTM" in c or "ZONA_UTM" in c]

    return {
        "tipo":        tipo,
        "df":          df,
        "n_registros": len(df),
        "n_tablas":    len(tablas),
        "utm_ok":      len(utm_cols) > 0,
        "texto_prev":  texto[:600].strip(),
    }


# ── Generacion de Excel con formato ANIN ─────────────────────────────────

ETIQUETAS_TIPO = {
    "centros_poblados":     "Centros Poblados",
    "bloques_intervencion": "Bloques de Intervencion",
    "areas_conservacion":   "Areas de Conservacion",
    "catastro_minero":      "Catastro Minero",
    "areas_degradadas":     "Areas Degradadas",
    "meteorologico":        "Estaciones Hidrometeorologicas",
    "generico":             "Reporte General",
}


def generar_excel(df: pd.DataFrame, tipo: str,
                  nombre_origen: str, opciones: dict = None) -> bytes:
    """
    Genera un Excel con formato institucional ANIN.
    opciones = {
        'deduplicar': bool,        # eliminar duplicados
        'solo_con_utm': bool,      # filtrar filas sin coordenadas
        'hoja_nombre': str,        # nombre de la hoja de calculo
    }
    Retorna bytes del archivo .xlsx.
    """
    if opciones is None:
        opciones = {}

    df = df.copy()

    # Aplicar opciones de limpieza
    if opciones.get("deduplicar", False) and not df.empty:
        df = df.drop_duplicates()

    utm_cols = [c for c in df.columns if "UTM" in c or "ZONA_UTM" in c]
    if opciones.get("solo_con_utm", False) and utm_cols:
        este_col = next((c for c in utm_cols if "ESTE" in c), None)
        if este_col:
            df = df[df[este_col].notna()]

    df = df.reset_index(drop=True)
    etiqueta   = ETIQUETAS_TIPO.get(tipo, "Reporte")
    hoja_nom   = opciones.get("hoja_nombre", etiqueta)[:31]
    utm_idxs   = {i + 1 for i, c in enumerate(df.columns)
                  if any(k in c.upper() for k in ["UTM_ESTE", "UTM_NORTE", "ZONA_UTM"])}

    wb = Workbook()
    ws = wb.active
    ws.title = hoja_nom
    ws.sheet_view.showGridLines = False

    n_cols    = max(len(df.columns), 3)
    last_col  = get_column_letter(n_cols)
    DATA_START = 5   # fila donde empiezan los headers

    # ── Fila 1: nombre institucional ─────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.merge_cells(f"A1:{last_col}1")
    _header_cell(ws["A1"],
                 "AUTORIDAD NACIONAL DE INFRAESTRUCTURA - ANIN",
                 bg=COLOR_HEADER, size=12)

    # ── Fila 2: division ─────────────────────────────────────────────
    ws.row_dimensions[2].height = 16
    ws.merge_cells(f"A2:{last_col}2")
    _header_cell(ws["A2"],
                 "DIRECCION DE INTERVENCIONES MULTISECTORIALES Y DE EMERGENCIA - DIME",
                 bg=COLOR_SUBHEADER, size=10)

    # ── Fila 3: titulo del reporte ───────────────────────────────────
    ws.row_dimensions[3].height = 18
    ws.merge_cells(f"A3:{last_col}3")
    _header_cell(ws["A3"],
                 f"PROYECTO IN PIURA — {etiqueta.upper()}",
                 bg=COLOR_SUBHEADER, size=11)

    # ── Fila 4: metadatos ─────────────────────────────────────────────
    ws.row_dimensions[4].height = 13
    ws.merge_cells(f"A4:{last_col}4")
    ws["A4"].value = (
        f"Origen: {nombre_origen}   |   "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   "
        f"Sistema: UTM WGS84 Zona 17S   |   "
        f"Registros: {len(df)}"
    )
    ws["A4"].font      = Font(name="Arial", size=8, italic=True, color=COLOR_CAPTION)
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Fila 5: cabecera de columnas ──────────────────────────────────
    ws.row_dimensions[DATA_START].height = 32
    for ci, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=DATA_START, column=ci)
        is_utm = ci in utm_idxs
        bg = COLOR_UTM_BG if is_utm else COLOR_HEADER
        fg = "1B4D2E"      if is_utm else COLOR_WHITE
        _header_cell(cell, col.replace("_", " "), bg=bg, fg=fg, size=9)

    # ── Filas de datos ────────────────────────────────────────────────
    for ri, (_, row) in enumerate(df.iterrows(), start=DATA_START + 1):
        ws.row_dimensions[ri].height = 13
        alt = ri % 2 == 0
        for ci, val in enumerate(row, start=1):
            cell    = ws.cell(row=ri, column=ci)
            is_utm  = ci in utm_idxs
            # Formatear numeros UTM
            if is_utm and val is not None and str(val).strip() not in ("", "None"):
                try:
                    val = round(float(str(val)), 2)
                    cell.number_format = "#,##0.00"
                except (ValueError, TypeError):
                    pass
            _data_cell(cell, val, alt=alt, utm=is_utm)

    # ── Fila de totales (columna de area si existe) ───────────────────
    area_ci = next(
        (ci for ci, c in enumerate(df.columns, start=1)
         if any(k in c.upper() for k in ["AREA", "HA", "HECTAREA", "SUPERFICIE", "INTERSEC"])),
        None,
    )
    total_row = DATA_START + len(df) + 1
    ws.row_dimensions[total_row].height = 15

    # Celda "TOTAL" en la primera columna
    t0 = ws.cell(row=total_row, column=1)
    t0.value = f"TOTAL  ({len(df)} registros)"
    t0.font  = Font(name="Arial", size=9, bold=True, color=COLOR_WHITE)
    t0.fill  = PatternFill("solid", fgColor=COLOR_SUBHEADER)
    t0.alignment = Alignment(horizontal="center", vertical="center")
    t0.border = _thin()

    if area_ci:
        try:
            area_vals  = pd.to_numeric(df.iloc[:, area_ci - 1], errors="coerce")
            total_area = area_vals.sum()
            tc = ws.cell(row=total_row, column=area_ci)
            tc.value  = round(total_area, 4)
            tc.number_format = "#,##0.0000"
            tc.font   = Font(name="Arial", size=9, bold=True, color=COLOR_WHITE)
            tc.fill   = PatternFill("solid", fgColor=COLOR_SUBHEADER)
            tc.alignment = Alignment(horizontal="center", vertical="center")
            tc.border = _thin()
        except Exception:
            pass

    # ── Ajuste de anchos ───────────────────────────────────────────────
    for ci, col in enumerate(df.columns, start=1):
        letra = get_column_letter(ci)
        max_contenido = df.iloc[:, ci - 1].astype(str).str.len().max() if not df.empty else 0
        ancho = min(max(len(col) + 2, int(max_contenido) + 2, 10), 40)
        ws.column_dimensions[letra].width = ancho

    ws.freeze_panes = f"A{DATA_START + 1}"

    # ── Nota metodologica al pie ───────────────────────────────────────
    footer_row = total_row + 2
    ws.merge_cells(f"A{footer_row}:{last_col}{footer_row}")
    ws[f"A{footer_row}"].value = (
        "Nota metodologica: Coordenadas UTM WGS84 Zona 17S. "
        "ESTE: 450,000-750,000 m | NORTE: 9,300,000-9,600,000 m. "
        "Fuente: Proyecto IN Piura CUI 2669244 | ANIN-DIME-SESDI 2026."
    )
    ws[f"A{footer_row}"].font = Font(name="Arial", size=7, italic=True, color=COLOR_CAPTION)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Utilidad: resumen de validacion de coordenadas ────────────────────────

def validar_utm(df: pd.DataFrame) -> dict:
    """
    Valida las columnas UTM de un DataFrame y retorna estadisticas.
    """
    este_col  = next((c for c in df.columns if "UTM_ESTE" in c.upper()), None)
    norte_col = next((c for c in df.columns if "UTM_NORTE" in c.upper()), None)

    if not este_col or not norte_col:
        return {"estado": "sin_utm", "total": len(df), "validos": 0, "invalidos": 0}

    estes  = pd.to_numeric(df[este_col],  errors="coerce")
    nortes = pd.to_numeric(df[norte_col], errors="coerce")

    validos = int(
        ((estes  >= 450_000) & (estes  <= 750_000) &
         (nortes >= 9_300_000) & (nortes <= 9_600_000)).sum()
    )

    return {
        "estado":   "ok",
        "total":    len(df),
        "validos":  validos,
        "invalidos": len(df) - validos,
        "cobertura": f"{validos / len(df) * 100:.1f}%" if len(df) > 0 else "0%",
    }

"""
IN Piura - Plan de Ingreso / Verificación de Campo
Módulo de generación de reportes PDF y Excel.
Incluye fichas de inspección, resumen de bloques, presupuesto y cronograma.
Cuenca alta del río Piura, Perú.
"""

import json
import os
from datetime import datetime

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

import database as db

REPORTES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reportes")


def _asegurar_directorio():
    os.makedirs(REPORTES_DIR, exist_ok=True)


# ── Estilos comunes Excel ────────────────────────────────────────────────

def _estilos_excel():
    encabezado_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    encabezado_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return encabezado_font, encabezado_fill, borde, centrado


# ── Reporte PDF - Ficha de Inspección por Bloque ──────────────────────────

class FichaInspeccionPDF(FPDF):
    """PDF personalizado para fichas de inspección IN Piura."""

    def header(self):
        self.set_font("Helvetica", "B", 14)
        self.cell(0, 8, "IN Piura - Plan de Ingreso - Verificacion de Campo", 0, 1, "C")
        self.set_font("Helvetica", "", 9)
        self.cell(0, 5, "Restauracion de Ecosistemas - Cuenca Alta del Rio Piura", 0, 1, "C")
        self.line(10, self.get_y() + 2, 200, self.get_y() + 2)
        self.ln(6)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 10,
                  f"Pagina {self.page_no()}/{{nb}}  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                  0, 0, "C")

    def _seccion(self, titulo):
        self.set_font("Helvetica", "B", 11)
        self.set_fill_color(44, 62, 80)
        self.set_text_color(255, 255, 255)
        self.cell(0, 7, f"  {titulo}", 0, 1, fill=True)
        self.set_text_color(0, 0, 0)
        self.ln(2)

    def _campo(self, etiqueta, valor, ancho_etiqueta=55):
        x = self.get_x()
        y = self.get_y()
        self.set_font("Helvetica", "B", 9)
        self.set_xy(x, y)
        self.cell(ancho_etiqueta, 6, etiqueta + ":")
        self.set_font("Helvetica", "", 9)
        texto = str(valor) if valor else "-"
        self.cell(0, 6, texto, 0, 1)

    def _campo_largo(self, etiqueta, valor):
        self.set_font("Helvetica", "B", 9)
        self.cell(0, 6, etiqueta + ":", 0, 1)
        self.set_font("Helvetica", "", 9)
        texto = str(valor) if valor else "Sin observaciones."
        x_margin = self.l_margin
        self.set_x(x_margin)
        self.multi_cell(190, 5, texto)
        self.ln(1)

    def _subficha(self, titulo):
        self.set_font("Helvetica", "BI", 9)
        self.set_fill_color(230, 236, 245)
        self.cell(0, 5, f"  {titulo}", 0, 1, fill=True)
        self.set_font("Helvetica", "", 9)

    def _tabla_json(self, etiqueta, raw_json, columnas):
        """Renderiza una lista JSON como tabla simple en el PDF.
        `columnas` es una lista de tuplas (clave, encabezado, ancho_mm).
        Solo se imprimen filas con al menos un campo no vacio."""
        try:
            data = json.loads(raw_json) if raw_json else []
        except (TypeError, ValueError):
            data = []
        filas = [r for r in data if isinstance(r, dict) and any(
            (str(r.get(k, "") or "").strip()) for k, _, _ in columnas)]
        if not filas:
            return
        self.set_font("Helvetica", "B", 9)
        self.cell(0, 5, etiqueta + ":", 0, 1)
        # Encabezado de tabla
        self.set_font("Helvetica", "B", 8)
        self.set_fill_color(220, 220, 220)
        for _, header, ancho in columnas:
            self.cell(ancho, 5, str(header)[:30], 1, 0, "C", fill=True)
        self.ln(5)
        # Filas
        self.set_font("Helvetica", "", 8)
        for fila in filas:
            for clave, _, ancho in columnas:
                val = str(fila.get(clave, "") or "").strip()
                # Truncar para evitar desbordes
                max_chars = max(8, int(ancho / 1.6))
                if len(val) > max_chars:
                    val = val[: max_chars - 1] + "."
                self.cell(ancho, 5, val, 1, 0)
            self.ln(5)
        self.set_font("Helvetica", "", 9)
        self.ln(1)


def generar_ficha_pdf(bloque_id, inspeccion_id=None):
    """Genera un PDF con la ficha de inspección de un bloque."""
    _asegurar_directorio()

    bloque = db.obtener_bloque_por_id(bloque_id)
    if not bloque:
        raise ValueError(f"Bloque con id {bloque_id} no encontrado.")

    inspecciones = db.obtener_inspecciones_por_bloque(bloque_id)
    if inspeccion_id:
        inspecciones = [i for i in inspecciones if i["id"] == inspeccion_id]

    pdf = FichaInspeccionPDF()
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # Datos del bloque
    pdf._seccion("1. Datos del Bloque de Intervencion")
    pdf._campo("Codigo de bloque", bloque["codigo"])
    pdf._campo("Tipo de intervencion", bloque["tipo_intervencion"])
    pdf._campo("Cuenca hidrografica", bloque["cuenca"])
    microcuenca = bloque.get("microcuenca", "") or ""
    if microcuenca:
        pdf._campo("Codigo microcuenca", microcuenca)
    provincia = bloque.get("provincia", "") or ""
    if provincia:
        pdf._campo("Provincia", provincia)
    pdf._campo("Distrito", bloque["distrito"])
    pdf._campo("Coordenadas UTM Este", f"{bloque['utm_este']:.2f} m")
    pdf._campo("Coordenadas UTM Norte", f"{bloque['utm_norte']:.2f} m")
    pdf._campo("Zona UTM", bloque["utm_zona"])
    altitud = bloque.get("altitud", 0) or 0
    pdf._campo("Altitud", f"{altitud:.0f} m.s.n.m.")
    pdf._campo("Area", f"{bloque['area_hectareas']:.4f} ha")
    responsable = bloque.get("responsable", "") or ""
    if responsable:
        pdf._campo("Responsable", responsable)
    pdf._campo("Estado", bloque["estado"])
    pdf.ln(3)

    # Presupuesto del bloque
    presupuesto = db.obtener_presupuesto_por_bloque(bloque_id)
    if presupuesto:
        pdf._seccion("1.1 Resumen Presupuestal del Bloque")
        total_plan = sum(p["monto_planificado"] for p in presupuesto)
        total_ejec = sum(p["monto_ejecutado"] for p in presupuesto)
        pct = (total_ejec / total_plan * 100) if total_plan > 0 else 0

        for p in presupuesto:
            pct_p = (p["monto_ejecutado"] / p["monto_planificado"] * 100) if p["monto_planificado"] > 0 else 0
            pdf._campo(p["categoria"],
                       f"Plan: S/ {p['monto_planificado']:,.2f} | "
                       f"Ejec: S/ {p['monto_ejecutado']:,.2f} ({pct_p:.1f}%)")

        pdf.ln(1)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(0, 6,
                 f"TOTAL: Planificado S/ {total_plan:,.2f} | "
                 f"Ejecutado S/ {total_ejec:,.2f} ({pct:.1f}%)", 0, 1)
        pdf.ln(3)

    if not inspecciones:
        pdf._seccion("2. Inspecciones de Campo")
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(0, 6, "No se han registrado inspecciones para este bloque.", ln=True)
    else:
        for idx, insp in enumerate(inspecciones, 1):
            pdf._seccion(f"2.{idx} Inspeccion - {insp['fecha_visita']}")
            pdf._campo("Fecha de visita", insp["fecha_visita"])
            pdf._campo("Inspector", insp["inspector"])
            pdf._campo("Condiciones climaticas", insp["condiciones_climaticas"])
            pdf._campo("Avance fisico", f"{insp['avance_fisico']:.1f} %")
            pdf._campo("Codigo de verificacion", insp["codigo_verificacion"])
            pdf._campo_largo("Observaciones tecnicas", insp["observaciones"])
            pdf._campo_largo("Desviaciones observadas al Plan de Trabajo", insp["desviaciones"])

            if insp["registro_fotografico"]:
                pdf.set_font("Helvetica", "B", 9)
                pdf.cell(0, 6, "Registro fotografico:", ln=True)
                pdf.set_font("Helvetica", "", 8)
                for ruta in insp["registro_fotografico"].split(";"):
                    ruta = ruta.strip()
                    if ruta:
                        pdf.cell(10)
                        pdf.cell(0, 5, f"- {ruta}", ln=True)
            pdf.ln(2)

            # Indicadores de calidad de esta inspección
            indicadores = db.obtener_indicadores_por_inspeccion(insp["id"])
            if indicadores:
                pdf._seccion(f"   Indicadores de Calidad - Inspeccion {insp['fecha_visita']}")
                mc_ind = indicadores.get("microcuenca", "") or ""
                if mc_ind:
                    pdf._campo("Codigo microcuenca", mc_ind)
                pct_cob = indicadores.get("porcentaje_cobertura_vegetal", 0) or 0
                pdf._campo("Cobertura vegetal", f"{pct_cob:.1f} %")
                tipo_cob = indicadores.get("tipo_cobertura_vegetal", "") or ""
                if tipo_cob:
                    pdf._campo("Tipo cobertura vegetal", tipo_cob)
                vigor_cob = indicadores.get("vigor_cobertura_vegetal", "") or ""
                if vigor_cob:
                    pdf._campo("Vigor cobertura vegetal", vigor_cob)
                pdf.ln(3)

    # Diagnostico Territorial del bloque (V5: F-DT-01..F-DT-05)
    # Deduplicar: solo el mas reciente por (ficha, evaluador)
    diagnosticos_dt_raw = db.obtener_diagnosticos_por_bloque(bloque_id)
    diagnosticos_dt = []
    _dt_vistas = set()
    for dt in diagnosticos_dt_raw:
        fichas = dt.get("ficha", "")
        clave = (fichas, dt.get("evaluador", ""))
        if clave not in _dt_vistas:
            _dt_vistas.add(clave)
            diagnosticos_dt.append(dt)
    if diagnosticos_dt:
        pdf._seccion("3. Diagnostico Territorial (Plantilla V5)")
        for dt in diagnosticos_dt:
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(0, 6, f"  Ficha(s) {dt.get('ficha', '')} - {dt.get('fecha_evaluacion', '')}", 0, 1)
            pdf.set_font("Helvetica", "", 9)
            # Datos generales V5
            datos_gen = [
                ("Microcuenca", "microcuenca"),
                ("Evaluador", "evaluador"),
                ("Brigada", "brigada"),
                ("Correlativo ficha", "ficha_correlativo"),
                ("Hora de registro", "hora_registro"),
                ("Centro poblado cercano", "centro_poblado_cercano"),
                ("Comunidad campesina", "comunidad_campesina_dt"),
                ("Altitud GPS (m)", "altitud_gps"),
                ("UTM Este", "utm_este_dt"),
                ("UTM Norte", "utm_norte_dt"),
            ]
            for label, key in datos_gen:
                val = dt.get(key, "") or ""
                if val:
                    pdf._campo(label, val)

            # ── F-DT-01: Datos Generales y Fisiografia ──
            campos_dt01 = [
                ("Forma del terreno", "forma_terreno"),
                ("Rango de pendiente", "pendiente"),
                ("Posicion fisiografica", "posicion_fisiografica"),
                ("Exposicion / Orientacion", "exposicion_orientacion"),
                ("Rango altitudinal", "rango_altitudinal"),
                ("Paisaje dominante", "paisaje_dominante"),
                ("Afloramientos rocosos", "dt01_afloramientos_rocosos"),
                ("Escarpes activos", "dt01_escarpes_activos"),
                ("Reptacion de suelo", "dt01_reptacion_suelo"),
                ("Deslizamientos antiguos", "dt01_deslizamientos_antiguos"),
                ("Remociones en masa activas", "dt01_remociones_masa_activas"),
            ]
            if any(dt.get(k) for _, k in campos_dt01) or dt.get("dt01_observaciones"):
                pdf._subficha("F-DT-01: Datos Generales y Fisiografia del Bloque")
                for label, key in campos_dt01:
                    val = dt.get(key, "") or ""
                    if val:
                        pdf._campo(label, val)
                if dt.get("dt01_observaciones"):
                    pdf._campo_largo("Observaciones F-DT-01", dt["dt01_observaciones"])

            # ── F-DT-02: Suelo y Procesos Erosivos ──
            campos_dt02 = [
                ("Sellamiento / Costra", "dt02_sellamiento_costra"),
                ("Compactacion por pisoteo", "dt02_compactacion_pisoteo"),
                ("Raices expuestas", "dt02_raices_expuestas"),
                ("Nivel general de erosion", "dt02_nivel_erosion_general"),
                ("Nivel erosion (sintesis)", "dt02_nivel_erosion_sintesis"),
                ("N° de carcavas", "dt02_num_carcavas"),
                ("Longitud total carcavas (m)", "dt02_longitud_total_carcavas"),
                ("% bloque con carcavas", "dt02_pct_bloque_carcavas"),
                ("Erosion laminar (%)", "dt02_erosion_laminar_pct"),
                ("Patron de carcavas", "dt02_patron_carcavas"),
                ("Socavamiento de cauce", "dt02_socavamiento_cauce"),
                ("Urgencia de control", "dt02_urgencia_control"),
            ]
            tiene_dt02 = (any(dt.get(k) for _, k in campos_dt02)
                          or dt.get("dt02_carcavas_json") or dt.get("dt02_observaciones"))
            if tiene_dt02:
                pdf._subficha("F-DT-02: Suelo y Procesos Erosivos")
                for label, key in campos_dt02:
                    val = dt.get(key, "") or ""
                    if val:
                        pdf._campo(label, val)
                pdf._tabla_json(
                    "Inventario de carcavas / surcos",
                    dt.get("dt02_carcavas_json", ""),
                    [
                        ("codigo", "Codigo", 18),
                        ("tipo", "Tipo", 22),
                        ("utm_e_ini", "UTM E ini", 22),
                        ("utm_n_ini", "UTM N ini", 22),
                        ("longitud_m", "Long.(m)", 16),
                        ("prof_m", "Prof.(m)", 16),
                        ("ancho_m", "Ancho(m)", 16),
                        ("estado", "Estado", 22),
                        ("causa", "Causa", 35),
                    ],
                )
                if dt.get("dt02_observaciones"):
                    pdf._campo_largo("Observaciones F-DT-02", dt["dt02_observaciones"])

            # ── F-DT-03: Ecosistema ──
            campos_dt03 = [
                ("Parcela de muestreo", "dt03_parcela_muestreo"),
                ("Dimensiones parcela (m)", "dt03_dim_parcela"),
                ("Pendiente parcela (%)", "dt03_pendiente_parcela"),
                ("Cobertura vegetal total (%)", "dt03_cobertura_total"),
                ("Tipo de ecosistema (MINAM)", "dt03_tipo_ecosistema"),
                ("Superficie ecosistema (ha)", "dt03_superficie_ecosistema"),
                ("Estado conservacion", "dt03_estado_conservacion_eco"),
                ("Uso dominante del suelo", "dt03_uso_dominante"),
                ("Cobertura dosel (%)", "dt03_cobertura_dosel"),
                ("Cobertura arbustiva (%)", "dt03_cobertura_arbustiva"),
                ("Cobertura herbacea (%)", "dt03_cobertura_herbacea"),
                ("Cobertura hojarasca (%)", "dt03_cobertura_hojarasca"),
                ("Suelo desnudo (%)", "dt03_suelo_desnudo"),
                ("Altura estrato dom. (m)", "dt03_altura_estrato_dom"),
                ("Altura maxima (m)", "dt03_altura_max"),
                ("DAP promedio (cm)", "dt03_dap_promedio"),
                ("Regeneracion natural", "dt03_regeneracion_natural"),
                ("Estado sanitario", "dt03_estado_sanitario"),
                ("Presencia epifitas", "dt03_presencia_epifitas"),
                ("Fenologia dominante", "dt03_fenologia_dominante"),
                ("Tipo cobertura dominante", "dt03_tipo_cobertura_dom"),
            ]
            tiene_dt03 = (any(dt.get(k) for _, k in campos_dt03)
                          or dt.get("dt03_floristica_json")
                          or dt.get("dt03_especies_clave_json")
                          or dt.get("dt03_observaciones"))
            if tiene_dt03:
                pdf._subficha("F-DT-03: Ecosistema - Composicion, Estructura y Valor Ecologico")
                for label, key in campos_dt03:
                    val = dt.get(key, "") or ""
                    if val:
                        pdf._campo(label, val)
                pdf._tabla_json(
                    "Composicion floristica",
                    dt.get("dt03_floristica_json", ""),
                    [
                        ("nombre_comun", "N. comun", 35),
                        ("nombre_cientifico", "N. cientifico", 45),
                        ("familia", "Familia", 28),
                        ("estrato", "Estrato", 20),
                        ("origen", "Origen", 18),
                        ("abundancia", "Abund.", 18),
                        ("dap_cm", "DAP", 12),
                        ("altura_m", "H(m)", 12),
                    ],
                )
                pdf._tabla_json(
                    "Especies clave / indicadoras",
                    dt.get("dt03_especies_clave_json", ""),
                    [
                        ("nombre", "Nombre", 45),
                        ("categoria", "Categoria", 30),
                        ("estado_uicn", "UICN/D.S.043", 28),
                        ("utm_e", "UTM E", 22),
                        ("utm_n", "UTM N", 22),
                        ("n_indiv", "N° ind.", 14),
                        ("foto", "Foto", 14),
                    ],
                )
                if dt.get("dt03_observaciones"):
                    pdf._campo_largo("Observaciones F-DT-03", dt["dt03_observaciones"])

            # ── F-DT-04: Causas e Indicadores de Degradacion ──
            campos_dt04 = [
                ("Causa subyacente principal", "dt04_causa_subyacente"),
                ("Velocidad de degradacion", "dt04_velocidad_degradacion"),
                ("Reversibilidad tecnica", "dt04_reversibilidad"),
                ("Urgencia de intervencion", "dt04_urgencia_intervencion"),
            ]
            tiene_dt04 = (any(dt.get(k) for _, k in campos_dt04)
                          or dt.get("dt04_causas_json")
                          or dt.get("dt04_indicadores_json")
                          or dt.get("dt04_causas_directas_texto")
                          or dt.get("dt04_observaciones"))
            if tiene_dt04:
                pdf._subficha("F-DT-04: Causas e Indicadores de Degradacion")
                pdf._tabla_json(
                    "Matriz de causas de degradacion",
                    dt.get("dt04_causas_json", ""),
                    [
                        ("n", "N°", 10),
                        ("causa", "Causa / Factor", 70),
                        ("presencia", "Pres.", 16),
                        ("intensidad", "Intens.", 22),
                        ("extension", "Ext.(%)", 18),
                        ("antiguedad", "Antig.", 16),
                        ("evidencia", "Evidencia", 38),
                    ],
                )
                pdf._tabla_json(
                    "Indicadores cuantitativos de degradacion",
                    dt.get("dt04_indicadores_json", ""),
                    [
                        ("n", "N°", 10),
                        ("indicador", "Indicador", 55),
                        ("unidad", "Unidad", 18),
                        ("valor", "Valor", 18),
                        ("fuente", "Fuente", 30),
                        ("umbral", "Umbral", 35),
                        ("nivel", "Nivel", 20),
                    ],
                )
                if dt.get("dt04_causas_directas_texto"):
                    pdf._campo_largo("Causas directas (sintesis)",
                                     dt["dt04_causas_directas_texto"])
                for label, key in campos_dt04:
                    val = dt.get(key, "") or ""
                    if val:
                        pdf._campo(label, val)
                if dt.get("dt04_observaciones"):
                    pdf._campo_largo("Observaciones F-DT-04", dt["dt04_observaciones"])

            # ── F-DT-05: Recursos Hidricos y Accesibilidad ──
            campos_dt05 = [
                ("Zona de recarga hidrica", "dt05_zona_recarga"),
                ("Humedad persistente", "dt05_humedad_persistente"),
                ("Escorrentia concentrada", "dt05_escorrentia_concentrada"),
                ("Distancia a captacion (m)", "dt05_dist_captacion"),
                ("JASS / captacion asociada", "dt05_jass_captacion"),
                ("Interferencia con riego", "dt05_interferencia_riego"),
                ("Sistema de riego", "dt05_sistema_riego_nombre"),
                ("Modalidad de acceso", "dt05_modalidad_acceso"),
                ("Via principal de acceso", "dt05_via_principal"),
                ("Tipo de via final", "dt05_tipo_via_final"),
                ("Transitabilidad - seca", "dt05_transitabilidad_seca"),
                ("Transitabilidad - lluviosa", "dt05_transitabilidad_lluviosa"),
                ("Tiempo desde capital distrital (min)", "dt05_tiempo_dist_capital"),
                ("Tiempo desde capital provincial (min)", "dt05_tiempo_prov_capital"),
                ("Senal celular", "dt05_senal_celular"),
                ("Operador celular dominante", "dt05_operador_celular"),
                ("Alojamiento rural disponible", "dt05_alojamiento"),
                ("Requiere autorizacion Ronda", "dt05_requiere_ronda"),
                ("Contacto Ronda", "dt05_contacto_ronda"),
            ]
            tiene_dt05 = (any(dt.get(k) for _, k in campos_dt05)
                          or dt.get("dt05_fuentes_agua_json")
                          or dt.get("dt05_observaciones"))
            if tiene_dt05:
                pdf._subficha("F-DT-05: Recursos Hidricos y Accesibilidad")
                pdf._tabla_json(
                    "Inventario de fuentes de agua",
                    dt.get("dt05_fuentes_agua_json", ""),
                    [
                        ("n", "N°", 10),
                        ("tipo", "Tipo", 28),
                        ("utm_e", "UTM E", 22),
                        ("utm_n", "UTM N", 22),
                        ("regimen", "Regimen", 22),
                        ("calidad", "Calidad", 22),
                        ("distancia_m", "Dist.(m)", 18),
                        ("uso_obs", "Uso/Obs.", 40),
                    ],
                )
                for label, key in campos_dt05:
                    val = dt.get(key, "") or ""
                    if val:
                        pdf._campo(label, val)
                if dt.get("dt05_observaciones"):
                    pdf._campo_largo("Observaciones F-DT-05", dt["dt05_observaciones"])

            # Observaciones generales del diagnostico
            if dt.get("observaciones_generales"):
                pdf._campo_largo("Observaciones generales", dt["observaciones_generales"])
            pdf.ln(2)

    # Diagnostico Social del bloque (deduplicar: solo el mas reciente por ficha)
    diagnosticos_ds_raw = db.obtener_diagnosticos_sociales_por_bloque(bloque_id)
    diagnosticos_ds = []
    _ds_vistas = set()
    for ds in diagnosticos_ds_raw:
        ficha = ds.get("ficha", "")
        clave = (ficha, ds.get("evaluador", ""))
        if clave not in _ds_vistas:
            _ds_vistas.add(clave)
            diagnosticos_ds.append(ds)
    if diagnosticos_ds:
        pdf._seccion("4. Diagnostico Social (F-DS-01..F-DS-05)")
        for ds in diagnosticos_ds:
            ficha_ds = ds.get("ficha", "")
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(0, 6, f"  Ficha {ficha_ds} - {ds.get('fecha_evaluacion', '')}", 0, 1)
            pdf.set_font("Helvetica", "", 9)
            # Datos generales (compartidos por todas las DS)
            datos_gen_ds = [
                ("Microcuenca", "microcuenca"),
                ("Evaluador / Responsable", "evaluador"),
                ("N° de ficha", "ficha_numero"),
                ("Provincia", "provincia"),
                ("Distrito", "distrito"),
                ("Centro poblado", "centro_poblado"),
                ("Comunidad campesina", "comunidad_campesina"),
                ("Coordenada UTM Este", "coordenada_este"),
                ("Coordenada UTM Norte", "coordenada_norte"),
                ("Altitud (msnm)", "altitud"),
                ("Codigo UBIGEO", "codigo_ubigeo"),
            ]
            for label, key in datos_gen_ds:
                val = ds.get(key, "") or ""
                if val and str(val).strip() not in ("0", "0.0"):
                    pdf._campo(label, val)

            # ── F-DS-01: Diagnostico Socioeconomico ──
            if ficha_ds == "F-DS-01":
                pdf._subficha("F-DS-01: Diagnostico Socioeconomico de Centro Poblado")
                campos_ds01 = [
                    ("N de familias / viviendas", "ds01_num_familias"),
                    ("Poblacion - Hombres", "ds01_poblacion_hombres"),
                    ("Poblacion - Mujeres", "ds01_poblacion_mujeres"),
                    ("Poblacion total", "ds01_poblacion_total"),
                    ("Idioma predominante", "ds01_idioma"),
                    ("Nivel educativo predominante", "ds01_nivel_educativo"),
                    ("Tasa de migracion", "ds01_tasa_migracion"),
                    ("Destino principal migracion", "ds01_destino_migracion"),
                    ("Organizacion comunal", "ds01_organizacion_comunal"),
                    ("Junta directiva vigente", "ds01_junta_directiva"),
                    ("Presidente/a de junta", "ds01_presidente_junta"),
                    ("Agua potable - tipo", "ds01_agua_potable_tipo"),
                    ("Agua potable - cobertura (%)", "ds01_agua_potable_cobertura"),
                    ("Saneamiento", "ds01_saneamiento"),
                    ("Energia - tipo", "ds01_energia_tipo"),
                    ("Energia - cobertura (%)", "ds01_energia_cobertura"),
                    ("Telecomunicaciones", "ds01_telecomunicaciones"),
                    ("Operador telecom", "ds01_telecom_operador"),
                    ("Acceso vial", "ds01_acceso_vial"),
                    ("Distancia a capital distrital (km)", "ds01_distancia_capital"),
                    ("Transporte", "ds01_transporte"),
                    ("Establecimiento de salud", "ds01_salud_tipo"),
                    ("Distancia a salud (km)", "ds01_salud_distancia"),
                    ("Institucion educativa", "ds01_educacion"),
                    ("Fuente principal de agua", "ds01_fuente_agua"),
                    ("Problemas con el agua", "ds01_problemas_agua"),
                    ("Uso de recursos forestales", "ds01_uso_recursos_forestales"),
                    ("Frecuencia uso forestal", "ds01_frecuencia_uso_forestal"),
                    ("Disposicion a participar", "ds01_disposicion_participar"),
                    ("Comentario disposicion", "ds01_comentario_disposicion"),
                    ("Area comunal (ha)", "ds01_tenencia_comunal_ha"),
                    ("Area privada (ha)", "ds01_tenencia_privada_ha"),
                    ("Area estatal (ha)", "ds01_tenencia_estatal_ha"),
                ]
                for label, key in campos_ds01:
                    val = ds.get(key, "") or ""
                    if val:
                        pdf._campo(label, val)
                if ds.get("ds01_percepcion_cambios"):
                    pdf._campo_largo("Percepcion de cambios ambientales",
                                     ds["ds01_percepcion_cambios"])
                if ds.get("ds01_activos_asociados"):
                    pdf._campo_largo("Activos asociados al bloque",
                                     ds["ds01_activos_asociados"])
                pdf._tabla_json(
                    "Actividades economicas",
                    ds.get("ds01_actividades_economicas", ""),
                    [
                        ("actividad", "Actividad", 45),
                        ("pct_familias", "% Fam.", 18),
                        ("productos", "Productos", 50),
                        ("destino", "Destino", 30),
                        ("ingreso", "Ingreso est.", 30),
                    ],
                )

            # ── F-DS-02: Actores clave ──
            elif ficha_ds == "F-DS-02":
                pdf._subficha("F-DS-02: Identificacion y Caracterizacion de Actores Clave")
                pdf._tabla_json(
                    "Registro de actores identificados",
                    ds.get("ds02_registro_actores", ""),
                    [
                        ("nombre", "Nombre / Organizacion", 50),
                        ("tipo", "Tipo", 25),
                        ("rol", "Rol / Funcion", 40),
                        ("relacion", "Rel.Proy.", 20),
                        ("influencia", "Influencia", 22),
                        ("interes", "Interes", 22),
                    ],
                )
                campos_ds02 = [
                    ("Gobierno Local", "ds02_actores_gob_local"),
                    ("Gobierno Regional", "ds02_actores_gob_regional"),
                    ("Gobierno Nacional", "ds02_actores_gob_nacional"),
                    ("Comunidades Campesinas", "ds02_actores_comunidades"),
                    ("Juntas de Usuarios / Riego", "ds02_actores_juntas_riego"),
                    ("Comites de Gestion / Cuenca", "ds02_actores_comites_cuenca"),
                    ("ONG / Cooperacion", "ds02_actores_ong"),
                    ("Empresa Privada", "ds02_actores_empresa"),
                    ("Instituciones Educativas", "ds02_actores_educacion"),
                    ("Organizaciones de Base", "ds02_actores_org_base"),
                ]
                for label, key in campos_ds02:
                    val = ds.get(key, "") or ""
                    if val:
                        pdf._campo(label, val)

            # ── F-DS-03: Entrevista semiestructurada ──
            elif ficha_ds == "F-DS-03":
                pdf._subficha("F-DS-03: Guia de Entrevista Semiestructurada")
                campos_ds03 = [
                    ("Nombre entrevistado/a", "ds03_nombre_entrevistado"),
                    ("Cargo / Funcion", "ds03_cargo_funcion"),
                    ("Institucion", "ds03_institucion"),
                    ("Telefono / Correo", "ds03_telefono_correo"),
                    ("Duracion entrevista", "ds03_duracion"),
                ]
                for label, key in campos_ds03:
                    val = ds.get(key, "") or ""
                    if val:
                        pdf._campo(label, val)
                # Respuestas (texto libre)
                resp_ds03 = [
                    ("1.1 Recursos naturales", "ds03_resp_recursos_naturales"),
                    ("1.2 Cambios ambientales", "ds03_resp_cambios_ambiente"),
                    ("1.3 Problemas ambientales", "ds03_resp_problemas_ambientales"),
                    ("1.4 Zonas de conservacion", "ds03_resp_zonas_conservacion"),
                    ("2.1 Actividades economicas", "ds03_resp_actividades_economicas"),
                    ("2.2 Abastecimiento de agua", "ds03_resp_abastecimiento_agua"),
                    ("2.3 Productos del bosque", "ds03_resp_productos_bosque"),
                    ("2.4 Cadenas productivas", "ds03_resp_cadenas_productivas"),
                    ("3.1 Organizaciones", "ds03_resp_organizaciones"),
                    ("3.2 Decisiones del territorio", "ds03_resp_decisiones_territorio"),
                    ("3.3 Conflictos", "ds03_resp_conflictos"),
                    ("3.4 Proyectos anteriores", "ds03_resp_proyectos_anteriores"),
                    ("3.5 Experiencia en reforestacion", "ds03_resp_experiencia_reforestacion"),
                    ("4.1 Conocimiento de restauracion", "ds03_resp_conocimiento_restauracion"),
                    ("4.2 Expectativas del proyecto", "ds03_resp_expectativas"),
                    ("4.3 Disposicion a participar", "ds03_resp_disposicion_participar"),
                    ("4.4 Condiciones / preocupaciones", "ds03_resp_condiciones"),
                    ("5.1 Conocimiento MERESE", "ds03_resp_conocimiento_merese"),
                    ("5.2 Beneficiarios", "ds03_resp_beneficiarios"),
                    ("5.3 Instituciones contribuyentes", "ds03_resp_instituciones_contribuyentes"),
                    ("5.4 Experiencias de pago / compensacion", "ds03_resp_experiencias_pago"),
                ]
                for label, key in resp_ds03:
                    val = ds.get(key, "") or ""
                    if val:
                        pdf._campo_largo(label, val)

            # ── F-DS-04: Taller participativo ──
            elif ficha_ds == "F-DS-04":
                pdf._subficha("F-DS-04: Acta de Taller Participativo")
                campos_ds04 = [
                    ("Lugar del taller", "ds04_lugar_taller"),
                    ("Convocante", "ds04_convocante"),
                    ("Hora de inicio", "ds04_hora_inicio"),
                    ("Hora de finalizacion", "ds04_hora_fin"),
                ]
                for label, key in campos_ds04:
                    val = ds.get(key, "") or ""
                    if val:
                        pdf._campo(label, val)
                if ds.get("ds04_objetivo"):
                    pdf._campo_largo("Objetivo", ds["ds04_objetivo"])
                pdf._tabla_json(
                    "Lista de participantes",
                    ds.get("ds04_lista_participantes", ""),
                    [
                        ("nombre", "Nombres y apellidos", 65),
                        ("dni", "DNI", 22),
                        ("institucion", "Inst./Comunidad", 50),
                        ("cargo", "Cargo", 25),
                        ("telefono", "Telefono", 25),
                    ],
                )
                if ds.get("ds04_presentacion"):
                    pdf._campo_largo("Presentacion del proyecto", ds["ds04_presentacion"])
                if ds.get("ds04_intervenciones"):
                    pdf._campo_largo("Principales intervenciones", ds["ds04_intervenciones"])
                if ds.get("ds04_preguntas_respuestas"):
                    pdf._campo_largo("Preguntas y respuestas", ds["ds04_preguntas_respuestas"])
                if ds.get("ds04_acuerdos"):
                    pdf._campo_largo("Acuerdos y compromisos", ds["ds04_acuerdos"])
                if ds.get("ds04_observaciones"):
                    pdf._campo_largo("Observaciones del taller", ds["ds04_observaciones"])

            # ── F-DS-05: Conflictos y oportunidades ──
            elif ficha_ds == "F-DS-05":
                pdf._subficha("F-DS-05: Conflictos y Oportunidades")
                pdf._tabla_json(
                    "Conflictos identificados",
                    ds.get("ds05_conflictos", ""),
                    [
                        ("tipo", "Tipo", 35),
                        ("actores", "Actores", 45),
                        ("nivel", "Nivel", 14),
                        ("estado", "Estado", 22),
                        ("descripcion", "Descripcion / Causa", 50),
                        ("impacto", "Impacto en proyecto", 35),
                    ],
                )
                pdf._tabla_json(
                    "Oportunidades identificadas",
                    ds.get("ds05_oportunidades", ""),
                    [
                        ("oportunidad", "Oportunidad", 50),
                        ("actores", "Actores", 40),
                        ("tipo", "Tipo", 25),
                        ("potencial", "Potencial", 22),
                        ("como_aprovechar", "Como aprovechar", 50),
                    ],
                )

            # Observaciones generales / archivos adjuntos
            if ds.get("observaciones_generales"):
                pdf._campo_largo("Observaciones generales", ds["observaciones_generales"])
            if ds.get("archivos_adjuntos"):
                pdf._campo("Archivos adjuntos", ds["archivos_adjuntos"])
            pdf.ln(2)

    # Elementos Expuestos del bloque (F-EE-01..F-EE-07)
    elementos_ee = db.obtener_elementos_expuestos_por_bloque(bloque_id)
    if elementos_ee:
        pdf._seccion("5. Elementos Expuestos / Analisis de Riesgos (F-EE-01..F-EE-07)")
        for ee in elementos_ee:
            ficha_ee = ee.get("ficha", "")
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(0, 6, f"  Ficha {ficha_ee} - {ee.get('fecha_campo', '')}", 0, 1)
            pdf.set_font("Helvetica", "", 9)
            # Datos generales EE
            datos_gen_ee = [
                ("Responsable / Brigada", "responsable_brigada"),
                ("Centro(s) poblado(s)", "centro_poblado"),
                ("UTM Este (m)", "coordenada_este"),
                ("UTM Norte (m)", "coordenada_norte"),
                ("Altitud (msnm)", "altitud"),
            ]
            for label, key in datos_gen_ee:
                val = ee.get(key, "") or ""
                if val and str(val).strip() not in ("0", "0.0"):
                    pdf._campo(label, val)

            # F-EE-01: Inventario de elementos expuestos
            if ficha_ee == "F-EE-01":
                pdf._subficha("F-EE-01: Inventario de Elementos Expuestos")
                pdf._tabla_json(
                    "Elementos registrados",
                    ee.get("ee01_registros", ""),
                    [
                        ("tipo_elemento", "Tipo", 30),
                        ("subtipo", "Subtipo", 30),
                        ("nombre", "Nombre", 35),
                        ("ubicacion_peligro", "Ubic. peligro", 28),
                        ("estado", "Estado", 18),
                        ("distancia_bloque", "Dist.(m)", 16),
                        ("beneficiarios", "N benef.", 16),
                        ("material", "Material", 25),
                    ],
                )

            # F-EE-02: Poblacion y viviendas
            elif ficha_ee == "F-EE-02":
                pdf._subficha("F-EE-02: Poblacion y Viviendas")
                pdf._tabla_json(
                    "Centros poblados en area de peligro",
                    ee.get("ee02_registros", ""),
                    [
                        ("centro_poblado", "Centro poblado", 45),
                        ("viviendas_total", "Viv. tot.", 18),
                        ("viviendas_peligro", "Viv. peligro", 22),
                        ("poblacion_total", "Pob. total", 20),
                        ("material_viviendas", "Material", 28),
                        ("agua_potable", "Agua", 14),
                        ("electricidad", "Elec.", 14),
                        ("antecedente_evento", "Antec.", 16),
                        ("nivel_danio", "Danio", 18),
                    ],
                )

            # F-EE-03: Infraestructura publica
            elif ficha_ee == "F-EE-03":
                pdf._subficha("F-EE-03: Infraestructura Publica Expuesta")
                pdf._tabla_json(
                    "Infraestructuras",
                    ee.get("ee03_registros", ""),
                    [
                        ("sector", "Sector", 24),
                        ("tipo_infraestructura", "Tipo", 35),
                        ("nombre", "Nombre", 40),
                        ("estado", "Estado", 18),
                        ("nivel_exposicion", "Nivel exp.", 22),
                        ("tipo_peligro", "Peligro", 28),
                        ("antecedente_danio", "Antec.", 16),
                        ("costo_activo", "S/ activo", 22),
                        ("costo_reposicion", "S/ reposic.", 22),
                    ],
                )

            # F-EE-04: Actividades economicas
            elif ficha_ee == "F-EE-04":
                pdf._subficha("F-EE-04: Actividades Economicas y Agropecuarias")
                pdf._tabla_json(
                    "Actividades expuestas",
                    ee.get("ee04_registros", ""),
                    [
                        ("tipo_actividad", "Tipo", 28),
                        ("descripcion", "Descripcion / Cultivo", 50),
                        ("area_ha", "Area (ha)", 18),
                        ("familias_dependientes", "N° fam.", 16),
                        ("valor_produccion", "Valor S/", 22),
                        ("nivel_exposicion", "Exp.", 18),
                        ("tipo_peligro", "Peligro", 28),
                        ("perdidas_anteriores", "Perd.ant.", 18),
                        ("monto_perdida", "S/ perd.", 22),
                    ],
                )

            # F-EE-05: Ecosistema y peligros
            elif ficha_ee == "F-EE-05":
                pdf._subficha("F-EE-05: Ecosistema (UP) y Activos Ambientales")
                campos_ee05 = [
                    ("Tipo de ecosistema (MINAM)", "ee05_tipo_ecosistema"),
                    ("Zona de vida (Holdridge)", "ee05_zona_vida"),
                    ("Cobertura vegetal predominante", "ee05_cobertura_vegetal"),
                    ("% cobertura vegetal", "ee05_pct_cobertura"),
                    ("Especies dominantes", "ee05_especies_dominantes"),
                    ("Evidencia de degradacion", "ee05_evidencia_degradacion"),
                    ("Tipo de degradacion", "ee05_tipo_degradacion"),
                    ("Nivel de degradacion (1-5)", "ee05_nivel_degradacion"),
                    ("Pendiente predominante (%)", "ee05_pendiente"),
                    ("Tipo de suelo", "ee05_tipo_suelo"),
                    ("Profundidad efectiva (cm)", "ee05_profundidad_efectiva"),
                    ("Presencia de carcavas/surcos", "ee05_presencia_carcavas"),
                    ("Presencia de quebrada/cauce", "ee05_presencia_quebrada"),
                    ("Nombre de quebrada", "ee05_nombre_quebrada"),
                    ("Fuentes de agua identificadas", "ee05_fuentes_agua"),
                ]
                for label, key in campos_ee05:
                    val = ee.get(key, "") or ""
                    if val:
                        pdf._campo(label, val)
                pdf._tabla_json(
                    "Peligros observados en campo",
                    ee.get("ee05_peligros_observados", ""),
                    [
                        ("tipo_peligro", "Tipo peligro", 35),
                        ("descripcion", "Descripcion / evidencia", 60),
                        ("nivel_estimado", "Nivel", 20),
                        ("probabilidad", "Prob. recur.", 25),
                        ("activos_amenazados", "Activos amenazados", 50),
                    ],
                )

            # F-EE-06: Resumen de vulnerabilidad
            elif ficha_ee == "F-EE-06":
                pdf._subficha("F-EE-06: Resumen de Vulnerabilidad del Bloque")
                pdf._tabla_json(
                    "Cuantificacion de elementos expuestos (gabinete vs campo)",
                    ee.get("ee06_cuantificacion", ""),
                    [
                        ("elemento", "Elemento", 60),
                        ("cantidad_gabinete", "Gabinete", 25),
                        ("cantidad_campo", "Campo", 25),
                        ("coincide", "Coincide", 22),
                        ("observaciones", "Observaciones", 55),
                    ],
                )
                pdf._tabla_json(
                    "Valoracion cualitativa de vulnerabilidad",
                    ee.get("ee06_valoracion_vulnerabilidad", ""),
                    [
                        ("factor", "Factor", 35),
                        ("descriptor", "Descriptor", 55),
                        ("nivel", "Nivel", 20),
                        ("peso", "Peso", 16),
                        ("justificacion", "Justificacion", 60),
                    ],
                )
                if ee.get("ee06_nivel_vulnerabilidad"):
                    pdf._campo("NIVEL DE VULNERABILIDAD DEL BLOQUE",
                               ee["ee06_nivel_vulnerabilidad"])
                if ee.get("ee06_nivel_riesgo"):
                    pdf._campo("NIVEL DE RIESGO PRELIMINAR DEL BLOQUE",
                               ee["ee06_nivel_riesgo"])

            # F-EE-07: Control fotografico
            elif ficha_ee == "F-EE-07":
                pdf._subficha("F-EE-07: Control de Registro Fotografico")
                pdf._tabla_json(
                    "Inventario de fotografias",
                    ee.get("ee07_registros", ""),
                    [
                        ("codigo_foto", "Codigo", 25),
                        ("fecha", "Fecha", 24),
                        ("hora", "Hora", 18),
                        ("elemento_fotografiado", "Elemento", 50),
                        ("formato_referencia", "Ref.", 22),
                        ("descripcion", "Descripcion", 55),
                    ],
                )

            if ee.get("observaciones_generales"):
                pdf._campo_largo("Observaciones generales", ee["observaciones_generales"])
            pdf.ln(2)

    # Cronograma del bloque
    actividades = db.obtener_actividades_por_bloque(bloque_id)
    if actividades:
        pdf._seccion("6. Cronograma de Actividades")
        for a in actividades:
            estado_act = a.get("estado", "Programado")
            avance_act = a.get("porcentaje_avance", 0)
            pdf._campo(a["actividad"],
                       f"{a['fecha_inicio_plan']} a {a['fecha_fin_plan']} | "
                       f"Estado: {estado_act} | Avance: {avance_act:.0f}%")
        pdf.ln(3)

    nombre_archivo = f"ficha_{bloque['codigo']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    ruta_pdf = os.path.join(REPORTES_DIR, nombre_archivo)
    pdf.output(ruta_pdf)
    return ruta_pdf


# ── Reporte Excel - Tabla Resumen ──────────────────────────────────────────

def generar_resumen_excel():
    """Genera un archivo Excel con la tabla resumen de todos los bloques inspeccionados.
    Incluye columnas UTM compatibles con importación a ArcGIS."""
    _asegurar_directorio()

    wb = Workbook()
    enc_font, enc_fill, borde, centrado = _estilos_excel()

    # ── Hoja 1: Resumen de Bloques ──
    ws = wb.active
    ws.title = "Resumen Bloques IN Piura"

    # Título
    ws.merge_cells("A1:T1")
    ws["A1"] = "IN Piura - Plan de Ingreso - Resumen de Bloques de Intervencion"
    ws["A1"].font = Font(name="Calibri", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:T2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Encabezados (fila 4)
    encabezados = [
        "Codigo Bloque", "Microcuenca", "Tipo Intervencion", "Cuenca Hidrografica",
        "Provincia", "Distrito", "UTM_Este", "UTM_Norte", "Zona_UTM", "Altitud (m.s.n.m.)",
        "Area (ha)", "Responsable", "Estado", "Total Inspecciones",
        "Ultima Visita", "Avance Fisico (%)",
        "Cobertura Vegetal (%)", "Tipo Cobertura", "Vigor Cobertura",
        "Fecha Registro"
    ]

    for col_idx, enc in enumerate(encabezados, 1):
        celda = ws.cell(row=4, column=col_idx, value=enc)
        celda.font = enc_font
        celda.fill = enc_fill
        celda.border = borde
        celda.alignment = centrado

    # Datos
    resumen = db.obtener_resumen_bloques()
    for row_idx, bloque in enumerate(resumen, 5):
        indicadores = db.obtener_indicadores_por_bloque(bloque["id"])
        ultimo_ind = indicadores[0] if indicadores else {}

        datos = [
            bloque["codigo"],
            bloque.get("microcuenca", "") or "",
            bloque["tipo_intervencion"],
            bloque["cuenca"],
            bloque.get("provincia", "") or "",
            bloque["distrito"],
            bloque["utm_este"],
            bloque["utm_norte"],
            bloque["utm_zona"],
            bloque.get("altitud", 0) or 0,
            bloque["area_hectareas"],
            bloque.get("responsable", "") or "",
            bloque["estado"],
            bloque.get("total_inspecciones", 0),
            bloque.get("ultima_visita", ""),
            bloque.get("ultimo_avance", 0),
            ultimo_ind.get("porcentaje_cobertura_vegetal", 0),
            ultimo_ind.get("tipo_cobertura_vegetal", "") or "",
            ultimo_ind.get("vigor_cobertura_vegetal", "") or "",
            bloque["fecha_registro"],
        ]

        for col_idx, valor in enumerate(datos, 1):
            celda = ws.cell(row=row_idx, column=col_idx, value=valor)
            celda.border = borde
            celda.alignment = Alignment(horizontal="center", vertical="center")

    anchos = [14, 18, 24, 22, 14, 16, 14, 14, 10, 14, 10, 16, 14, 14, 14, 14, 16, 16, 16, 18]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # ── Hoja 2: Resumen Presupuestal ──
    ws_pres = wb.create_sheet("Resumen Presupuestal")

    ws_pres.merge_cells("A1:G1")
    ws_pres["A1"] = "IN Piura - Resumen Presupuestal por Bloque"
    ws_pres["A1"].font = Font(name="Calibri", bold=True, size=13)
    ws_pres["A1"].alignment = Alignment(horizontal="center")

    enc_pres = ["Codigo Bloque", "Tipo Intervencion", "Distrito",
                "Planificado (S/)", "Ejecutado (S/)", "% Ejecucion", "N Partidas"]
    for col_idx, enc in enumerate(enc_pres, 1):
        celda = ws_pres.cell(row=3, column=col_idx, value=enc)
        celda.font = enc_font
        celda.fill = enc_fill
        celda.border = borde
        celda.alignment = centrado

    resumen_pres = db.obtener_resumen_presupuesto()
    total_plan_global = 0
    total_ejec_global = 0
    for row_idx, r in enumerate(resumen_pres, 4):
        pct = (r["total_ejecutado"] / r["total_planificado"] * 100) if r["total_planificado"] > 0 else 0
        total_plan_global += r["total_planificado"]
        total_ejec_global += r["total_ejecutado"]

        datos_p = [
            r["codigo"], r["tipo_intervencion"], r["distrito"],
            r["total_planificado"], r["total_ejecutado"],
            round(pct, 1), r["num_partidas"],
        ]
        for col_idx, valor in enumerate(datos_p, 1):
            celda = ws_pres.cell(row=row_idx, column=col_idx, value=valor)
            celda.border = borde
            celda.alignment = Alignment(horizontal="center", vertical="center")

    # Fila de totales
    fila_total = len(resumen_pres) + 4
    ws_pres.cell(row=fila_total, column=1, value="TOTAL PROYECTO").font = Font(bold=True)
    ws_pres.cell(row=fila_total, column=4, value=total_plan_global).font = Font(bold=True)
    ws_pres.cell(row=fila_total, column=5, value=total_ejec_global).font = Font(bold=True)
    pct_global = (total_ejec_global / total_plan_global * 100) if total_plan_global > 0 else 0
    ws_pres.cell(row=fila_total, column=6, value=round(pct_global, 1)).font = Font(bold=True)

    anchos_pres = [14, 24, 16, 16, 16, 12, 12]
    for i, ancho in enumerate(anchos_pres, 1):
        ws_pres.column_dimensions[get_column_letter(i)].width = ancho

    # ── Hoja 3: Cronograma ──
    ws_crono = wb.create_sheet("Cronograma")

    ws_crono.merge_cells("A1:H1")
    ws_crono["A1"] = "IN Piura - Cronograma de Actividades"
    ws_crono["A1"].font = Font(name="Calibri", bold=True, size=13)
    ws_crono["A1"].alignment = Alignment(horizontal="center")

    enc_crono = ["Bloque", "Actividad", "Inicio Plan.", "Fin Plan.",
                 "Inicio Real", "Fin Real", "Avance %", "Estado"]
    for col_idx, enc in enumerate(enc_crono, 1):
        celda = ws_crono.cell(row=3, column=col_idx, value=enc)
        celda.font = enc_font
        celda.fill = enc_fill
        celda.border = borde
        celda.alignment = centrado

    actividades = db.obtener_todas_actividades()
    for row_idx, a in enumerate(actividades, 4):
        datos_a = [
            a.get("bloque_codigo", ""),
            a["actividad"],
            a["fecha_inicio_plan"],
            a["fecha_fin_plan"],
            a["fecha_inicio_real"] or "-",
            a["fecha_fin_real"] or "-",
            a["porcentaje_avance"],
            a["estado"],
        ]
        for col_idx, valor in enumerate(datos_a, 1):
            celda = ws_crono.cell(row=row_idx, column=col_idx, value=valor)
            celda.border = borde
            celda.alignment = Alignment(horizontal="center", vertical="center")

    anchos_crono = [14, 28, 14, 14, 14, 14, 10, 14]
    for i, ancho in enumerate(anchos_crono, 1):
        ws_crono.column_dimensions[get_column_letter(i)].width = ancho

    nombre_archivo = f"resumen_bloques_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ruta_excel = os.path.join(REPORTES_DIR, nombre_archivo)
    wb.save(ruta_excel)
    return ruta_excel


def generar_excel_arcgis():
    """Genera un Excel simplificado con columnas UTM listas para importar a ArcGIS."""
    _asegurar_directorio()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bloques_UTM_ArcGIS"

    encabezados_arcgis = [
        "OBJECTID", "COD_BLOQUE", "MICROCUENCA", "TIPO_INTERV", "CUENCA",
        "PROVINCIA", "DISTRITO", "POINT_X", "POINT_Y", "ZONA_UTM", "ALTITUD",
        "AREA_HA", "RESPONSABLE", "ESTADO", "AVANCE_PCT",
        "COB_VEG_PCT", "TIPO_COB_VEG", "VIGOR_COB_VEG"
    ]

    encabezado_font = Font(name="Calibri", bold=True, size=10)
    for col_idx, enc in enumerate(encabezados_arcgis, 1):
        celda = ws.cell(row=1, column=col_idx, value=enc)
        celda.font = encabezado_font

    resumen = db.obtener_resumen_bloques()
    for row_idx, bloque in enumerate(resumen, 2):
        indicadores = db.obtener_indicadores_por_bloque(bloque["id"])
        ultimo_ind = indicadores[0] if indicadores else {}

        datos = [
            row_idx - 1,
            bloque["codigo"],
            bloque.get("microcuenca", "") or "",
            bloque["tipo_intervencion"],
            bloque["cuenca"],
            bloque.get("provincia", "") or "",
            bloque["distrito"],
            bloque["utm_este"],
            bloque["utm_norte"],
            bloque["utm_zona"],
            bloque.get("altitud", 0) or 0,
            bloque["area_hectareas"],
            bloque.get("responsable", "") or "",
            bloque["estado"],
            bloque.get("ultimo_avance", 0) or 0,
            ultimo_ind.get("porcentaje_cobertura_vegetal", 0),
            ultimo_ind.get("tipo_cobertura_vegetal", "") or "",
            ultimo_ind.get("vigor_cobertura_vegetal", "") or "",
        ]

        for col_idx, valor in enumerate(datos, 1):
            ws.cell(row=row_idx, column=col_idx, value=valor)

    anchos = [10, 14, 18, 24, 20, 14, 16, 14, 14, 10, 10, 10, 16, 14, 12, 14, 16, 16]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    nombre_archivo = f"bloques_arcgis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ruta_excel = os.path.join(REPORTES_DIR, nombre_archivo)
    wb.save(ruta_excel)
    return ruta_excel

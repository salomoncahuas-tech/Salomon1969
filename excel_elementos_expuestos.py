"""
Modulo de importacion/exportacion Excel para Elementos Expuestos (AdR).
Proyecto IN Piura CUI 2669244 | ANIN - DIME - SESDI

Genera plantillas Excel estandarizadas para las fichas F-EE-01 a F-EE-07
y parsea archivos Excel llenados por tecnicos de campo para autocompletar
los formularios del aplicativo.
"""

import io
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# -- Estilos (consistentes con excel_diagnostico_social.py) ----------------
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
LABEL_FILL = PatternFill(start_color="D5DBDB", end_color="D5DBDB", fill_type="solid")
LABEL_FONT = Font(name="Calibri", size=10, bold=True)
VALUE_FILL = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
VALUE_FONT = Font(name="Calibri", size=10)
SECTION_FILL = PatternFill(start_color="1ABC9C", end_color="1ABC9C", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TABLE_HEADER_FILL = PatternFill(start_color="5DADE2", end_color="5DADE2", fill_type="solid")
TABLE_HEADER_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
INSTRUCCION_FONT = Font(name="Calibri", size=9, italic=True, color="7F8C8D")
AUTOFILL_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
AUTOFILL_FONT = Font(name="Calibri", size=10, italic=True)

# -- Opciones de listas desplegables --------------------------------------
FEE01_TIPO_ELEMENTO = [
    "POBLACION", "VIVIENDA", "EDUCACION", "SALUD", "TRANSPORTE",
    "RIEGO", "SANEAMIENTO", "TELECOMUNICACIONES", "AGROPECUARIO", "AMBIENTAL"]
FEE01_UBICACION_PELIGRO = [
    "DENTRO_ZONA_ALTA", "DENTRO_ZONA_MEDIA", "DENTRO_ZONA_BAJA", "FUERA"]
FEE_ESTADO = ["B", "R", "M"]
FEE_NIVEL_AMB = ["A", "M", "B"]
FEE_SI_NO = ["Si", "No"]
FEE02_MATERIAL_VIV = ["Adobe", "Material noble", "Quincha", "Madera", "Otro"]
FEE03_SECTOR = [
    "EDUCACION", "SALUD", "TRANSPORTE", "RIEGO", "SANEAMIENTO",
    "TELECOMUNICACIONES", "GOBIERNO", "OTRO"]
FEE_TIPO_PELIGRO = [
    "MM", "EPH", "INUND", "SEQUIA", "HELADA", "INCENDIO", "SISMO", "OTRO"]
FEE04_TIPO_ACTIVIDAD = [
    "AGRICOLA", "PECUARIO", "ACUICOLA", "FORESTAL", "COMERCIO", "TURISMO", "OTRO"]
FEE05_TIPO_DEGRADACION = [
    "Deforestacion", "Sobrepastoreo", "Erosion", "Incendio", "Otro"]
FEE05_FUENTES_AGUA = ["Manantial", "Quebrada", "Rio", "Canal", "Ninguna"]
FEE05_PROB_RECURRENCIA = ["Alta", "Media", "Baja"]
FEE06_NIVEL_VULN = ["Muy Alto", "Alto", "Medio", "Bajo", "Muy Bajo"]

# Elementos fijos para F-EE-06 Seccion A
FEE06_ELEMENTOS = [
    "Poblacion total expuesta (hab.)",
    "N Viviendas expuestas",
    "N Instituciones Educativas expuestas",
    "N Establecimientos de Salud expuestos",
    "Longitud de vias expuestas (km)",
    "N Infraestructura de riego expuesta",
    "Area agricola expuesta (ha)",
    "N Cabezas de ganado expuesto",
    "Area de ecosistema degradado (ha)",
    "N Fuentes de agua expuestas",
]

# Factores fijos para F-EE-06 Seccion B
FEE06_FACTORES = [
    ("EXPOSICION", "Localizacion de activos respecto a zona de peligro"),
    ("FRAGILIDAD - Dimension Fisica", "Material constructivo, estado, antiguedad"),
    ("FRAGILIDAD - Dimension Social", "Nivel educativo, organizacion comunal"),
    ("FRAGILIDAD - Dimension Economica", "Dependencia economica, diversificacion"),
    ("FRAGILIDAD - Dimension Ambiental", "Degradacion del ecosistema, perdida de cobertura"),
    ("RESILIENCIA - Dimension Fisica", "Existencia de obras de proteccion"),
    ("RESILIENCIA - Dimension Social", "Capacidad de respuesta, conocimiento de GdR"),
    ("RESILIENCIA - Dimension Economica", "Acceso a seguros, ahorros, empleo alternativo"),
    ("RESILIENCIA - Dimension Ambiental", "Capacidad regenerativa del ecosistema"),
]


# =========================================================================
# FUNCIONES AUXILIARES
# =========================================================================

def _cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def _add_header(ws, ficha_titulo, max_col=4):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    _cell(ws, 1, 1, "PROYECTO IN PIURA - CUI 2669244",
          HEADER_FONT, HEADER_FILL, Alignment(horizontal="center", vertical="center"))
    for c in range(1, max_col + 1):
        ws.cell(row=1, column=c).fill = HEADER_FILL
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    _cell(ws, 2, 1, ficha_titulo,
          SUBHEADER_FONT, SUBHEADER_FILL, Alignment(horizontal="center", vertical="center"))
    for c in range(1, max_col + 1):
        ws.cell(row=2, column=c).fill = SUBHEADER_FILL
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
    _cell(ws, 3, 1, "ANIN - DIME - SESDI | Complete las celdas naranja",
          INSTRUCCION_FONT, None, Alignment(horizontal="center"))


def _add_datos_sheet(wb, bloques_data):
    ws = wb.create_sheet("_Datos")
    ws.cell(row=1, column=1, value="Codigo")
    ws.cell(row=1, column=2, value="Microcuenca")
    ws.cell(row=1, column=3, value="Provincia")
    ws.cell(row=1, column=4, value="Distrito")
    for i, (codigo, mc, prov, dist) in enumerate(bloques_data, start=2):
        ws.cell(row=i, column=1, value=codigo)
        ws.cell(row=i, column=2, value=mc)
        ws.cell(row=i, column=3, value=prov)
        ws.cell(row=i, column=4, value=dist)
    ws.sheet_state = "hidden"
    return len(bloques_data)


def _add_datos_generales(ws, start_row=5, num_bloques=0, max_col=4):
    r = start_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    _cell(ws, r, 1, "DATOS DE IDENTIFICACION", SECTION_FONT, SECTION_FILL,
          Alignment(horizontal="center"))
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    # Codigo de Bloque
    bloque_row = r
    _cell(ws, r, 1, "Codigo del Bloque", LABEL_FONT, LABEL_FILL, border=THIN_BORDER)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=max_col)
    _cell(ws, r, 2, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
    if num_bloques > 0:
        last = num_bloques + 1
        dv = DataValidation(type="list", formula1=f"_Datos!$A$2:$A${last}", allow_blank=True)
        dv.error = "Seleccione un bloque registrado"
        dv.errorTitle = "Bloque no valido"
        ws.add_data_validation(dv)
        dv.add(f"B{r}")
    r += 1

    # Microcuenca (auto)
    _cell(ws, r, 1, "Microcuenca (auto)", LABEL_FONT, LABEL_FILL, border=THIN_BORDER)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=max_col)
    if num_bloques > 0:
        last = num_bloques + 1
        ws.cell(row=r, column=2).value = (
            f'=IFERROR(VLOOKUP(B{bloque_row},_Datos!$A$2:$D${last},2,FALSE),"")')
        ws.cell(row=r, column=2).font = AUTOFILL_FONT
        ws.cell(row=r, column=2).fill = AUTOFILL_FILL
        ws.cell(row=r, column=2).border = THIN_BORDER
    else:
        _cell(ws, r, 2, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
    r += 1

    # Provincia (auto)
    _cell(ws, r, 1, "Provincia (auto)", LABEL_FONT, LABEL_FILL, border=THIN_BORDER)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=max_col)
    if num_bloques > 0:
        last = num_bloques + 1
        ws.cell(row=r, column=2).value = (
            f'=IFERROR(VLOOKUP(B{bloque_row},_Datos!$A$2:$D${last},3,FALSE),"")')
        ws.cell(row=r, column=2).font = AUTOFILL_FONT
        ws.cell(row=r, column=2).fill = AUTOFILL_FILL
        ws.cell(row=r, column=2).border = THIN_BORDER
    else:
        _cell(ws, r, 2, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
    r += 1

    # Distrito (auto)
    _cell(ws, r, 1, "Distrito (auto)", LABEL_FONT, LABEL_FILL, border=THIN_BORDER)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=max_col)
    if num_bloques > 0:
        last = num_bloques + 1
        ws.cell(row=r, column=2).value = (
            f'=IFERROR(VLOOKUP(B{bloque_row},_Datos!$A$2:$D${last},4,FALSE),"")')
        ws.cell(row=r, column=2).font = AUTOFILL_FONT
        ws.cell(row=r, column=2).fill = AUTOFILL_FILL
        ws.cell(row=r, column=2).border = THIN_BORDER
    else:
        _cell(ws, r, 2, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
    r += 1

    for label in ("Centro(s) Poblado(s)", "Coordenada UTM Este (m)",
                  "Coordenada UTM Norte (m)", "Altitud (msnm)",
                  "Fecha de campo (DD/MM/AAAA)", "Responsable / Brigada"):
        _cell(ws, r, 1, label, LABEL_FONT, LABEL_FILL, border=THIN_BORDER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=max_col)
        _cell(ws, r, 2, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        r += 1

    return r + 1


def _add_validation(ws, col_letter, min_row, max_row, options):
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Seleccione un valor de la lista"
    dv.errorTitle = "Valor no valido"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{min_row}:{col_letter}{max_row}")



# =========================================================================
# GENERACION DE PLANTILLAS
# =========================================================================

def _generar_ee01(wb):
    """F-EE-01: Inventario de Elementos Expuestos."""
    ws = wb.create_sheet("F-EE-01")
    max_col = 14
    for i, w in enumerate([5, 18, 15, 18, 12, 12, 10, 12, 18, 10, 14, 10, 13, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _add_header(ws, "F-EE-01: INVENTARIO DE ELEMENTOS EXPUESTOS", max_col)
    r = _add_datos_generales(ws, num_bloques=wb._num_bloques, max_col=max_col)

    # Tabla
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    _cell(ws, r, 1, "REGISTRO DE ELEMENTOS EXPUESTOS", SECTION_FONT, SECTION_FILL)
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    headers = ["N", "Tipo de Elemento\nExpuesto", "Subtipo /\nCategoria",
               "Nombre o\nIdentificador", "Coord. UTM\nEste (m)", "Coord. UTM\nNorte (m)",
               "Altitud\n(msnm)", "Distancia al\nBloque (m)",
               "Ubicacion respecto\nal Peligro", "Estado Actual\n(B/R/M)",
               "Material\nPredominante", "Antiguedad\n(anios)",
               "N Beneficiarios\no Usuarios", "Registro\nFotografico\n(Codigo Foto)"]
    for i, h in enumerate(headers):
        _cell(ws, r, i + 1, h, TABLE_HEADER_FONT, TABLE_HEADER_FILL,
              Alignment(horizontal="center", wrap_text=True, vertical="center"), THIN_BORDER)
    r += 1
    data_start = r
    for row_n in range(20):
        _cell(ws, r, 1, row_n + 1, VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        for c in range(2, max_col + 1):
            _cell(ws, r, c, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        r += 1
    data_end = data_start + 19

    _add_validation(ws, "B", data_start, data_end, FEE01_TIPO_ELEMENTO)
    _add_validation(ws, "I", data_start, data_end, FEE01_UBICACION_PELIGRO)
    _add_validation(ws, "J", data_start, data_end, FEE_ESTADO)

    r += 1
    _cell(ws, r, 1, "LEYENDA:", LABEL_FONT, LABEL_FILL)
    r += 1
    _cell(ws, r, 1, "Tipo de Elemento: " + " / ".join(FEE01_TIPO_ELEMENTO), INSTRUCCION_FONT)
    r += 1
    _cell(ws, r, 1, "Ubicacion: " + " / ".join(FEE01_UBICACION_PELIGRO), INSTRUCCION_FONT)
    r += 1
    _cell(ws, r, 1, "Estado Actual: B=Bueno / R=Regular / M=Malo", INSTRUCCION_FONT)


def _generar_ee02(wb):
    """F-EE-02: Poblacion y Viviendas."""
    ws = wb.create_sheet("F-EE-02")
    max_col = 16
    for i, w in enumerate([5, 20, 12, 12, 14, 12, 14, 16, 10, 12, 12, 12, 14, 10, 12, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _add_header(ws, "F-EE-02: ELEMENTOS EXPUESTOS - POBLACION Y VIVIENDAS", max_col)
    r = _add_datos_generales(ws, num_bloques=wb._num_bloques, max_col=max_col)

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    _cell(ws, r, 1, "DETALLE DE CENTROS POBLADOS, POBLACION Y VIVIENDAS", SECTION_FONT, SECTION_FILL)
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    headers = ["N", "Centro Poblado\n(Nombre)", "Codigo\nINEI", "N Viviendas\nTotal",
               "N Viviendas en\nArea de Peligro", "Poblacion\nTotal",
               "Poblacion en\nArea de Peligro", "Material de\nViviendas\n(Predominante)",
               "N Pisos\n(Predominante)", "Acceso a Agua\nPotable (Si/No)",
               "Acceso a\nElectricidad\n(Si/No)", "Antecedente de\nEvento (Si/No)",
               "Tipo Evento\nAnterior", "Anio Ultimo\nEvento",
               "Nivel de Danio\nAnterior (A/M/B)", "Observaciones"]
    for i, h in enumerate(headers):
        _cell(ws, r, i + 1, h, TABLE_HEADER_FONT, TABLE_HEADER_FILL,
              Alignment(horizontal="center", wrap_text=True, vertical="center"), THIN_BORDER)
    r += 1
    data_start = r
    for row_n in range(15):
        _cell(ws, r, 1, row_n + 1, VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        for c in range(2, max_col + 1):
            _cell(ws, r, c, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        r += 1
    data_end = data_start + 14

    _add_validation(ws, "H", data_start, data_end, FEE02_MATERIAL_VIV)
    _add_validation(ws, "J", data_start, data_end, FEE_SI_NO)
    _add_validation(ws, "K", data_start, data_end, FEE_SI_NO)
    _add_validation(ws, "L", data_start, data_end, FEE_SI_NO)
    _add_validation(ws, "O", data_start, data_end, FEE_NIVEL_AMB)

    # Fila TOTAL
    r += 1
    _cell(ws, r, 1, "TOTAL", LABEL_FONT, LABEL_FILL, border=THIN_BORDER)
    for c in range(2, max_col + 1):
        _cell(ws, r, c, "", LABEL_FONT, LABEL_FILL, border=THIN_BORDER)


def _generar_ee03(wb):
    """F-EE-03: Infraestructura Publica."""
    ws = wb.create_sheet("F-EE-03")
    max_col = 18
    for i, w in enumerate([5, 14, 16, 18, 12, 12, 12, 14, 8, 10, 13, 12, 14, 10, 16, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _add_header(ws, "F-EE-03: ELEMENTOS EXPUESTOS - INFRAESTRUCTURA PUBLICA", max_col)
    r = _add_datos_generales(ws, num_bloques=wb._num_bloques, max_col=max_col)

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    _cell(ws, r, 1, "INFRAESTRUCTURA PUBLICA EXPUESTA", SECTION_FONT, SECTION_FILL)
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    headers = ["N", "Sector", "Tipo de\nInfraestructura", "Nombre /\nIdentificador",
               "Codigo\nOficial", "Coord. UTM\nEste (m)", "Coord. UTM\nNorte (m)",
               "Material\nConstructivo", "Estado\n(B/R/M)", "Antiguedad\n(anios)",
               "N Usuarios /\nBeneficiarios", "Nivel de\nExposicion\n(A/M/B)",
               "Tipo de Peligro\nque lo afecta", "Antecedente\nde Danio\n(Si/No)",
               "Descripcion\ndel Danio", "Costo Estimado\ndel Activo (S/)",
               "Costo Estimado\nde Reposicion (S/)", "Registro\nFotografico"]
    for i, h in enumerate(headers):
        _cell(ws, r, i + 1, h, TABLE_HEADER_FONT, TABLE_HEADER_FILL,
              Alignment(horizontal="center", wrap_text=True, vertical="center"), THIN_BORDER)
    r += 1
    data_start = r
    for row_n in range(15):
        _cell(ws, r, 1, row_n + 1, VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        for c in range(2, max_col + 1):
            _cell(ws, r, c, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        r += 1
    data_end = data_start + 14

    _add_validation(ws, "B", data_start, data_end, FEE03_SECTOR)
    _add_validation(ws, "I", data_start, data_end, FEE_ESTADO)
    _add_validation(ws, "L", data_start, data_end, FEE_NIVEL_AMB)
    _add_validation(ws, "M", data_start, data_end, FEE_TIPO_PELIGRO)
    _add_validation(ws, "N", data_start, data_end, FEE_SI_NO)

    r += 1
    _cell(ws, r, 1, "LEYENDA:", LABEL_FONT, LABEL_FILL)
    r += 1
    _cell(ws, r, 1, "Sector: " + " / ".join(FEE03_SECTOR), INSTRUCCION_FONT)
    r += 1
    _cell(ws, r, 1, "Tipo Peligro: " + " / ".join(FEE_TIPO_PELIGRO), INSTRUCCION_FONT)


def _generar_ee04(wb):
    """F-EE-04: Actividades Economicas y Agropecuarias."""
    ws = wb.create_sheet("F-EE-04")
    max_col = 16
    for i, w in enumerate([5, 16, 20, 10, 12, 12, 16, 18, 18, 12, 12, 12, 14, 12, 14, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _add_header(ws, "F-EE-04: ELEMENTOS EXPUESTOS - ACTIVIDADES ECONOMICAS Y AGROPECUARIAS", max_col)
    r = _add_datos_generales(ws, num_bloques=wb._num_bloques, max_col=max_col)

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    _cell(ws, r, 1, "ACTIVIDADES ECONOMICAS Y AGROPECUARIAS EXPUESTAS", SECTION_FONT, SECTION_FILL)
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    headers = ["N", "Tipo de Actividad", "Descripcion /\nCultivo Principal", "Area (ha)",
               "N Predios\nAfectables", "N Familias\nDependientes",
               "Valor Estimado\nProduccion\nAnual (S/)", "Infraestructura\nProductiva\nAsociada",
               "Ganado\n(Tipo / N Cabezas)", "Coord. UTM\nEste", "Coord. UTM\nNorte",
               "Nivel de\nExposicion\n(A/M/B)", "Tipo Peligro\nque lo Afecta",
               "Perdidas por\nEventos Anteriores\n(Si/No)", "Monto Estimado\nPerdida (S/)",
               "Observaciones"]
    for i, h in enumerate(headers):
        _cell(ws, r, i + 1, h, TABLE_HEADER_FONT, TABLE_HEADER_FILL,
              Alignment(horizontal="center", wrap_text=True, vertical="center"), THIN_BORDER)
    r += 1
    data_start = r
    for row_n in range(14):
        _cell(ws, r, 1, row_n + 1, VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        for c in range(2, max_col + 1):
            _cell(ws, r, c, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        r += 1
    data_end = data_start + 13

    _add_validation(ws, "B", data_start, data_end, FEE04_TIPO_ACTIVIDAD)
    _add_validation(ws, "L", data_start, data_end, FEE_NIVEL_AMB)
    _add_validation(ws, "M", data_start, data_end, FEE_TIPO_PELIGRO)
    _add_validation(ws, "N", data_start, data_end, FEE_SI_NO)



def _generar_ee05(wb):
    """F-EE-05: Ecosistema (UP) y Activos Ambientales."""
    ws = wb.create_sheet("F-EE-05")
    max_col = 11
    for i, w in enumerate([5, 22, 22, 12, 12, 14, 16, 18, 12, 14, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _add_header(ws, "F-EE-05: ECOSISTEMA (UP) Y ACTIVOS AMBIENTALES", max_col)
    r = _add_datos_generales(ws, num_bloques=wb._num_bloques, max_col=max_col)

    # Seccion A: Caracterizacion
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    _cell(ws, r, 1, "A. CARACTERIZACION DEL ECOSISTEMA EN EL BLOQUE", SECTION_FONT, SECTION_FILL)
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    campos_eco = [
        ("Tipo de Ecosistema (MINAM)", None),
        ("Zona de Vida (Holdridge)", None),
        ("Cobertura Vegetal Predominante", None),
        ("% Cobertura Vegetal Estimado", None),
        ("Especies Dominantes (listar)", None),
        ("Evidencia de Degradacion", "SI / NO"),
        ("Tipo de Degradacion", ", ".join(FEE05_TIPO_DEGRADACION)),
        ("Nivel de Degradacion (1-5)", "1=Leve ... 5=Severa"),
        ("Pendiente Predominante (%)", None),
        ("Tipo de Suelo Observado", None),
        ("Profundidad Efectiva Estimada (cm)", None),
        ("Presencia de Carcavas/Surcos", "SI / NO"),
        ("Presencia de Quebrada/Cauce", "SI / NO"),
        ("Nombre de Quebrada", None),
        ("Fuentes de Agua Identificadas", ", ".join(FEE05_FUENTES_AGUA)),
    ]
    for label, hint in campos_eco:
        _cell(ws, r, 1, label, LABEL_FONT, LABEL_FILL, border=THIN_BORDER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=max_col)
        _cell(ws, r, 2, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        r += 1

    # Validaciones para campos con lista
    si_no_rows = []
    for i, (label, _) in enumerate(campos_eco):
        row_num = r - len(campos_eco) + i
        if label == "Evidencia de Degradacion":
            _add_validation(ws, "B", row_num, row_num, FEE_SI_NO)
        elif label == "Tipo de Degradacion":
            _add_validation(ws, "B", row_num, row_num, FEE05_TIPO_DEGRADACION)
        elif label == "Nivel de Degradacion (1-5)":
            _add_validation(ws, "B", row_num, row_num, ["1", "2", "3", "4", "5"])
        elif label == "Presencia de Carcavas/Surcos":
            _add_validation(ws, "B", row_num, row_num, FEE_SI_NO)
        elif label == "Presencia de Quebrada/Cauce":
            _add_validation(ws, "B", row_num, row_num, FEE_SI_NO)
        elif label == "Fuentes de Agua Identificadas":
            _add_validation(ws, "B", row_num, row_num, FEE05_FUENTES_AGUA)

    r += 1

    # Seccion B: Peligros observados
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    _cell(ws, r, 1, "B. EVIDENCIAS DE PELIGROS OBSERVADOS EN CAMPO", SECTION_FONT, SECTION_FILL)
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    headers_b = ["N", "Tipo de Peligro\nObservado", "Descripcion del\nIndicio/Evidencia",
                 "Coord. UTM\nEste", "Coord. UTM\nNorte", "Dimension\nEstimada (m)",
                 "Direccion de\nMovimiento/Flujo", "Activos\nAmenazados",
                 "Nivel Estimado\n(A/M/B)", "Probabilidad\nde Recurrencia", "Codigo Foto"]
    for i, h in enumerate(headers_b):
        _cell(ws, r, i + 1, h, TABLE_HEADER_FONT, TABLE_HEADER_FILL,
              Alignment(horizontal="center", wrap_text=True, vertical="center"), THIN_BORDER)
    r += 1
    data_start = r
    for row_n in range(8):
        _cell(ws, r, 1, row_n + 1, VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        for c in range(2, max_col + 1):
            _cell(ws, r, c, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        r += 1
    data_end = data_start + 7

    _add_validation(ws, "B", data_start, data_end, FEE_TIPO_PELIGRO)
    _add_validation(ws, "I", data_start, data_end, FEE_NIVEL_AMB)
    _add_validation(ws, "J", data_start, data_end, FEE05_PROB_RECURRENCIA)


def _generar_ee06(wb):
    """F-EE-06: Resumen de Vulnerabilidad del Bloque."""
    ws = wb.create_sheet("F-EE-06")
    max_col = 7
    for i, w in enumerate([5, 28, 32, 16, 12, 12, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _add_header(ws, "F-EE-06: RESUMEN DE VULNERABILIDAD DEL BLOQUE", max_col)
    r = _add_datos_generales(ws, num_bloques=wb._num_bloques, max_col=max_col)

    # Seccion A: Cuantificacion
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    _cell(ws, r, 1, "A. CUANTIFICACION DE ELEMENTOS EXPUESTOS", SECTION_FONT, SECTION_FILL)
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    headers_a = ["N", "Elemento Expuesto", "Cantidad\nGabinete\n(Secundaria)",
                 "Cantidad\nVerificada\n(Campo)", "Coincide\n(Si/No)", "Observaciones"]
    # Solo 6 columnas para seccion A
    for i, h in enumerate(headers_a):
        _cell(ws, r, i + 1, h, TABLE_HEADER_FONT, TABLE_HEADER_FILL,
              Alignment(horizontal="center", wrap_text=True, vertical="center"), THIN_BORDER)
    r += 1
    data_start_a = r
    for idx, elem in enumerate(FEE06_ELEMENTOS):
        _cell(ws, r, 1, idx + 1, VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        _cell(ws, r, 2, elem, LABEL_FONT, LABEL_FILL, border=THIN_BORDER)
        for c in range(3, 7):
            _cell(ws, r, c, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        r += 1
    data_end_a = data_start_a + 9
    _add_validation(ws, "E", data_start_a, data_end_a, FEE_SI_NO)

    r += 1

    # Seccion B: Valoracion cualitativa
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    _cell(ws, r, 1, "B. VALORACION CUALITATIVA DE VULNERABILIDAD", SECTION_FONT, SECTION_FILL)
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    headers_b = ["N", "Factor / Dimension", "Descriptor o Variable",
                 "Nivel\n(MB/B/M/A/MA)", "Peso\nPonderado", "Valor\nPonderado",
                 "Justificacion / Evidencia de Campo"]
    for i, h in enumerate(headers_b):
        _cell(ws, r, i + 1, h, TABLE_HEADER_FONT, TABLE_HEADER_FILL,
              Alignment(horizontal="center", wrap_text=True, vertical="center"), THIN_BORDER)
    r += 1
    data_start_b = r
    for idx, (factor, descriptor) in enumerate(FEE06_FACTORES):
        _cell(ws, r, 1, idx + 1, VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        _cell(ws, r, 2, factor, LABEL_FONT, LABEL_FILL, border=THIN_BORDER)
        _cell(ws, r, 3, descriptor, VALUE_FONT, LABEL_FILL,
              Alignment(wrap_text=True), THIN_BORDER)
        for c in range(4, 8):
            _cell(ws, r, c, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        r += 1
    data_end_b = data_start_b + 8
    _add_validation(ws, "D", data_start_b, data_end_b, FEE06_NIVEL_VULN)

    r += 1

    # Resumen final
    _cell(ws, r, 1, "NIVEL DE VULNERABILIDAD DEL BLOQUE:", LABEL_FONT, LABEL_FILL, border=THIN_BORDER)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=max_col)
    _cell(ws, r, 2, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
    _add_validation(ws, "B", r, r, FEE06_NIVEL_VULN)
    r += 1
    _cell(ws, r, 1, "NIVEL DE RIESGO PRELIMINAR DEL BLOQUE:", LABEL_FONT, LABEL_FILL, border=THIN_BORDER)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=max_col)
    _cell(ws, r, 2, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
    _add_validation(ws, "B", r, r, FEE06_NIVEL_VULN)


def _generar_ee07(wb):
    """F-EE-07: Control de Registro Fotografico."""
    ws = wb.create_sheet("F-EE-07")
    max_col = 10
    for i, w in enumerate([5, 14, 12, 10, 12, 12, 10, 20, 16, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _add_header(ws, "F-EE-07: CONTROL DE REGISTRO FOTOGRAFICO", max_col)
    r = _add_datos_generales(ws, num_bloques=wb._num_bloques, max_col=max_col)

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    _cell(ws, r, 1, "INVENTARIO DE FOTOGRAFIAS GEOREFERENCIADAS", SECTION_FONT, SECTION_FILL)
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    r += 1

    headers = ["N", "Codigo Foto", "Fecha", "Hora", "Coord. UTM\nEste",
               "Coord. UTM\nNorte", "Altitud\n(msnm)", "Elemento\nFotografiado",
               "Formato (F-EE-XX)\nde Referencia", "Descripcion / Detalle"]
    for i, h in enumerate(headers):
        _cell(ws, r, i + 1, h, TABLE_HEADER_FONT, TABLE_HEADER_FILL,
              Alignment(horizontal="center", wrap_text=True, vertical="center"), THIN_BORDER)
    r += 1
    for row_n in range(25):
        _cell(ws, r, 1, row_n + 1, VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        for c in range(2, max_col + 1):
            _cell(ws, r, c, "", VALUE_FONT, VALUE_FILL, border=THIN_BORDER)
        r += 1


# =========================================================================
# FUNCION PUBLICA: GENERAR PLANTILLA COMPLETA
# =========================================================================

def generar_plantilla_ee(bloques_data=None):
    """Genera un workbook con todas las hojas F-EE-01 a F-EE-07."""
    wb = Workbook()
    wb.remove(wb.active)

    if bloques_data:
        wb._num_bloques = _add_datos_sheet(wb, bloques_data)
    else:
        wb._num_bloques = 0

    _generar_ee01(wb)
    _generar_ee02(wb)
    _generar_ee03(wb)
    _generar_ee04(wb)
    _generar_ee05(wb)
    _generar_ee06(wb)
    _generar_ee07(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()



# =========================================================================
# PARSERS DE IMPORTACION
# =========================================================================

def _val(ws, row, col):
    """Lee el valor de una celda como texto limpio."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    s = str(v).strip()
    if s.startswith("="):
        return ""
    return s


def _find_label_row(ws, keyword, col=1, start=1):
    """Busca una fila cuyo col contenga keyword."""
    kw = keyword.lower()
    for r in range(start, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v and kw in str(v).lower():
            return r
    return None


def _read_label_value_pairs(ws, start, end, label_col=1, val_col=2):
    """Lee pares etiqueta:valor de un rango."""
    pairs = {}
    for r in range(start, end + 1):
        label = _val(ws, r, label_col)
        if label:
            pairs[label] = _val(ws, r, val_col)
    return pairs


def _parse_datos_generales(ws):
    """Parsea la cabecera comun de identificacion."""
    pairs = _read_label_value_pairs(ws, 1, min(20, ws.max_row))
    datos = {}

    def _g(*kws):
        for k, v in pairs.items():
            kl = k.lower()
            for kw in kws:
                if kw.lower() in kl:
                    return v
        return ""

    datos["codigo_bloque"] = _g("codigo del bloque", "codigo bloque")
    datos["centro_poblado"] = _g("centro(s) poblado", "centro poblado")
    datos["coordenada_este"] = _g("coordenada utm este", "utm este")
    datos["coordenada_norte"] = _g("coordenada utm norte", "utm norte")
    datos["altitud"] = _g("altitud")
    datos["fecha_campo"] = _g("fecha de campo", "fecha campo")
    datos["responsable_brigada"] = _g("responsable", "brigada")
    return datos


def _parse_table(ws, header_row, num_rows, col_map):
    """Parsea una tabla generica dado el mapeo de columnas."""
    registros = []
    for r in range(header_row + 1, header_row + 1 + num_rows):
        row_data = {}
        has_data = False
        for col_idx, field_name in col_map.items():
            v = _val(ws, r, col_idx)
            row_data[field_name] = v
            if v and field_name != "numero":
                has_data = True
        if has_data:
            registros.append(row_data)
    return registros


def _parse_ee01(ws):
    """Parsea F-EE-01: Inventario."""
    datos = _parse_datos_generales(ws)
    header = _find_label_row(ws, "Tipo de Elemento") or _find_label_row(ws, "N")
    if not header:
        return datos
    col_map = {
        1: "numero", 2: "tipo_elemento", 3: "subtipo", 4: "nombre",
        5: "utm_este", 6: "utm_norte", 7: "altitud", 8: "distancia_bloque",
        9: "ubicacion_peligro", 10: "estado", 11: "material",
        12: "antiguedad", 13: "beneficiarios", 14: "foto"
    }
    datos["ee01_registros"] = _parse_table(ws, header, 20, col_map)
    return datos


def _parse_ee02(ws):
    """Parsea F-EE-02: Poblacion y Viviendas."""
    datos = _parse_datos_generales(ws)
    header = _find_label_row(ws, "Centro Poblado") or _find_label_row(ws, "N")
    if not header:
        return datos
    col_map = {
        1: "numero", 2: "centro_poblado", 3: "codigo_inei",
        4: "viviendas_total", 5: "viviendas_peligro", 6: "poblacion_total",
        7: "poblacion_peligro", 8: "material_viviendas", 9: "pisos",
        10: "agua_potable", 11: "electricidad", 12: "antecedente_evento",
        13: "tipo_evento", 14: "anio_evento", 15: "nivel_danio", 16: "observaciones"
    }
    datos["ee02_registros"] = _parse_table(ws, header, 15, col_map)
    # Totales
    total_row = _find_label_row(ws, "TOTAL")
    if total_row:
        datos["ee02_total_viviendas"] = _val(ws, total_row, 4)
        datos["ee02_total_poblacion"] = _val(ws, total_row, 6)
    return datos


def _parse_ee03(ws):
    """Parsea F-EE-03: Infraestructura Publica."""
    datos = _parse_datos_generales(ws)
    header = _find_label_row(ws, "Sector") or _find_label_row(ws, "Tipo de\nInfraestructura")
    if not header:
        return datos
    col_map = {
        1: "numero", 2: "sector", 3: "tipo_infraestructura", 4: "nombre",
        5: "codigo_oficial", 6: "utm_este", 7: "utm_norte", 8: "material",
        9: "estado", 10: "antiguedad", 11: "beneficiarios", 12: "nivel_exposicion",
        13: "tipo_peligro", 14: "antecedente_danio", 15: "descripcion_danio",
        16: "costo_activo", 17: "costo_reposicion", 18: "foto"
    }
    datos["ee03_registros"] = _parse_table(ws, header, 15, col_map)
    return datos


def _parse_ee04(ws):
    """Parsea F-EE-04: Actividades Economicas."""
    datos = _parse_datos_generales(ws)
    header = _find_label_row(ws, "Tipo de Actividad") or _find_label_row(ws, "N")
    if not header:
        return datos
    col_map = {
        1: "numero", 2: "tipo_actividad", 3: "descripcion", 4: "area_ha",
        5: "predios_afectables", 6: "familias_dependientes",
        7: "valor_produccion", 8: "infraestructura_productiva",
        9: "ganado", 10: "utm_este", 11: "utm_norte",
        12: "nivel_exposicion", 13: "tipo_peligro",
        14: "perdidas_anteriores", 15: "monto_perdida", 16: "observaciones"
    }
    datos["ee04_registros"] = _parse_table(ws, header, 14, col_map)
    return datos


def _parse_ee05(ws):
    """Parsea F-EE-05: Ecosistema."""
    datos = _parse_datos_generales(ws)
    pairs = _read_label_value_pairs(ws, 1, ws.max_row)

    def _g(*kws):
        for k, v in pairs.items():
            kl = k.lower()
            for kw in kws:
                if kw.lower() in kl:
                    return v
        return ""

    datos["ee05_tipo_ecosistema"] = _g("tipo de ecosistema")
    datos["ee05_zona_vida"] = _g("zona de vida")
    datos["ee05_cobertura_vegetal"] = _g("cobertura vegetal predominante")
    datos["ee05_pct_cobertura"] = _g("% cobertura")
    datos["ee05_especies_dominantes"] = _g("especies dominantes")
    datos["ee05_evidencia_degradacion"] = _g("evidencia de degradacion")
    datos["ee05_tipo_degradacion"] = _g("tipo de degradacion")
    datos["ee05_nivel_degradacion"] = _g("nivel de degradacion")
    datos["ee05_pendiente"] = _g("pendiente predominante")
    datos["ee05_tipo_suelo"] = _g("tipo de suelo")
    datos["ee05_profundidad_efectiva"] = _g("profundidad efectiva")
    datos["ee05_presencia_carcavas"] = _g("presencia de carcavas")
    datos["ee05_presencia_quebrada"] = _g("presencia de quebrada")
    datos["ee05_nombre_quebrada"] = _g("nombre de quebrada")
    datos["ee05_fuentes_agua"] = _g("fuentes de agua")

    # Tabla B: Peligros observados
    header = _find_label_row(ws, "Tipo de Peligro\nObservado") or \
             _find_label_row(ws, "Tipo de Peligro")
    if header:
        col_map = {
            1: "numero", 2: "tipo_peligro", 3: "descripcion",
            4: "utm_este", 5: "utm_norte", 6: "dimension",
            7: "direccion", 8: "activos_amenazados",
            9: "nivel_estimado", 10: "probabilidad", 11: "foto"
        }
        datos["ee05_peligros_observados"] = _parse_table(ws, header, 8, col_map)
    return datos


def _parse_ee06(ws):
    """Parsea F-EE-06: Resumen de Vulnerabilidad."""
    datos = _parse_datos_generales(ws)

    # Seccion A: Cuantificacion
    header_a = _find_label_row(ws, "Elemento Expuesto")
    cuantificacion = []
    if header_a:
        for r in range(header_a + 1, header_a + 11):
            elem = _val(ws, r, 2)
            if elem:
                cuantificacion.append({
                    "elemento": elem,
                    "cantidad_gabinete": _val(ws, r, 3),
                    "cantidad_campo": _val(ws, r, 4),
                    "coincide": _val(ws, r, 5),
                    "observaciones": _val(ws, r, 6),
                })
    datos["ee06_cuantificacion"] = cuantificacion

    # Seccion B: Valoracion cualitativa
    header_b = _find_label_row(ws, "Factor / Dimension") or \
               _find_label_row(ws, "Factor")
    valoracion = []
    if header_b:
        for r in range(header_b + 1, header_b + 10):
            factor = _val(ws, r, 2)
            if factor:
                valoracion.append({
                    "factor": factor,
                    "descriptor": _val(ws, r, 3),
                    "nivel": _val(ws, r, 4),
                    "peso": _val(ws, r, 5),
                    "valor_ponderado": _val(ws, r, 6),
                    "justificacion": _val(ws, r, 7),
                })
    datos["ee06_valoracion_vulnerabilidad"] = valoracion

    # Resumen final
    vuln_row = _find_label_row(ws, "NIVEL DE VULNERABILIDAD")
    if vuln_row:
        datos["ee06_nivel_vulnerabilidad"] = _val(ws, vuln_row, 2)
    riesgo_row = _find_label_row(ws, "NIVEL DE RIESGO")
    if riesgo_row:
        datos["ee06_nivel_riesgo"] = _val(ws, riesgo_row, 2)
    return datos


def _parse_ee07(ws):
    """Parsea F-EE-07: Control Fotografico."""
    datos = _parse_datos_generales(ws)
    header = _find_label_row(ws, "Codigo Foto") or _find_label_row(ws, "N")
    if not header:
        return datos
    col_map = {
        1: "numero", 2: "codigo_foto", 3: "fecha", 4: "hora",
        5: "utm_este", 6: "utm_norte", 7: "altitud",
        8: "elemento_fotografiado", 9: "formato_referencia", 10: "descripcion"
    }
    datos["ee07_registros"] = _parse_table(ws, header, 25, col_map)
    return datos


# =========================================================================
# FUNCION PUBLICA: PARSEAR EXCEL CARGADO
# =========================================================================

_PARSERS = {
    "F-EE-01": _parse_ee01,
    "F-EE-02": _parse_ee02,
    "F-EE-03": _parse_ee03,
    "F-EE-04": _parse_ee04,
    "F-EE-05": _parse_ee05,
    "F-EE-06": _parse_ee06,
    "F-EE-07": _parse_ee07,
}


def parsear_excel_ee(file_bytes):
    """Parsea un archivo Excel cargado y devuelve dict de datos por ficha.

    Se abre en modo read_only y cada hoja se copia una sola vez a una
    `_SheetGrid` acotada (ver excel_diagnostico_territorial); asi un
    archivo con el rango usado inflado por Excel no agota CPU/memoria."""
    from excel_diagnostico_territorial import _SheetGrid
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    try:
        resultados = {}
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("_") or sheet_name.upper() == "INSTRUCTIVO":
                continue
            # Determinar tipo de ficha
            ficha = None
            for key in _PARSERS:
                if key in sheet_name:
                    ficha = key
                    break
            if ficha and ficha in _PARSERS:
                datos = _PARSERS[ficha](_SheetGrid(wb[sheet_name]))
                datos["ficha"] = ficha
                # Serializar listas a JSON
                for k, v in datos.items():
                    if isinstance(v, list):
                        datos[k] = json.dumps(v, ensure_ascii=False)
                resultados[ficha] = datos
        return resultados
    finally:
        wb.close()


# =========================================================================
# MAPEO A SESSION STATE
# =========================================================================

def mapear_a_session_state(datos_parseados, bloques_map):
    """Mapea datos parseados a claves de session_state de Streamlit."""
    ss = {}
    if not datos_parseados:
        return ss

    # Tomar la primera ficha disponible para datos generales
    first = next(iter(datos_parseados.values()))
    codigo_bloque = first.get("codigo_bloque", "")
    if codigo_bloque and bloques_map:
        for label, bid in bloques_map.items():
            if codigo_bloque in label:
                ss["ee_bl"] = label
                break

    ss["ee_fecha"] = first.get("fecha_campo", "")
    ss["ee_resp"] = first.get("responsable_brigada", "")
    ss["ee_cp"] = first.get("centro_poblado", "")
    ss["ee_este"] = first.get("coordenada_este", "")
    ss["ee_norte"] = first.get("coordenada_norte", "")
    ss["ee_alt"] = first.get("altitud", "")

    for ficha, datos in datos_parseados.items():
        ss[f"ee_ficha_sel"] = ficha

        if ficha == "F-EE-05":
            field_map = {
                "e05_eco": "ee05_tipo_ecosistema", "e05_zv": "ee05_zona_vida",
                "e05_cv": "ee05_cobertura_vegetal", "e05_pcv": "ee05_pct_cobertura",
                "e05_esp": "ee05_especies_dominantes",
                "e05_deg": "ee05_evidencia_degradacion",
                "e05_tdeg": "ee05_tipo_degradacion",
                "e05_ndeg": "ee05_nivel_degradacion",
                "e05_pend": "ee05_pendiente", "e05_suelo": "ee05_tipo_suelo",
                "e05_prof": "ee05_profundidad_efectiva",
                "e05_carc": "ee05_presencia_carcavas",
                "e05_queb": "ee05_presencia_quebrada",
                "e05_nqueb": "ee05_nombre_quebrada",
                "e05_fag": "ee05_fuentes_agua",
            }
            for sk, dk in field_map.items():
                ss[sk] = datos.get(dk, "")

        elif ficha == "F-EE-06":
            ss["e06_nvuln"] = datos.get("ee06_nivel_vulnerabilidad", "")
            ss["e06_nriesgo"] = datos.get("ee06_nivel_riesgo", "")

    return ss

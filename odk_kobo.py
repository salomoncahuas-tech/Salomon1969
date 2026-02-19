"""
IN Piura - Módulo de Integración ODK / KoBoToolbox
Generación de formularios XLSForm para colecta en campo sin conexión,
importación de datos recolectados (CSV/Excel) y sincronización
con API de KoBoToolbox.
Cuenca Alta del Río Piura, Perú.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import csv
import json
import os
import urllib.request
import urllib.error
import urllib.parse
import ssl

import database as db

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False


# ── Generador de formularios XLSForm ──────────────────────────────────────

XLSFORM_SURVEY = [
    # type, name, label::Spanish, required, constraint, appearance, calculation
    ("start", "inicio", "", "", "", "", ""),
    ("end", "fin", "", "", "", "", ""),
    ("today", "fecha_hoy", "", "", "", "", ""),
    ("deviceid", "id_dispositivo", "", "", "", "", ""),
    ("geopoint", "ubicacion_gps", "Ubicación GPS del punto de verificación", "yes", "", "maps", ""),
    ("begin_group", "datos_bloque", "Datos del Bloque de Intervención", "", "", "field-list", ""),
    ("text", "codigo_bloque", "Código del bloque", "yes", "", "", ""),
    ("select_one tipo_intervencion", "tipo_intervencion", "Tipo de intervención", "yes", "", "", ""),
    ("text", "cuenca", "Cuenca hidrográfica", "yes", "", "", ""),
    ("text", "distrito", "Distrito", "yes", "", "", ""),
    ("decimal", "utm_este", "Coordenada UTM Este (m)", "yes", ". > 0", "", ""),
    ("decimal", "utm_norte", "Coordenada UTM Norte (m)", "yes", ". > 0", "", ""),
    ("select_one zona_utm", "utm_zona", "Zona UTM", "yes", "", "", ""),
    ("decimal", "area_hectareas", "Área (hectáreas)", "yes", ". > 0", "", ""),
    ("select_one estado_bloque", "estado", "Estado actual", "yes", "", "", ""),
    ("end_group", "", "", "", "", "", ""),
    ("begin_group", "inspeccion_campo", "Inspección de Campo", "", "", "field-list", ""),
    ("date", "fecha_visita", "Fecha de visita de campo", "yes", "", "", ""),
    ("text", "inspector", "Nombre del inspector", "yes", "", "", ""),
    ("select_one condicion_clima", "condiciones_climaticas", "Condiciones climáticas", "yes", "", "", ""),
    ("integer", "avance_fisico", "Avance físico (%)", "yes", ". >= 0 and . <= 100", "", ""),
    ("text", "observaciones", "Observaciones técnicas", "no", "", "multiline", ""),
    ("text", "desviaciones", "Desviaciones al expediente técnico", "no", "", "multiline", ""),
    ("image", "foto_1", "Fotografía 1 - Vista general", "no", "", "", ""),
    ("image", "foto_2", "Fotografía 2 - Detalle", "no", "", "", ""),
    ("image", "foto_3", "Fotografía 3 - Evidencia adicional", "no", "", "", ""),
    ("end_group", "", "", "", "", "", ""),
    ("begin_group", "indicadores_calidad", "Indicadores de Calidad", "", "", "field-list", ""),
    ("decimal", "densidad_planificada", "Densidad de plantación planificada (pl/ha)", "no", ". >= 0", "", ""),
    ("decimal", "densidad_lograda", "Densidad de plantación lograda (pl/ha)", "no", ". >= 0", "", ""),
    ("decimal", "sobrevivencia_especies", "Sobrevivencia de especies (%)", "no", ". >= 0 and . <= 100", "", ""),
    ("decimal", "longitud_zanjas", "Longitud ejecutada de zanjas (ml)", "no", ". >= 0", "", ""),
    ("decimal", "volumen_retencion", "Volumen retención sedimentos (m³)", "no", ". >= 0", "", ""),
    ("end_group", "", "", "", "", "", ""),
    ("calculate", "codigo_verificacion", "", "", "", "", "concat('VER-', format-date(${fecha_hoy}, '%Y%m%d'), '-', substr(${id_dispositivo}, 0, 8))"),
    ("note", "nota_verificacion", "Código de verificación: ${codigo_verificacion}", "", "", "", ""),
]

XLSFORM_CHOICES = [
    # list_name, name, label::Spanish
    ("tipo_intervencion", "revegetacion", "Revegetación"),
    ("tipo_intervencion", "zanjas_infiltracion", "Zanjas de infiltración"),
    ("tipo_intervencion", "terrazas_formacion", "Terrazas de formación lenta"),
    ("tipo_intervencion", "diques_mamposteria", "Diques de mampostería"),
    ("zona_utm", "17S", "17S"),
    ("zona_utm", "17N", "17N"),
    ("zona_utm", "18S", "18S"),
    ("estado_bloque", "pendiente", "Pendiente"),
    ("estado_bloque", "en_progreso", "En progreso"),
    ("estado_bloque", "verificado", "Verificado"),
    ("condicion_clima", "despejado", "Despejado"),
    ("condicion_clima", "parcialmente_nublado", "Parcialmente nublado"),
    ("condicion_clima", "nublado", "Nublado"),
    ("condicion_clima", "lluvia_ligera", "Lluvia ligera"),
    ("condicion_clima", "lluvia_moderada", "Lluvia moderada"),
    ("condicion_clima", "lluvia_intensa", "Lluvia intensa"),
    ("condicion_clima", "neblina", "Neblina"),
]

XLSFORM_SETTINGS = {
    "form_title": "IN Piura - Verificación de Campo",
    "form_id": "in_piura_verificacion_campo",
    "version": datetime.now().strftime("%Y%m%d%H%M"),
    "default_language": "Spanish",
    "style": "pages",
}

# Mapeo de choices name -> label para importación
CHOICES_LABEL_MAP = {}
for list_name, name, label in XLSFORM_CHOICES:
    CHOICES_LABEL_MAP[name] = label

CHOICES_NAME_MAP = {}
for list_name, name, label in XLSFORM_CHOICES:
    CHOICES_NAME_MAP[label.lower()] = label
    CHOICES_NAME_MAP[name.lower()] = label


def generar_xlsform(ruta_salida=None):
    """Genera un archivo XLSForm (.xlsx) compatible con ODK/KoBoToolbox."""
    if not OPENPYXL_DISPONIBLE:
        raise ImportError("Se requiere openpyxl para generar formularios XLSForm.")

    if ruta_salida is None:
        directorio = os.path.join(os.path.dirname(os.path.abspath(__file__)), "formularios")
        os.makedirs(directorio, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta_salida = os.path.join(directorio, f"IN_Piura_XLSForm_{timestamp}.xlsx")

    wb = Workbook()

    # ── Hoja survey ──
    ws_survey = wb.active
    ws_survey.title = "survey"

    encabezados_survey = ["type", "name", "label::Spanish", "required",
                          "constraint", "appearance", "calculation"]
    for col, header in enumerate(encabezados_survey, 1):
        celda = ws_survey.cell(row=1, column=col, value=header)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="2C3E50", fill_type="solid")
        celda.alignment = Alignment(horizontal="center")

    for row_idx, fila in enumerate(XLSFORM_SURVEY, 2):
        for col_idx, valor in enumerate(fila, 1):
            ws_survey.cell(row=row_idx, column=col_idx, value=valor)

    # Ajustar anchos
    ws_survey.column_dimensions["A"].width = 28
    ws_survey.column_dimensions["B"].width = 26
    ws_survey.column_dimensions["C"].width = 50
    ws_survey.column_dimensions["D"].width = 10
    ws_survey.column_dimensions["E"].width = 25
    ws_survey.column_dimensions["F"].width = 14
    ws_survey.column_dimensions["G"].width = 60

    # ── Hoja choices ──
    ws_choices = wb.create_sheet("choices")

    encabezados_choices = ["list_name", "name", "label::Spanish"]
    for col, header in enumerate(encabezados_choices, 1):
        celda = ws_choices.cell(row=1, column=col, value=header)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="3498DB", fill_type="solid")
        celda.alignment = Alignment(horizontal="center")

    for row_idx, (list_name, name, label) in enumerate(XLSFORM_CHOICES, 2):
        ws_choices.cell(row=row_idx, column=1, value=list_name)
        ws_choices.cell(row=row_idx, column=2, value=name)
        ws_choices.cell(row=row_idx, column=3, value=label)

    ws_choices.column_dimensions["A"].width = 22
    ws_choices.column_dimensions["B"].width = 24
    ws_choices.column_dimensions["C"].width = 30

    # ── Hoja settings ──
    ws_settings = wb.create_sheet("settings")

    for col, header in enumerate(XLSFORM_SETTINGS.keys(), 1):
        celda = ws_settings.cell(row=1, column=col, value=header)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="27AE60", fill_type="solid")

    for col, valor in enumerate(XLSFORM_SETTINGS.values(), 1):
        ws_settings.cell(row=2, column=col, value=valor)

    wb.save(ruta_salida)
    return ruta_salida


# ── Importador de datos ODK/KoBoToolbox ───────────────────────────────────

def _normalizar_tipo(valor):
    """Normaliza el valor del tipo de intervención desde ODK."""
    mapa = {
        "revegetacion": "Revegetación",
        "zanjas_infiltracion": "Zanjas de infiltración",
        "terrazas_formacion": "Terrazas de formación lenta",
        "diques_mamposteria": "Diques de mampostería",
    }
    return mapa.get(valor.lower().strip(), valor)


def _normalizar_estado(valor):
    """Normaliza el valor del estado desde ODK."""
    mapa = {
        "pendiente": "Pendiente",
        "en_progreso": "En progreso",
        "verificado": "Verificado",
    }
    return mapa.get(valor.lower().strip(), valor)


def _normalizar_clima(valor):
    """Normaliza condiciones climáticas desde ODK."""
    mapa = {
        "despejado": "Despejado",
        "parcialmente_nublado": "Parcialmente nublado",
        "nublado": "Nublado",
        "lluvia_ligera": "Lluvia ligera",
        "lluvia_moderada": "Lluvia moderada",
        "lluvia_intensa": "Lluvia intensa",
        "neblina": "Neblina",
    }
    return mapa.get(valor.lower().strip(), valor)


def importar_csv_odk(ruta_csv):
    """Importa datos desde un archivo CSV exportado de ODK/KoBoToolbox.
    Retorna un dict con estadísticas de importación."""
    resultados = {
        "bloques_nuevos": 0,
        "bloques_actualizados": 0,
        "inspecciones_creadas": 0,
        "indicadores_creados": 0,
        "errores": [],
        "total_filas": 0,
    }

    with open(ruta_csv, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        filas = list(reader)

    resultados["total_filas"] = len(filas)

    for i, fila in enumerate(filas, 1):
        try:
            # Normalizar claves (ODK puede usar prefijos de grupo)
            datos = {}
            for k, v in fila.items():
                # Remover prefijos de grupo (e.g., "datos_bloque/codigo_bloque" -> "codigo_bloque")
                key_limpia = k.split("/")[-1].strip()
                datos[key_limpia] = v.strip() if v else ""

            codigo = datos.get("codigo_bloque", "")
            if not codigo:
                resultados["errores"].append(f"Fila {i}: código de bloque vacío")
                continue

            # Datos del bloque
            tipo = _normalizar_tipo(datos.get("tipo_intervencion", "Revegetación"))
            cuenca = datos.get("cuenca", "Cuenca Alta del Río Piura")
            distrito = datos.get("distrito", "")
            estado = _normalizar_estado(datos.get("estado", "Pendiente"))

            # Coordenadas: pueden venir del geopoint o de campos UTM
            utm_este = 0.0
            utm_norte = 0.0
            utm_zona = datos.get("utm_zona", "17S")

            # Intentar usar campos UTM directos
            try:
                utm_este = float(datos.get("utm_este", "0"))
                utm_norte = float(datos.get("utm_norte", "0"))
            except ValueError:
                pass

            # Si hay geopoint, extraer lat/lon y convertir a UTM
            geopoint = datos.get("ubicacion_gps", "")
            if geopoint and (utm_este == 0 or utm_norte == 0):
                partes = geopoint.split()
                if len(partes) >= 2:
                    try:
                        lat = float(partes[0])
                        lon = float(partes[1])
                        from georeferenciacion import latlon_a_utm
                        utm_este, utm_norte, utm_zona = latlon_a_utm(lat, lon)
                    except (ValueError, ImportError):
                        pass

            try:
                area = float(datos.get("area_hectareas", "0"))
            except ValueError:
                area = 0.0

            # Crear o actualizar bloque
            bloque_existente = db.obtener_bloque_por_codigo(codigo)
            if bloque_existente:
                db.actualizar_bloque(
                    bloque_existente["id"], codigo, tipo, cuenca,
                    distrito, utm_este, utm_norte, utm_zona, area, estado
                )
                bloque_id = bloque_existente["id"]
                resultados["bloques_actualizados"] += 1
            else:
                bloque_id = db.insertar_bloque(
                    codigo, tipo, cuenca, distrito,
                    utm_este, utm_norte, utm_zona, area, estado
                )
                resultados["bloques_nuevos"] += 1

            # Crear inspección si hay datos
            fecha_visita = datos.get("fecha_visita", "")
            inspector = datos.get("inspector", "")
            if fecha_visita and inspector:
                clima = _normalizar_clima(datos.get("condiciones_climaticas", "Despejado"))
                try:
                    avance = float(datos.get("avance_fisico", "0"))
                except ValueError:
                    avance = 0.0

                observaciones = datos.get("observaciones", "")
                desviaciones = datos.get("desviaciones", "")
                fotos = "; ".join(filter(None, [
                    datos.get("foto_1", ""),
                    datos.get("foto_2", ""),
                    datos.get("foto_3", ""),
                ]))
                codigo_ver = datos.get("codigo_verificacion", "")

                inspeccion_id = db.insertar_inspeccion(
                    bloque_id, fecha_visita, inspector, clima,
                    avance, observaciones, desviaciones, fotos, codigo_ver
                )
                resultados["inspecciones_creadas"] += 1

                # Crear indicadores si hay datos
                try:
                    dens_plan = float(datos.get("densidad_planificada", "0"))
                    dens_logr = float(datos.get("densidad_lograda", "0"))
                    sobrev = float(datos.get("sobrevivencia_especies", "0"))
                    zanjas = float(datos.get("longitud_zanjas", "0"))
                    vol_ret = float(datos.get("volumen_retencion", "0"))

                    if any([dens_plan, dens_logr, sobrev, zanjas, vol_ret]):
                        db.insertar_indicadores(
                            bloque_id, inspeccion_id, dens_plan,
                            dens_logr, sobrev, zanjas, vol_ret
                        )
                        resultados["indicadores_creados"] += 1
                except ValueError:
                    pass

        except Exception as e:
            resultados["errores"].append(f"Fila {i}: {str(e)}")

    return resultados


# ── Cliente API KoBoToolbox ───────────────────────────────────────────────

class KoBoClient:
    """Cliente para interactuar con la API de KoBoToolbox."""

    def __init__(self, url_servidor, token_api):
        self.url_servidor = url_servidor.rstrip("/")
        self.token_api = token_api
        # Crear contexto SSL que no verifica certificados (para entornos de desarrollo)
        self.ssl_context = ssl.create_default_context()
        self.ssl_context.check_hostname = False
        self.ssl_context.verify_mode = ssl.CERT_NONE

    def _hacer_peticion(self, endpoint, metodo="GET", datos=None):
        """Realiza una petición a la API de KoBoToolbox."""
        url = f"{self.url_servidor}/api/v2/{endpoint}"
        headers = {
            "Authorization": f"Token {self.token_api}",
            "Content-Type": "application/json",
        }

        if datos:
            data_bytes = json.dumps(datos).encode("utf-8")
        else:
            data_bytes = None

        req = urllib.request.Request(url, data=data_bytes, headers=headers,
                                     method=metodo)

        try:
            with urllib.request.urlopen(req, context=self.ssl_context,
                                        timeout=30) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            error_body = e.read().decode("utf-8", errors="replace")
            raise ConnectionError(
                f"Error HTTP {e.code}: {error_body}"
            )
        except urllib.error.URLError as e:
            raise ConnectionError(f"Error de conexión: {e.reason}")

    def listar_formularios(self):
        """Lista los formularios disponibles en la cuenta."""
        resultado = self._hacer_peticion("assets/?format=json")
        formularios = []
        for asset in resultado.get("results", []):
            if asset.get("asset_type") == "survey":
                formularios.append({
                    "uid": asset["uid"],
                    "nombre": asset.get("name", "Sin nombre"),
                    "fecha_modificacion": asset.get("date_modified", ""),
                    "envios": asset.get("deployment__submission_count", 0),
                    "desplegado": asset.get("has_deployment", False),
                })
        return formularios

    def obtener_envios(self, formulario_uid):
        """Obtiene todos los envíos (submissions) de un formulario."""
        resultado = self._hacer_peticion(
            f"assets/{formulario_uid}/data/?format=json"
        )
        return resultado.get("results", [])

    def test_conexion(self):
        """Verifica que la conexión y el token son válidos."""
        try:
            self._hacer_peticion("assets/?format=json&limit=1")
            return True, "Conexión exitosa"
        except ConnectionError as e:
            return False, str(e)
        except Exception as e:
            return False, f"Error inesperado: {str(e)}"


def importar_desde_kobo(url_servidor, token_api, formulario_uid):
    """Importa datos directamente desde la API de KoBoToolbox."""
    cliente = KoBoClient(url_servidor, token_api)
    envios = cliente.obtener_envios(formulario_uid)

    resultados = {
        "bloques_nuevos": 0,
        "bloques_actualizados": 0,
        "inspecciones_creadas": 0,
        "indicadores_creados": 0,
        "errores": [],
        "total_filas": len(envios),
    }

    for i, envio in enumerate(envios, 1):
        try:
            # Normalizar claves (KoBo usa prefijos de grupo con /)
            datos = {}
            for k, v in envio.items():
                key_limpia = k.split("/")[-1].strip()
                if isinstance(v, str):
                    datos[key_limpia] = v.strip()
                else:
                    datos[key_limpia] = v

            codigo = str(datos.get("codigo_bloque", ""))
            if not codigo:
                resultados["errores"].append(f"Envío {i}: código de bloque vacío")
                continue

            tipo = _normalizar_tipo(str(datos.get("tipo_intervencion", "Revegetación")))
            cuenca = str(datos.get("cuenca", "Cuenca Alta del Río Piura"))
            distrito = str(datos.get("distrito", ""))
            estado = _normalizar_estado(str(datos.get("estado", "Pendiente")))

            utm_este = 0.0
            utm_norte = 0.0
            utm_zona = str(datos.get("utm_zona", "17S"))

            try:
                utm_este = float(datos.get("utm_este", 0))
                utm_norte = float(datos.get("utm_norte", 0))
            except (ValueError, TypeError):
                pass

            geopoint = str(datos.get("ubicacion_gps", ""))
            if geopoint and (utm_este == 0 or utm_norte == 0):
                partes = geopoint.split()
                if len(partes) >= 2:
                    try:
                        lat = float(partes[0])
                        lon = float(partes[1])
                        from georeferenciacion import latlon_a_utm
                        utm_este, utm_norte, utm_zona = latlon_a_utm(lat, lon)
                    except (ValueError, ImportError):
                        pass

            try:
                area = float(datos.get("area_hectareas", 0))
            except (ValueError, TypeError):
                area = 0.0

            bloque_existente = db.obtener_bloque_por_codigo(codigo)
            if bloque_existente:
                db.actualizar_bloque(
                    bloque_existente["id"], codigo, tipo, cuenca,
                    distrito, utm_este, utm_norte, utm_zona, area, estado
                )
                bloque_id = bloque_existente["id"]
                resultados["bloques_actualizados"] += 1
            else:
                bloque_id = db.insertar_bloque(
                    codigo, tipo, cuenca, distrito,
                    utm_este, utm_norte, utm_zona, area, estado
                )
                resultados["bloques_nuevos"] += 1

            fecha_visita = str(datos.get("fecha_visita", ""))
            inspector = str(datos.get("inspector", ""))
            if fecha_visita and inspector:
                clima = _normalizar_clima(str(datos.get("condiciones_climaticas", "Despejado")))
                try:
                    avance = float(datos.get("avance_fisico", 0))
                except (ValueError, TypeError):
                    avance = 0.0

                observaciones = str(datos.get("observaciones", ""))
                desviaciones = str(datos.get("desviaciones", ""))
                codigo_ver = str(datos.get("codigo_verificacion", ""))

                inspeccion_id = db.insertar_inspeccion(
                    bloque_id, fecha_visita, inspector, clima,
                    avance, observaciones, desviaciones, "", codigo_ver
                )
                resultados["inspecciones_creadas"] += 1

                try:
                    dens_plan = float(datos.get("densidad_planificada", 0))
                    dens_logr = float(datos.get("densidad_lograda", 0))
                    sobrev = float(datos.get("sobrevivencia_especies", 0))
                    zanjas = float(datos.get("longitud_zanjas", 0))
                    vol_ret = float(datos.get("volumen_retencion", 0))

                    if any([dens_plan, dens_logr, sobrev, zanjas, vol_ret]):
                        db.insertar_indicadores(
                            bloque_id, inspeccion_id, dens_plan,
                            dens_logr, sobrev, zanjas, vol_ret
                        )
                        resultados["indicadores_creados"] += 1
                except (ValueError, TypeError):
                    pass

        except Exception as e:
            resultados["errores"].append(f"Envío {i}: {str(e)}")

    return resultados


# ── Tab ODK / KoBoToolbox ─────────────────────────────────────────────────

class TabODKKobo(ttk.Frame):
    """Pestaña de integración con ODK/KoBoToolbox para colecta en campo."""

    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.kobo_client = None
        self._crear_widgets()

    def _crear_widgets(self):
        # Canvas con scroll
        canvas = tk.Canvas(self, bg="#ECF0F1", highlightthickness=0)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.scroll_frame = ttk.Frame(canvas)
        self.scroll_frame.bind("<Configure>",
                               lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        frame = self.scroll_frame
        frame.columnconfigure(1, weight=1)

        # ── Sección 1: Generador de Formularios XLSForm ──
        ttk.Label(frame, text="ODK / KoBoToolbox - Colecta en Campo",
                  style="Header.TLabel").grid(
            row=0, column=0, columnspan=3, sticky="w", padx=12, pady=(12, 4))

        ttk.Label(frame,
                  text="Genere formularios XLSForm para recolección de datos sin conexión "
                       "con ODK Collect o KoBoCollect.",
                  wraplength=700).grid(
            row=1, column=0, columnspan=3, sticky="w", padx=12, pady=(0, 10))

        # Sub-sección: Generar formulario
        sec1 = ttk.LabelFrame(frame, text=" Generar Formulario XLSForm ", padding=12)
        sec1.grid(row=2, column=0, columnspan=3, sticky="ew", padx=12, pady=6)
        sec1.columnconfigure(1, weight=1)

        ttk.Label(sec1,
                  text="El formulario incluye campos para: ubicación GPS, datos del bloque,\n"
                       "inspección de campo, registro fotográfico e indicadores de calidad.",
                  wraplength=600).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))

        ttk.Button(sec1, text="Generar Formulario XLSForm",
                   command=self._generar_formulario,
                   style="Accent.TButton").grid(row=1, column=0, sticky="w", padx=0, pady=4)

        ttk.Label(sec1,
                  text="Compatible con: ODK Collect, KoBoCollect, Enketo",
                  foreground="#7F8C8D").grid(row=1, column=1, sticky="w", padx=12)

        self.label_estado_form = ttk.Label(sec1, text="", wraplength=600)
        self.label_estado_form.grid(row=2, column=0, columnspan=3, sticky="w", pady=4)

        # ── Sección 2: Importar datos CSV ──
        sec2 = ttk.LabelFrame(frame, text=" Importar Datos desde CSV / Excel ", padding=12)
        sec2.grid(row=3, column=0, columnspan=3, sticky="ew", padx=12, pady=6)
        sec2.columnconfigure(1, weight=1)

        ttk.Label(sec2,
                  text="Importe datos recolectados en campo exportados como CSV desde\n"
                       "ODK Central, KoBoToolbox u otro servidor ODK compatible.",
                  wraplength=600).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))

        btn_csv_frame = ttk.Frame(sec2)
        btn_csv_frame.grid(row=1, column=0, columnspan=3, sticky="w", pady=4)

        ttk.Button(btn_csv_frame, text="Seleccionar archivo CSV...",
                   command=self._importar_csv,
                   style="Accent.TButton").pack(side="left", padx=(0, 8))

        ttk.Button(btn_csv_frame, text="Descargar plantilla CSV",
                   command=self._descargar_plantilla_csv).pack(side="left", padx=4)

        self.label_estado_csv = ttk.Label(sec2, text="", wraplength=600)
        self.label_estado_csv.grid(row=2, column=0, columnspan=3, sticky="w", pady=4)

        # Tabla de resultados de importación
        self.frame_resultados = ttk.Frame(sec2)
        self.frame_resultados.grid(row=3, column=0, columnspan=3, sticky="ew", pady=4)

        # ── Sección 3: Conexión API KoBoToolbox ──
        sec3 = ttk.LabelFrame(frame, text=" Sincronización con KoBoToolbox (API) ", padding=12)
        sec3.grid(row=4, column=0, columnspan=3, sticky="ew", padx=12, pady=6)
        sec3.columnconfigure(1, weight=1)

        ttk.Label(sec3,
                  text="Conecte directamente con su cuenta de KoBoToolbox para\n"
                       "descargar envíos automáticamente (requiere conexión a internet).",
                  wraplength=600).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))

        ttk.Label(sec3, text="Servidor:").grid(row=1, column=0, sticky="w", pady=3)
        self.combo_servidor = ttk.Combobox(sec3, values=[
            "https://kf.kobotoolbox.org",
            "https://kobo.humanitarianresponse.info",
        ], width=40)
        self.combo_servidor.grid(row=1, column=1, sticky="w", padx=6, pady=3)
        self.combo_servidor.set("https://kf.kobotoolbox.org")

        ttk.Label(sec3, text="Token API:").grid(row=2, column=0, sticky="w", pady=3)
        self.entry_token = ttk.Entry(sec3, width=42, show="*")
        self.entry_token.grid(row=2, column=1, sticky="w", padx=6, pady=3)

        ttk.Label(sec3, text="",
                  foreground="#7F8C8D").grid(row=3, column=0, columnspan=2, sticky="w")

        btn_api_frame = ttk.Frame(sec3)
        btn_api_frame.grid(row=4, column=0, columnspan=3, sticky="w", pady=6)

        ttk.Button(btn_api_frame, text="Probar Conexión",
                   command=self._probar_conexion).pack(side="left", padx=(0, 8))
        ttk.Button(btn_api_frame, text="Listar Formularios",
                   command=self._listar_formularios).pack(side="left", padx=4)
        ttk.Button(btn_api_frame, text="Importar Envíos",
                   command=self._importar_envios_kobo,
                   style="Accent.TButton").pack(side="left", padx=4)

        self.label_estado_api = ttk.Label(sec3, text="", wraplength=600)
        self.label_estado_api.grid(row=5, column=0, columnspan=3, sticky="w", pady=4)

        # Lista de formularios
        ttk.Label(sec3, text="Formularios disponibles:").grid(
            row=6, column=0, columnspan=3, sticky="w", pady=(8, 2))

        columnas_form = ("uid", "nombre", "envios", "estado")
        self.tree_formularios = ttk.Treeview(sec3, columns=columnas_form,
                                              show="headings", height=5)
        self.tree_formularios.heading("uid", text="UID")
        self.tree_formularios.heading("nombre", text="Nombre")
        self.tree_formularios.heading("envios", text="Envíos")
        self.tree_formularios.heading("estado", text="Estado")
        self.tree_formularios.column("uid", width=120)
        self.tree_formularios.column("nombre", width=280)
        self.tree_formularios.column("envios", width=70)
        self.tree_formularios.column("estado", width=100)

        self.tree_formularios.grid(row=7, column=0, columnspan=3, sticky="ew", pady=4)

        # ── Sección 4: Guía rápida ──
        sec4 = ttk.LabelFrame(frame, text=" Guía Rápida de Uso ", padding=12)
        sec4.grid(row=5, column=0, columnspan=3, sticky="ew", padx=12, pady=(6, 12))

        guia_texto = (
            "Flujo de trabajo para colecta en campo sin conexión:\n\n"
            "1. GENERAR: Cree el formulario XLSForm desde esta pestaña\n"
            "2. SUBIR: Cargue el archivo .xlsx en KoBoToolbox (kobotoolbox.org) o ODK Central\n"
            "3. DESPLEGAR: Active el formulario en la plataforma\n"
            "4. RECOLECTAR: Use KoBoCollect u ODK Collect en sus dispositivos móviles\n"
            "   - Los datos se guardan localmente sin necesidad de internet\n"
            "   - Incluye captura GPS automática y fotografías\n"
            "5. SINCRONIZAR: Al tener conexión, envíe los datos al servidor\n"
            "6. IMPORTAR: Descargue el CSV desde la plataforma e impórtelo aquí,\n"
            "   o use la conexión API directa para sincronizar automáticamente"
        )
        ttk.Label(sec4, text=guia_texto, wraplength=680,
                  justify="left").pack(anchor="w")

    def _generar_formulario(self):
        """Genera el formulario XLSForm."""
        try:
            ruta = generar_xlsform()
            self.label_estado_form.config(
                text=f"Formulario generado exitosamente:\n{ruta}",
                foreground="#27AE60")
            messagebox.showinfo("XLSForm Generado",
                                f"Formulario listo para subir a KoBoToolbox/ODK Central:\n\n{ruta}")
        except ImportError:
            messagebox.showerror("Dependencia faltante",
                                 "Se requiere openpyxl para generar formularios XLSForm.\n"
                                 "Instale con: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el formulario:\n{e}")

    def _importar_csv(self):
        """Importa datos desde un archivo CSV."""
        ruta = filedialog.askopenfilename(
            title="Seleccionar archivo CSV de ODK/KoBoToolbox",
            filetypes=[("CSV", "*.csv"), ("Todos", "*.*")]
        )
        if not ruta:
            return

        try:
            resultados = importar_csv_odk(ruta)
            self._mostrar_resultados_importacion(resultados)
            self.app.refrescar_todo()
        except Exception as e:
            messagebox.showerror("Error de importación",
                                 f"No se pudo importar el archivo:\n{e}")

    def _descargar_plantilla_csv(self):
        """Genera un archivo CSV de plantilla con los encabezados esperados."""
        ruta = filedialog.asksaveasfilename(
            title="Guardar plantilla CSV",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile="plantilla_in_piura_odk.csv"
        )
        if not ruta:
            return

        encabezados = [
            "codigo_bloque", "tipo_intervencion", "cuenca", "distrito",
            "utm_este", "utm_norte", "utm_zona", "area_hectareas", "estado",
            "ubicacion_gps", "fecha_visita", "inspector",
            "condiciones_climaticas", "avance_fisico",
            "observaciones", "desviaciones",
            "foto_1", "foto_2", "foto_3",
            "densidad_planificada", "densidad_lograda",
            "sobrevivencia_especies", "longitud_zanjas", "volumen_retencion",
            "codigo_verificacion",
        ]

        # Fila de ejemplo
        ejemplo = [
            "BLQ-001", "revegetacion", "Cuenca Alta del Río Piura",
            "Canchaque", "622150.50", "9436720.30", "17S", "2.5000",
            "pendiente", "-5.3456 -79.6123 1850 10", "2026-01-15",
            "Juan Pérez", "despejado", "45",
            "Plantación en buen estado", "Ninguna",
            "", "", "", "1100", "850", "78.5", "120.5", "35.2",
            "VER-20260115-ABC12345",
        ]

        with open(ruta, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(encabezados)
            writer.writerow(ejemplo)

        self.label_estado_csv.config(
            text=f"Plantilla guardada en: {ruta}",
            foreground="#27AE60")
        messagebox.showinfo("Plantilla CSV",
                            f"Plantilla descargada correctamente:\n{ruta}")

    def _mostrar_resultados_importacion(self, resultados):
        """Muestra los resultados de la importación."""
        # Limpiar resultados anteriores
        for widget in self.frame_resultados.winfo_children():
            widget.destroy()

        total = resultados["total_filas"]
        nuevos = resultados["bloques_nuevos"]
        actualizados = resultados["bloques_actualizados"]
        insp = resultados["inspecciones_creadas"]
        indic = resultados["indicadores_creados"]
        errores = resultados["errores"]

        resumen = (
            f"Importación completada: {total} registro(s) procesados\n"
            f"  Bloques nuevos: {nuevos}\n"
            f"  Bloques actualizados: {actualizados}\n"
            f"  Inspecciones creadas: {insp}\n"
            f"  Indicadores creados: {indic}"
        )

        if errores:
            resumen += f"\n  Errores: {len(errores)}"

        color = "#27AE60" if not errores else "#F39C12"
        self.label_estado_csv.config(text=resumen, foreground=color)

        if errores:
            ttk.Label(self.frame_resultados, text="Errores:",
                      font=("Segoe UI", 9, "bold")).pack(anchor="w")
            for err in errores[:10]:
                ttk.Label(self.frame_resultados, text=f"  - {err}",
                          foreground="#E74C3C", wraplength=580).pack(anchor="w")
            if len(errores) > 10:
                ttk.Label(self.frame_resultados,
                          text=f"  ... y {len(errores) - 10} errores más",
                          foreground="#E74C3C").pack(anchor="w")

        if total > 0 and not errores:
            messagebox.showinfo("Importación Exitosa", resumen)

    def _probar_conexion(self):
        """Prueba la conexión con KoBoToolbox."""
        url = self.combo_servidor.get().strip()
        token = self.entry_token.get().strip()

        if not url or not token:
            messagebox.showwarning("Validación",
                                   "Ingrese la URL del servidor y el token API.")
            return

        self.label_estado_api.config(text="Conectando...", foreground="#3498DB")
        self.update_idletasks()

        cliente = KoBoClient(url, token)
        exito, mensaje = cliente.test_conexion()

        if exito:
            self.kobo_client = cliente
            self.label_estado_api.config(text=f"Conexión exitosa con {url}",
                                         foreground="#27AE60")
        else:
            self.label_estado_api.config(text=f"Error: {mensaje}",
                                         foreground="#E74C3C")

    def _listar_formularios(self):
        """Lista los formularios disponibles en KoBoToolbox."""
        if not self.kobo_client:
            self._probar_conexion()
            if not self.kobo_client:
                return

        try:
            self.label_estado_api.config(text="Cargando formularios...",
                                         foreground="#3498DB")
            self.update_idletasks()

            formularios = self.kobo_client.listar_formularios()

            # Limpiar tabla
            for item in self.tree_formularios.get_children():
                self.tree_formularios.delete(item)

            for form in formularios:
                estado = "Desplegado" if form["desplegado"] else "Borrador"
                self.tree_formularios.insert("", "end", iid=form["uid"], values=(
                    form["uid"],
                    form["nombre"],
                    form["envios"],
                    estado,
                ))

            self.label_estado_api.config(
                text=f"{len(formularios)} formulario(s) encontrados",
                foreground="#27AE60")

        except Exception as e:
            self.label_estado_api.config(text=f"Error: {e}",
                                         foreground="#E74C3C")

    def _importar_envios_kobo(self):
        """Importa envíos del formulario seleccionado en KoBoToolbox."""
        if not self.kobo_client:
            self._probar_conexion()
            if not self.kobo_client:
                return

        sel = self.tree_formularios.selection()
        if not sel:
            messagebox.showwarning("Selección",
                                   "Seleccione un formulario de la lista para importar.")
            return

        formulario_uid = sel[0]

        if not messagebox.askyesno("Confirmar Importación",
                                    f"¿Importar envíos del formulario {formulario_uid}?\n"
                                    "Los bloques existentes con el mismo código serán actualizados."):
            return

        try:
            self.label_estado_api.config(text="Importando datos...",
                                         foreground="#3498DB")
            self.update_idletasks()

            url = self.combo_servidor.get().strip()
            token = self.entry_token.get().strip()

            resultados = importar_desde_kobo(url, token, formulario_uid)
            self._mostrar_resultados_importacion(resultados)
            self.app.refrescar_todo()

            self.label_estado_api.config(
                text=f"Importación completada: {resultados['total_filas']} envíos procesados",
                foreground="#27AE60")

        except Exception as e:
            self.label_estado_api.config(text=f"Error: {e}",
                                         foreground="#E74C3C")
            messagebox.showerror("Error de Importación",
                                 f"No se pudieron importar los datos:\n{e}")

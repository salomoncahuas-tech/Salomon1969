"""
Modulo de importacion/exportacion Excel para Diagnostico Territorial V5.
Proyecto IN Piura CUI 2669244 | ANIN - DIME - SESDI

Reemplaza el formato anterior (F-DT-01..06) por la plantilla V5 oficial:
F-DT-01 a F-DT-05 con inventarios tabulares (carcavas, floristica, especies
clave, fuentes de agua, matriz de causas e indicadores cuantitativos).

API publica:
    generar_plantilla_dt(fichas, bloques_data) -> bytes
    parsear_excel_dt(file_bytes, ficha=None) -> list[dict]
    mapear_dt_a_session_state(datos_parseados, bloques_map) -> dict
"""

import io
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ─── Plantilla V5 oficial ─────────────────────────────────────────────────
# El archivo .xlsx con la plantilla V5 oficial (con todas las hojas, listas
# de validacion, etc.) vive junto al codigo en el repo. Lo usamos como base
# para la descarga; solo reemplazamos la hoja oculta '_Datos' / 'Bloques V4'
# con los bloques actuales del aplicativo (V5).
PLANTILLA_V5_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "Plantilla_DT_Campo_Check_Validada_V5.xlsx",
)


# ─── Generacion de plantilla descargable ──────────────────────────────────

def generar_plantilla_dt(fichas=None, bloques_data=None):
    """Genera el archivo Excel V5 listo para descarga.

    Toma como base el archivo V5 oficial del repo y, opcionalmente,
    sobreescribe la hoja 'Bloques V4' con la lista de bloques que el
    aplicativo conoce (asi el dropdown del Excel siempre refleja la BD).

    Args:
        fichas: ignorado en V5; el archivo trae las 5 fichas siempre. Se
                conserva en la firma por compatibilidad.
        bloques_data: lista de tuplas
            (codigo, microcuenca, provincia, distrito, area_ha,
             utm_este, utm_norte, zona)

    Returns:
        bytes con el contenido del .xlsx.
    """
    if not os.path.exists(PLANTILLA_V5_PATH):
        raise FileNotFoundError(
            "No se encontro la plantilla V5: "
            f"{PLANTILLA_V5_PATH}. Verifique el repositorio.")

    # Trabajar sobre una copia en memoria para no tocar el archivo de disco.
    with open(PLANTILLA_V5_PATH, "rb") as fh:
        raw = fh.read()
    buf_in = io.BytesIO(raw)
    wb = load_workbook(buf_in)

    if bloques_data and "Bloques V4" in wb.sheetnames:
        ws = wb["Bloques V4"]
        # Conservar fila 1 (cabecera) y limpiar el resto.
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
        # Cabecera esperada V5: codigo | microcuenca | provincia | distrito
        # | zona | este_utm | norte_utm | area_ha
        for i, b in enumerate(bloques_data, start=2):
            codigo      = b[0] if len(b) > 0 else ""
            microcuenca = b[1] if len(b) > 1 else ""
            provincia   = b[2] if len(b) > 2 else ""
            distrito    = b[3] if len(b) > 3 else ""
            area_ha     = b[4] if len(b) > 4 else ""
            utm_e       = b[5] if len(b) > 5 else ""
            utm_n       = b[6] if len(b) > 6 else ""
            zona        = b[7] if len(b) > 7 else ""
            ws.cell(row=i, column=1, value=codigo)
            ws.cell(row=i, column=2, value=microcuenca)
            ws.cell(row=i, column=3, value=provincia)
            ws.cell(row=i, column=4, value=distrito)
            ws.cell(row=i, column=5, value=zona)
            ws.cell(row=i, column=6, value=utm_e)
            ws.cell(row=i, column=7, value=utm_n)
            ws.cell(row=i, column=8, value=area_ha)

    buf_out = io.BytesIO()
    wb.save(buf_out)
    buf_out.seek(0)
    return buf_out.getvalue()


# ─── Parseo del Excel V5 llenado ──────────────────────────────────────────

def _val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    return str(v).strip()


# Hints que indican que una celda es una ETIQUETA (no un valor capturado).
# Usados por _scan_label_value para no devolver el texto del rotulo de al
# lado cuando el campo objetivo esta vacio o su formula (auto) no tiene
# valor en cache.
_LABEL_HINTS = (
    "(auto)", "(dd-mm-aaaa)", "(msnm)", "(ha)", "(m × m)",
    "(% bloque)", "(% area)", "(ua/ha)", "(años aprox.)",
    "(min)", "(síntesis)", "(sintesis)", "(cualitativa)",
    "(preliminar)", "(descripción libre)", "(descripcion libre)",
    "(si aplica)", "(motor)", "(m)", "(cm)", "(%)",
    "(época de visita)", "(epoca de visita)",
)


def _looks_like_label(text):
    """Heuristica: ¿la celda parece ser un rotulo/etiqueta y no un valor
    capturado por el usuario? Se usa para detener el barrido lateral de
    `_scan_label_value` cuando llega a la siguiente etiqueta."""
    if not text:
        return False
    t = str(text).strip()
    if not t:
        return False
    if t.endswith("*") or t.endswith("?") or t.endswith(":"):
        return True
    if " *" in t or "*  " in t:
        return True
    tl = t.lower()
    for h in _LABEL_HINTS:
        if h in tl:
            return True
    if t.startswith("☐") or "▼" in t:
        return True
    return False


def _collect_label_anchors(ws, vocabulary):
    """Pre-escanea el worksheet y devuelve un set de (row, col) con las
    celdas que parecen ser etiquetas. Una celda es ancla si:
      - su texto contiene cualquier keyword del `vocabulary` (lista de
        keywords usados por el parser de la ficha), o
      - su texto cumple `_looks_like_label` (marcadores *, ?, :, (auto)...).
    Estas anclas se usan para que el barrido lateral de
    `_scan_label_value` se detenga ANTES de devolver el rotulo vecino
    como valor."""
    vocab_l = [k.lower() for k in vocabulary]
    anchors = set()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            # Saltar textos largos (subtitulos/descripciones): no son
            # rotulos de campo individual.
            if len(s) > 120:
                continue
            sl = s.lower()
            if _looks_like_label(s) or any(kw in sl for kw in vocab_l):
                anchors.add((r, c))
    return anchors


def _scan_label_value(ws, label_keywords, search_offsets=(1, 2, 3, 4, 5, 6),
                     anchors=None):
    """Busca un label cuyo texto contenga alguna palabra clave y devuelve
    el primer valor poblado a la derecha. Si al barrer a la derecha se
    encuentra otra etiqueta (`anchors` o `_looks_like_label`) antes que
    un valor, devuelve cadena vacia. Asi se evita devolver el rotulo
    vecino cuando el campo objetivo esta vacio o su formula `(auto)` no
    tiene valor en cache."""
    label_keywords_l = [k.lower() for k in label_keywords]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cv = ws.cell(row=r, column=c).value
            if not cv:
                continue
            cv_s = str(cv).strip()
            # Saltar textos de subtitulo/descripcion: son demasiado largos
            # para ser un rotulo y suelen contener varias palabras clave
            # (ej. la cabecera de cada ficha en la fila 5).
            if len(cv_s) > 120:
                continue
            cv_l = cv_s.lower()
            if any(kw in cv_l for kw in label_keywords_l):
                for off in search_offsets:
                    nc = c + off
                    nv = ws.cell(row=r, column=nc).value
                    if nv is None:
                        continue
                    s = str(nv).strip()
                    if s == "":
                        continue
                    if anchors is not None and (r, nc) in anchors:
                        return ""
                    if _looks_like_label(s):
                        return ""
                    return s
                return ""
    return ""


def _scan_observaciones(ws, label_row=None, anchors=None):
    """Busca la celda etiquetada 'OBSERVACIONES' (exacta, no como
    sub-cadena dentro de otro rotulo) y devuelve el texto escrito DEBAJO
    del rotulo. Si `label_row` se pasa, se usa esa fila directamente
    (evita falsos positivos con encabezados de columna del tipo
    'Uso observado / Observaciones')."""
    target_row = None
    target_col = None
    if label_row is not None:
        # Validar que efectivamente hay un rotulo OBSERVACIONES en esa fila.
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=label_row, column=c).value
            if v and str(v).strip().lower() == "observaciones":
                target_row = label_row
                target_col = c
                break
    if target_row is None:
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if not v:
                    continue
                # Coincidencia EXACTA con 'observaciones' (admite mayusculas).
                if str(v).strip().lower() == "observaciones":
                    target_row = r
                    target_col = c
                    break
            if target_row is not None:
                break
    if target_row is None:
        return ""
    pieces = []
    for dr in range(target_row + 1, min(target_row + 8, ws.max_row + 1)):
        for dc in range(1, ws.max_column + 1):
            v = ws.cell(row=dr, column=dc).value
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            if anchors is not None and (dr, dc) in anchors:
                continue
            if _looks_like_label(s):
                continue
            pieces.append(s)
    return " ".join(pieces)


def _scan_table(ws, header_keywords, num_data_rows, num_cols, start_col=None):
    """Busca una fila que contenga las cabeceras dadas (todas presentes)
    y devuelve hasta `num_data_rows` filas siguientes como lista de listas
    de strings (longitud `num_cols`).

    `start_col`: si se omite, se infiere como la columna NO VACIA mas a
    la izquierda dentro de la fila cabecera detectada. Antes se usaba la
    columna de la primera keyword, pero eso desplazaba la tabla cuando la
    primera keyword no aparecia en la columna 1 (p.ej. 'Categoría' en
    DT-03 esta en C3, pero la tabla empieza en C1 con 'N°')."""
    header_keywords_l = [k.lower() for k in header_keywords]
    for r in range(1, ws.max_row + 1):
        row_vals = [str(ws.cell(row=r, column=c).value or "").lower()
                    for c in range(1, ws.max_column + 1)]
        # Cada keyword debe aparecer en una COLUMNA distinta (asi se evita
        # confundir el rotulo cabecera real con la fila de "Use los
        # desplegables..." que mete todos los nombres en una sola celda).
        matched_cols = set()
        ok = True
        for kw in header_keywords_l:
            hit = False
            for c, v in enumerate(row_vals, start=1):
                if c in matched_cols:
                    continue
                if kw in v:
                    matched_cols.add(c)
                    hit = True
                    break
            if not hit:
                ok = False
                break
        if ok:
            if start_col is None:
                # Columna no-vacia mas a la izquierda en la fila cabecera.
                sc = 1
                for c, v in enumerate(row_vals, start=1):
                    if v.strip():
                        sc = c
                        break
            else:
                sc = start_col
            data = []
            for dr in range(r + 1, r + 1 + num_data_rows):
                row = []
                for dc in range(sc, sc + num_cols):
                    v = ws.cell(row=dr, column=dc).value
                    row.append("" if v is None else str(v).strip())
                if any(x for x in row):
                    data.append(row)
            return data
    return []


def _parse_dt01(ws):
    label_map = {
        "fecha_evaluacion": ["fecha"],
        "hora_registro": ["hora de registro"],
        "evaluador": ["brigada / responsable", "responsable", "evaluador"],
        "ficha_correlativo": ["ficha n", "correlativo"],
        "codigo_bloque": ["código del bloque", "codigo del bloque"],
        "microcuenca": ["microcuenca"],
        "provincia": ["provincia"],
        "distrito": ["distrito"],
        "utm_este_dt": ["utm este punto", "utm este"],
        "utm_norte_dt": ["utm norte punto", "utm norte"],
        "altitud_gps": ["altitud gps", "altitud"],
        "centro_poblado_cercano": ["centro poblado"],
        "comunidad_campesina_dt": ["comunidad campesina"],
        "forma_terreno": ["forma predominante", "forma del terreno"],
        "pendiente": ["rango de pendiente", "pendiente dominante"],
        "posicion_fisiografica": ["posición fisiográfica", "posicion fisiografica"],
        "exposicion_orientacion": ["exposición", "exposicion"],
        "rango_altitudinal": ["rango altitudinal"],
        "paisaje_dominante": ["paisaje dominante"],
        "dt01_afloramientos_rocosos": ["afloramientos rocosos"],
        "dt01_escarpes_activos": ["escarpes activos"],
        "dt01_reptacion_suelo": ["reptación", "reptacion"],
        "dt01_deslizamientos_antiguos": ["deslizamientos antiguos"],
        "dt01_remociones_masa_activas": ["remociones en masa activas"],
    }
    vocab = [kw for kws in label_map.values() for kw in kws]
    anchors = _collect_label_anchors(ws, vocab)
    out = {k: _scan_label_value(ws, kws, anchors=anchors)
           for k, kws in label_map.items()}
    out["dt01_observaciones"] = _scan_observaciones(ws, anchors=anchors)
    return out


def _parse_dt02(ws):
    label_map = {
        "fecha_evaluacion": ["fecha"],
        "evaluador": ["responsable", "evaluador"],
        "codigo_bloque": ["código del bloque", "codigo del bloque"],
        "microcuenca": ["microcuenca"],
        "distrito": ["distrito"],
        "altitud_gps": ["altitud"],
        "utm_este_dt": ["utm este"],
        "utm_norte_dt": ["utm norte"],
        "dt02_sellamiento_costra": ["sellamiento", "costra"],
        "dt02_compactacion_pisoteo": ["compactación", "compactacion"],
        "dt02_raices_expuestas": ["raíces expuestas", "raices expuestas"],
        "dt02_nivel_erosion_general": ["nivel general de erosión observado", "nivel general de erosion observado"],
        "dt02_nivel_erosion_sintesis": ["nivel general de erosión (síntesis)", "nivel general de erosion (sintesis)"],
        "dt02_num_carcavas": ["n° total de cárcavas", "n total de carcavas"],
        "dt02_longitud_total_carcavas": ["longitud total de cárcavas", "longitud total de carcavas"],
        "dt02_pct_bloque_carcavas": ["% del bloque afectado"],
        "dt02_erosion_laminar_pct": ["erosión laminar observada", "erosion laminar observada"],
        "dt02_patron_carcavas": ["patrón de cárcavas", "patron de carcavas"],
        "dt02_socavamiento_cauce": ["socavamiento de cauce"],
        "dt02_urgencia_control": ["urgencia de control"],
    }
    vocab = [kw for kws in label_map.values() for kw in kws]
    anchors = _collect_label_anchors(ws, vocab)
    base = {k: _scan_label_value(ws, kws, anchors=anchors)
            for k, kws in label_map.items()}
    base["dt02_observaciones"] = _scan_observaciones(ws, anchors=anchors)
    # Tabla de carcavas (12 columnas: codigo, tipo, utm_e_ini, utm_n_ini,
    # utm_e_fin, utm_n_fin, longitud, prof, ancho, estado, causa, foto)
    table = _scan_table(ws, ["código", "tipo"], num_data_rows=10, num_cols=12)
    keys = ["codigo", "tipo", "utm_e_ini", "utm_n_ini", "utm_e_fin",
            "utm_n_fin", "longitud_m", "prof_m", "ancho_m", "estado",
            "causa", "foto"]
    base["dt02_carcavas"] = [
        {k: v for k, v in zip(keys, row)} for row in table
    ]
    return base


def _parse_dt03(ws):
    label_map = {
        "fecha_evaluacion": ["fecha"],
        "evaluador": ["responsable", "evaluador"],
        "codigo_bloque": ["código del bloque", "codigo del bloque"],
        "microcuenca": ["microcuenca"],
        "distrito": ["distrito"],
        "dt03_parcela_muestreo": ["parcela de muestreo"],
        "dt03_dim_parcela": ["dimensiones de parcela"],
        "altitud_gps": ["altitud"],
        "dt03_pendiente_parcela": ["pendiente promedio"],
        "dt03_cobertura_total": ["cobertura vegetal total"],
        "utm_este_dt": ["utm este"],
        "utm_norte_dt": ["utm norte"],
        "dt03_tipo_ecosistema": ["tipo de ecosistema"],
        "dt03_superficie_ecosistema": ["superficie del ecosistema"],
        "dt03_estado_conservacion_eco": ["estado de conservación general", "estado de conservacion general"],
        "dt03_uso_dominante": ["uso actual dominante"],
        "dt03_cobertura_dosel": ["cobertura del dosel arbóreo", "cobertura del dosel arboreo"],
        "dt03_cobertura_arbustiva": ["cobertura arbustiva"],
        "dt03_cobertura_herbacea": ["cobertura herbácea", "cobertura herbacea"],
        "dt03_cobertura_hojarasca": ["cobertura de hojarasca"],
        "dt03_suelo_desnudo": ["cobertura de suelo desnudo"],
        "dt03_altura_estrato_dom": ["altura promedio estrato"],
        "dt03_altura_max": ["altura máxima", "altura maxima"],
        "dt03_dap_promedio": ["dap promedio"],
        "dt03_regeneracion_natural": ["regeneración natural", "regeneracion natural"],
        "dt03_estado_sanitario": ["estado sanitario general"],
        "dt03_presencia_epifitas": ["presencia de epífitas", "presencia de epifitas"],
        "dt03_fenologia_dominante": ["fenología dominante", "fenologia dominante"],
        "dt03_tipo_cobertura_dom": ["tipo de cobertura vegetal dominante"],
    }
    vocab = [kw for kws in label_map.values() for kw in kws]
    anchors = _collect_label_anchors(ws, vocab)
    base = {k: _scan_label_value(ws, kws, anchors=anchors)
            for k, kws in label_map.items()}
    base["dt03_observaciones"] = _scan_observaciones(ws, anchors=anchors)
    # Tabla de floristica (9 columnas: n, nombre_comun, nombre_cient, familia, estrato, origen, abundancia, dap, altura)
    flora = _scan_table(ws, ["n°", "nombre común", "estrato"], num_data_rows=15, num_cols=9)
    flora_keys = ["n", "nombre_comun", "nombre_cientifico", "familia",
                  "estrato", "origen", "abundancia", "dap_cm", "altura_m"]
    base["dt03_floristica"] = [
        {k: v for k, v in zip(flora_keys, row)} for row in flora
    ]
    # Tabla de especies clave (9 columnas)
    esp = _scan_table(ws, ["categoría", "uicn"], num_data_rows=10, num_cols=9)
    esp_keys = ["n", "nombre", "categoria", "estado_uicn", "utm_e",
                "utm_n", "n_indiv", "foto", "observacion"]
    base["dt03_especies_clave"] = [
        {k: v for k, v in zip(esp_keys, row)} for row in esp
    ]
    return base


def _parse_dt04(ws):
    label_map = {
        "fecha_evaluacion": ["fecha"],
        "evaluador": ["responsable", "evaluador"],
        "codigo_bloque": ["código del bloque", "codigo del bloque"],
        "microcuenca": ["microcuenca"],
        "dt04_causas_directas_texto": ["principales causas directas"],
        "dt04_causa_subyacente": ["causa subyacente"],
        "dt04_velocidad_degradacion": ["velocidad de degradación", "velocidad de degradacion"],
        "dt04_reversibilidad": ["reversibilidad técnica", "reversibilidad tecnica"],
        "dt04_urgencia_intervencion": ["urgencia de intervención", "urgencia de intervencion"],
    }
    vocab = [kw for kws in label_map.values() for kw in kws]
    anchors = _collect_label_anchors(ws, vocab)
    base = {k: _scan_label_value(ws, kws, anchors=anchors)
            for k, kws in label_map.items()}
    base["dt04_observaciones"] = _scan_observaciones(ws, anchors=anchors)
    # Matriz de causas (7 cols): n, causa, presencia, intensidad, extension, antiguedad, evidencia
    causas = _scan_table(ws, ["causa", "presencia", "intensidad"],
                         num_data_rows=16, num_cols=7)
    causas_keys = ["n", "causa", "presencia", "intensidad", "extension",
                   "antiguedad", "evidencia"]
    base["dt04_causas"] = [
        {k: v for k, v in zip(causas_keys, row)} for row in causas
    ]
    # Indicadores cuantitativos (7 cols): n, indicador, unidad, valor, fuente, umbral, nivel
    inds = _scan_table(ws, ["indicador", "unidad", "valor"],
                       num_data_rows=8, num_cols=7)
    ind_keys = ["n", "indicador", "unidad", "valor", "fuente", "umbral", "nivel"]
    base["dt04_indicadores"] = [
        {k: v for k, v in zip(ind_keys, row)} for row in inds
    ]
    return base


def _parse_dt05(ws):
    label_map = {
        "fecha_evaluacion": ["fecha"],
        "evaluador": ["responsable", "evaluador"],
        "codigo_bloque": ["código del bloque", "codigo del bloque"],
        "microcuenca": ["microcuenca"],
        "distrito": ["distrito"],
        "provincia": ["provincia"],
        "dt05_zona_recarga": ["zona de recarga"],
        "dt05_humedad_persistente": ["humedad persistente"],
        "dt05_escorrentia_concentrada": ["escorrentía concentrada", "escorrentia concentrada"],
        "dt05_dist_captacion": ["distancia a captación", "distancia a captacion"],
        "dt05_jass_captacion": ["jass", "captación asociada", "captacion asociada"],
        "dt05_interferencia_riego": ["interferencia con obras"],
        "dt05_sistema_riego_nombre": ["sistema de riego"],
        "dt05_modalidad_acceso": ["modalidad de acceso"],
        "dt05_via_principal": ["vía principal", "via principal"],
        "dt05_tipo_via_final": ["tipo de vía final", "tipo de via final"],
        "dt05_transitabilidad_seca": ["transitabilidad — época seca", "transitabilidad - epoca seca"],
        "dt05_transitabilidad_lluviosa": ["transitabilidad — época lluviosa", "transitabilidad - epoca lluviosa"],
        "dt05_tiempo_dist_capital": ["tiempo desde capital distrital"],
        "dt05_tiempo_prov_capital": ["tiempo desde capital provincial"],
        "dt05_senal_celular": ["señal celular", "senal celular"],
        "dt05_operador_celular": ["operador celular"],
        "dt05_alojamiento": ["alojamiento rural"],
        "dt05_requiere_ronda": ["autorización de ronda", "autorizacion de ronda"],
        "dt05_contacto_ronda": ["contacto responsable de ronda"],
    }
    vocab = [kw for kws in label_map.values() for kw in kws]
    anchors = _collect_label_anchors(ws, vocab)
    base = {k: _scan_label_value(ws, kws, anchors=anchors)
            for k, kws in label_map.items()}
    base["dt05_observaciones"] = _scan_observaciones(ws, anchors=anchors)
    # Inventario de fuentes de agua (8 cols)
    fuentes = _scan_table(ws, ["tipo de fuente", "régimen"],
                          num_data_rows=10, num_cols=8)
    if not fuentes:
        fuentes = _scan_table(ws, ["tipo de fuente", "regimen"],
                              num_data_rows=10, num_cols=8)
    fuente_keys = ["n", "tipo", "utm_e", "utm_n", "regimen", "calidad",
                   "distancia_m", "uso_obs"]
    base["dt05_fuentes_agua"] = [
        {k: v for k, v in zip(fuente_keys, row)} for row in fuentes
    ]
    return base


_PARSERS = {
    "F-DT-01": _parse_dt01,
    "F-DT-02": _parse_dt02,
    "F-DT-03": _parse_dt03,
    "F-DT-04": _parse_dt04,
    "F-DT-05": _parse_dt05,
}


def parsear_excel_dt(file_bytes, ficha=None):
    """Parsea un archivo Excel V5 y devuelve [{ficha, datos}, ...]."""
    wb = load_workbook(file_bytes, data_only=True)
    resultados = []
    if ficha and ficha in _PARSERS:
        for sn in wb.sheetnames:
            if ficha.lower().replace("-", "") in sn.lower().replace("-", ""):
                resultados.append({"ficha": ficha, "datos": _PARSERS[ficha](wb[sn])})
                return resultados
        return resultados
    for sn in wb.sheetnames:
        ficha_det = None
        sn_clean = sn.lower().replace("-", "").replace(" ", "")
        for fid in _PARSERS:
            if fid.lower().replace("-", "") in sn_clean:
                ficha_det = fid
                break
        if not ficha_det:
            for r in range(1, 6):
                cv = str(wb[sn].cell(row=r, column=1).value or "")
                for fid in _PARSERS:
                    if fid in cv:
                        ficha_det = fid
                        break
                if ficha_det:
                    break
        if ficha_det:
            resultados.append({"ficha": ficha_det,
                               "datos": _PARSERS[ficha_det](wb[sn])})
    return resultados


# ─── Mapeo a session_state del aplicativo ─────────────────────────────────

def mapear_dt_a_session_state(datos_parseados, bloques_map):
    """Convierte el resultado del parser a un dict {session_state_key: valor}
    listo para inyectar en st.session_state. `datos_parseados` debe traer
    'ficha' (puede ser una concatenacion 'F-DT-01, F-DT-02, ...') y
    'datos' (dict con todos los campos de todas las fichas detectadas)."""
    import json as _json
    datos = datos_parseados.get("datos", {}) or {}
    ss = {}

    # Datos generales
    if datos.get("evaluador"):
        ss["dt_eval"] = datos["evaluador"]
    if datos.get("ficha_correlativo"):
        ss["dt_corr"] = datos["ficha_correlativo"]
    if datos.get("hora_registro"):
        ss["dt_hora"] = datos["hora_registro"]
    if datos.get("altitud_gps"):
        ss["dt_altitud"] = datos["altitud_gps"]
    if datos.get("centro_poblado_cercano"):
        ss["dt_cp"] = datos["centro_poblado_cercano"]
    if datos.get("comunidad_campesina_dt"):
        ss["dt_cc"] = datos["comunidad_campesina_dt"]
    if datos.get("utm_este_dt"):
        ss["dt_utm_e"] = datos["utm_este_dt"]
    if datos.get("utm_norte_dt"):
        ss["dt_utm_n"] = datos["utm_norte_dt"]

    fecha_str = str(datos.get("fecha_evaluacion", ""))
    if fecha_str:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                ss["dt_fecha"] = datetime.strptime(fecha_str.split(" ")[0], fmt).date()
                break
            except (ValueError, TypeError):
                continue

    codigo_bloque = datos.get("codigo_bloque", "")
    if codigo_bloque:
        for label in bloques_map:
            if codigo_bloque in label:
                ss["dt_bl"] = label
                break

    if datos.get("microcuenca"):
        ss["dt_mc"] = datos["microcuenca"]

    # ── F-DT-01 ──
    _set_if(ss, "f01_ft", datos.get("forma_terreno"))
    _set_if(ss, "f01_pe", datos.get("pendiente"))
    _set_if(ss, "f01_pf", datos.get("posicion_fisiografica"))
    _set_if(ss, "f01_ex", datos.get("exposicion_orientacion"))
    _set_if(ss, "f01_ra", datos.get("rango_altitudinal"))
    _set_if(ss, "f01_pa", datos.get("paisaje_dominante"))
    _set_if(ss, "f01_af", datos.get("dt01_afloramientos_rocosos"))
    _set_if(ss, "f01_es", datos.get("dt01_escarpes_activos"))
    _set_if(ss, "f01_rp", datos.get("dt01_reptacion_suelo"))
    _set_if(ss, "f01_de", datos.get("dt01_deslizamientos_antiguos"))
    _set_if(ss, "f01_rm", datos.get("dt01_remociones_masa_activas"))
    _set_if(ss, "f01_obs", datos.get("dt01_observaciones"))

    # ── F-DT-02 ──
    _set_if(ss, "f02_sell", datos.get("dt02_sellamiento_costra"))
    _set_if(ss, "f02_compa", datos.get("dt02_compactacion_pisoteo"))
    _set_if(ss, "f02_raices", datos.get("dt02_raices_expuestas"))
    _set_if(ss, "f02_nivel", datos.get("dt02_nivel_erosion_general"))
    _set_if(ss, "f02_sint", datos.get("dt02_nivel_erosion_sintesis"))
    _set_if(ss, "f02_num", datos.get("dt02_num_carcavas"))
    _set_if(ss, "f02_long", datos.get("dt02_longitud_total_carcavas"))
    _set_if(ss, "f02_pct", datos.get("dt02_pct_bloque_carcavas"))
    _set_if(ss, "f02_lam", datos.get("dt02_erosion_laminar_pct"))
    _set_if(ss, "f02_pat", datos.get("dt02_patron_carcavas"))
    _set_if(ss, "f02_soc", datos.get("dt02_socavamiento_cauce"))
    _set_if(ss, "f02_urg", datos.get("dt02_urgencia_control"))
    _set_if(ss, "f02_obs", datos.get("dt02_observaciones"))

    # ── F-DT-03 ──
    _set_if(ss, "f03_parcela", datos.get("dt03_parcela_muestreo"))
    _set_if(ss, "f03_dim", datos.get("dt03_dim_parcela"))
    _set_if(ss, "f03_pend", datos.get("dt03_pendiente_parcela"))
    _set_if(ss, "f03_cobtot", datos.get("dt03_cobertura_total"))
    _set_if(ss, "f03_uso", datos.get("dt03_uso_dominante"))
    _set_if(ss, "f03_eco", datos.get("dt03_tipo_ecosistema"))
    _set_if(ss, "f03_supe", datos.get("dt03_superficie_ecosistema"))
    _set_if(ss, "f03_cons", datos.get("dt03_estado_conservacion_eco"))
    _set_if(ss, "f03_dosel", datos.get("dt03_cobertura_dosel"))
    _set_if(ss, "f03_arbus", datos.get("dt03_cobertura_arbustiva"))
    _set_if(ss, "f03_herb", datos.get("dt03_cobertura_herbacea"))
    _set_if(ss, "f03_hoja", datos.get("dt03_cobertura_hojarasca"))
    _set_if(ss, "f03_desn", datos.get("dt03_suelo_desnudo"))
    _set_if(ss, "f03_haltd", datos.get("dt03_altura_estrato_dom"))
    _set_if(ss, "f03_hmax", datos.get("dt03_altura_max"))
    _set_if(ss, "f03_dap", datos.get("dt03_dap_promedio"))
    _set_if(ss, "f03_regen", datos.get("dt03_regeneracion_natural"))
    _set_if(ss, "f03_san", datos.get("dt03_estado_sanitario"))
    _set_if(ss, "f03_epif", datos.get("dt03_presencia_epifitas"))
    _set_if(ss, "f03_feno", datos.get("dt03_fenologia_dominante"))
    _set_if(ss, "f03_tipocob", datos.get("dt03_tipo_cobertura_dom"))
    _set_if(ss, "f03_obs", datos.get("dt03_observaciones"))

    # ── F-DT-04 ──
    _set_if(ss, "f04_dir", datos.get("dt04_causas_directas_texto"))
    _set_if(ss, "f04_sub", datos.get("dt04_causa_subyacente"))
    _set_if(ss, "f04_vel", datos.get("dt04_velocidad_degradacion"))
    _set_if(ss, "f04_rev", datos.get("dt04_reversibilidad"))
    _set_if(ss, "f04_urg", datos.get("dt04_urgencia_intervencion"))
    _set_if(ss, "f04_obs", datos.get("dt04_observaciones"))

    # ── F-DT-05 ──
    _set_if(ss, "f05_recarga", datos.get("dt05_zona_recarga"))
    _set_if(ss, "f05_humedad", datos.get("dt05_humedad_persistente"))
    _set_if(ss, "f05_escor", datos.get("dt05_escorrentia_concentrada"))
    _set_if(ss, "f05_distcap", datos.get("dt05_dist_captacion"))
    _set_if(ss, "f05_jass", datos.get("dt05_jass_captacion"))
    _set_if(ss, "f05_inter", datos.get("dt05_interferencia_riego"))
    _set_if(ss, "f05_riego", datos.get("dt05_sistema_riego_nombre"))
    _set_if(ss, "f05_modo", datos.get("dt05_modalidad_acceso"))
    _set_if(ss, "f05_via", datos.get("dt05_via_principal"))
    _set_if(ss, "f05_tvia", datos.get("dt05_tipo_via_final"))
    _set_if(ss, "f05_tseca", datos.get("dt05_transitabilidad_seca"))
    _set_if(ss, "f05_tllu", datos.get("dt05_transitabilidad_lluviosa"))
    _set_if(ss, "f05_tcap", datos.get("dt05_tiempo_dist_capital"))
    _set_if(ss, "f05_tprov", datos.get("dt05_tiempo_prov_capital"))
    _set_if(ss, "f05_senal", datos.get("dt05_senal_celular"))
    _set_if(ss, "f05_oper", datos.get("dt05_operador_celular"))
    _set_if(ss, "f05_aloj", datos.get("dt05_alojamiento"))
    _set_if(ss, "f05_ronda", datos.get("dt05_requiere_ronda"))
    _set_if(ss, "f05_contacto", datos.get("dt05_contacto_ronda"))
    _set_if(ss, "f05_obs", datos.get("dt05_observaciones"))

    # Las tablas no se inyectan al session_state directamente porque
    # st.data_editor maneja su propio estado; el usuario debera revisar
    # las tablas manualmente al pasar a la pestana de Registro.

    return ss


def _set_if(ss, key, value):
    if value is not None and str(value).strip():
        ss[key] = str(value).strip()

"""
Regenera `centros_poblados.py` a partir de los Excel INEI del repositorio.

Es idempotente: toma la relacion de centros poblados y comunidades campesinas
del propio `centros_poblados.py` (que es la version curada, con la ortografia
oficial de cada nombre) y vuelve a derivar la demografia desde el Excel. Asi,
cuando llegue una fuente actualizada basta con reemplazar el Excel y ejecutar:

    python generar_centros_poblados.py

Para incorporar bloques o centros poblados NUEVOS, agreguelos primero a
`centros_poblados.py` (nombre del centro poblado y bloque al que pertenece) y
luego ejecute este script para que complete su demografia.
"""

import io
import os
import re
import sys
import unicodedata

import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_DEMOGRAFIA = os.path.join(BASE, "CentrosPoblados-INEI-Bloques_Adicionales.xlsx")
HOJA_PRINCIPAL = "centrospobladosineibloques_v5"
HOJA_ADICIONALES = "Hoja2"          # bloques 83 a 87
DESTINO = os.path.join(BASE, "centros_poblados.py")


def norm(s):
    """Nombre comparable: sin acentos, sin puntuacion y en minusculas."""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", " ", s).strip()


def cod_bloque(v):
    """'BLOQUE - M3B6' -> 'M3B6'."""
    return re.sub(r"^BLOQUE\s*-?\s*", "", str(v).strip(), flags=re.I).strip()


def entero(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def decimal(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return 0.0


def leer_demografia():
    """Devuelve (por_bloque, unicos_por_nombre) leidos del Excel INEI."""
    wb = openpyxl.load_workbook(EXCEL_DEMOGRAFIA, data_only=True)
    filas = list(wb[HOJA_PRINCIPAL].iter_rows(values_only=True))
    ix = {h: i for i, h in enumerate(filas[0])}

    def registro(r):
        return {
            "poblacion_total": entero(r[ix["POBTOTAL"]]),
            "hombres": entero(r[ix["TOTHOMBRES"]]),
            "mujeres": entero(r[ix["TOTMUJERES"]]),
            "poblacion_vulnerable": entero(r[ix["POBVULNERA"]]),
            "viviendas": entero(r[ix["VIV_PARTIC"]]),
            "tipo": str(r[ix["TIPO"]] or ""),
            "utm_este": decimal(r[ix["ESTE_UTM"]]),
            "utm_norte": decimal(r[ix["NORTE_UTM"]]),
        }

    por_bloque, por_nombre = {}, {}
    for r in filas[1:]:
        if not r[ix["BLOQUE"]] or not r[ix["NOMB_CCPP"]]:
            continue
        por_bloque.setdefault(cod_bloque(r[ix["BLOQUE"]]), {})[norm(r[ix["NOMB_CCPP"]])] = registro(r)
        por_nombre.setdefault(norm(r[ix["NOMB_CCPP"]]), []).append(r)

    # Un centro poblado cuyo nombre corresponde a un unico IDCCPP sirve para
    # completar los bloques donde la fuente trae la fila vacia.
    unicos = {}
    for nombre_n, rs in por_nombre.items():
        if len({str(r[ix["IDCCPP"]]) for r in rs}) == 1:
            unicos[nombre_n] = registro(rs[0])

    # Bloques 83-87: la hoja auxiliar solo trae poblacion total.
    for r in wb[HOJA_ADICIONALES].iter_rows(values_only=True):
        if not r or not r[5] or not r[4]:
            continue
        por_bloque.setdefault(cod_bloque(r[5]), {})[norm(r[4])] = {
            "poblacion_total": entero(r[3]), "hombres": 0, "mujeres": 0,
            "poblacion_vulnerable": 0, "viviendas": 0, "tipo": "Rural",
            "utm_este": 0.0, "utm_norte": 0.0,
        }
    return por_bloque, unicos


def construir(catalogo, por_bloque, unicos):
    salida, sin_datos, inferidos = {}, [], []
    for codigo, info in catalogo.items():
        nombres = list(info.get("centros_poblados", []))
        del_bloque = por_bloque.get(codigo, {})
        detalle = []
        for nombre in nombres:
            d = del_bloque.get(norm(nombre))
            inferido = False
            if d is None:
                d = unicos.get(norm(nombre))
                inferido = d is not None
            if d is None:
                sin_datos.append((codigo, nombre))
                continue
            fila = dict(centro_poblado=nombre, **d)
            if inferido:
                fila["inferido"] = True
                inferidos.append((codigo, nombre))
            detalle.append(fila)
        salida[codigo] = {
            "centros_poblados": nombres,
            "comunidades_campesinas": list(info.get("comunidades_campesinas", [])),
            "poblacion_total": sum(d["poblacion_total"] for d in detalle),
            "hombres": sum(d["hombres"] for d in detalle),
            "mujeres": sum(d["mujeres"] for d in detalle),
            "poblacion_vulnerable": sum(d["poblacion_vulnerable"] for d in detalle),
            "viviendas": sum(d["viviendas"] for d in detalle),
            "demografia": detalle,
        }
    return salida, sin_datos, inferidos


ENCABEZADO = '''"""
IN Piura - Catalogo de Centros Poblados por Bloque de Intervencion.

ARCHIVO GENERADO por `generar_centros_poblados.py`. Para actualizar la
demografia, reemplace el Excel de origen y vuelva a ejecutar ese script; para
agregar un bloque o un centro poblado nuevo, escribalo aqui (nombre y bloque)
y luego ejecute el script para que complete sus datos.

Fuentes:
  - CentrosPoblados-INEI-Bloques_v5.xlsx (hoja "Lista CPs INEI Bloques V5"):
    relacion de centros poblados y comunidades campesinas por bloque.
  - CentrosPoblados-INEI-Bloques_Adicionales.xlsx (hoja
    "centrospobladosineibloques_v5"): poblacion total y por sexo, poblacion
    vulnerable, viviendas particulares y coordenadas de cada centro poblado;
    hoja "Hoja2" para los bloques 83 a 87, que solo trae poblacion total.

Estructura de cada entrada:
    "<codigo de bloque>": {
        "centros_poblados": [...],        # nombres, en el orden de la fuente
        "comunidades_campesinas": [...],
        "poblacion_total": int,           # suma de los CCPP del bloque
        "hombres": int, "mujeres": int,
        "poblacion_vulnerable": int, "viviendas": int,
        "demografia": [ {datos por centro poblado} ],
    }

Un mismo centro poblado puede estar asociado a varios bloques: los totales son
"poblacion de los centros poblados asociados al bloque", no poblacion residente
dentro del poligono, y no deben sumarse entre bloques.

Un centro poblado con "inferido": True es aquel cuya fila venia vacia en su
bloque y cuyos datos se tomaron de otro bloque donde el mismo centro poblado
(mismo IDCCPP) si figura completo.
"""

CENTROS_POBLADOS_BLOQUE = {'''

PIE = '''}


def datos_bloque(codigo):
    """Entrada del catalogo para un bloque, o una vacia si no figura."""
    return CENTROS_POBLADOS_BLOQUE.get(codigo, {
        "centros_poblados": [], "comunidades_campesinas": [],
        "poblacion_total": 0, "hombres": 0, "mujeres": 0,
        "poblacion_vulnerable": 0, "viviendas": 0, "demografia": [],
    })
'''


def escribir(salida):
    def orden(c):
        return (0, int(c), "") if c.isdigit() else (1, 0, c)

    def lista(xs):
        return "[%s]" % ", ".join('"%s"' % x for x in xs)

    lineas = [ENCABEZADO]
    for codigo in sorted(salida, key=orden):
        v = salida[codigo]
        lineas.append('    "%s": {' % codigo)
        lineas.append('        "centros_poblados": %s,' % lista(v["centros_poblados"]))
        lineas.append('        "comunidades_campesinas": %s,' % lista(v["comunidades_campesinas"]))
        lineas.append('        "poblacion_total": %d, "hombres": %d, "mujeres": %d,'
                      % (v["poblacion_total"], v["hombres"], v["mujeres"]))
        lineas.append('        "poblacion_vulnerable": %d, "viviendas": %d,'
                      % (v["poblacion_vulnerable"], v["viviendas"]))
        if not v["demografia"]:
            lineas.append('        "demografia": [],')
        else:
            lineas.append('        "demografia": [')
            for d in v["demografia"]:
                lineas.append('            {"centro_poblado": "%s", "poblacion_total": %d, '
                              '"hombres": %d, "mujeres": %d,'
                              % (d["centro_poblado"], d["poblacion_total"],
                                 d["hombres"], d["mujeres"]))
                lineas.append('             "poblacion_vulnerable": %d, "viviendas": %d, "tipo": "%s",'
                              % (d["poblacion_vulnerable"], d["viviendas"], d["tipo"]))
                lineas.append('             "utm_este": %r, "utm_norte": %r%s,'
                              % (d["utm_este"], d["utm_norte"],
                                 ', "inferido": True}' if d.get("inferido") else "}"))
            lineas.append('        ],')
        lineas.append('    },')
    lineas.append(PIE)
    io.open(DESTINO, "w", encoding="utf-8").write("\n".join(lineas))


def main():
    sys.path.insert(0, BASE)
    from centros_poblados import CENTROS_POBLADOS_BLOQUE as catalogo

    por_bloque, unicos = leer_demografia()
    salida, sin_datos, inferidos = construir(catalogo, por_bloque, unicos)
    escribir(salida)

    print(f"Bloques: {len(salida)}")
    print(f"Centros poblados con demografia: "
          f"{sum(len(v['demografia']) for v in salida.values())}")
    print(f"Poblacion agregada (con duplicidad entre bloques): "
          f"{sum(v['poblacion_total'] for v in salida.values()):,}")
    if inferidos:
        print(f"Completados desde otro bloque: {inferidos}")
    if sin_datos:
        print(f"SIN demografia en la fuente: {sin_datos}")
    print(f"Escrito {DESTINO}")


if __name__ == "__main__":
    main()

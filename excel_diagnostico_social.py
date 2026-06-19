"""
Modulo de importacion/exportacion Excel para Diagnostico Social V3.
Proyecto IN Piura CUI 2669244 | ANIN - DIME - SESDI

Reemplaza los 5 formatos anteriores (F-DS-01..05 generados por codigo) por la
plantilla oficial validada V3 con 7 fichas (F-DS-01..07), hojas auxiliares
(_Listas, _Codigos, _Datos) y celdas de validacion (desplegables) nativas.

Estrategia (igual que Diagnostico Territorial V5): se sirve el archivo .xlsx
oficial que vive en el repo y solo se inyecta la lista de bloques actual en la
hoja oculta `_Datos`, conservando intactas todas las listas de validacion.

API publica (estable):
    generar_plantilla_ds(fichas=None, bloques_data=None) -> bytes
    parsear_excel_ds(file_bytes, ficha=None) -> list[dict]
    mapear_a_session_state(resultado, bloques_map) -> dict
"""

import io
import os
import re
import zipfile
import warnings
from datetime import datetime
from xml.sax.saxutils import escape

from openpyxl import load_workbook


# ─── Plantilla oficial (vive junto al codigo en el repo) ───────────────────
# Version actual: V4. Se mantiene el alias PLANTILLA_V3_PATH por compatibilidad.
PLANTILLA_DS_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "Plantilla_Diagnostico_Social_IN_Piura_V4.xlsx",
)
PLANTILLA_V3_PATH = PLANTILLA_DS_PATH

FICHAS_HOJAS = ["F-DS-01", "F-DS-02", "F-DS-03", "F-DS-04",
                "F-DS-05", "F-DS-06", "F-DS-07"]


# ─── Generacion de la plantilla descargable ────────────────────────────────

_DATOS_HEADER = ["codigo", "microcuenca", "provincia", "distrito",
                 "zona", "este_utm", "norte_utm", "area_ha"]


def _col_letter(idx0):
    """Letra de columna 0-based (0->A)."""
    s = ""
    n = idx0
    while True:
        s = chr(ord("A") + n % 26) + s
        n = n // 26 - 1
        if n < 0:
            break
    return s


def _build_datos_sheetdata(bloques_data):
    """Construye el bloque <sheetData> de la hoja _Datos (cabecera + filas),
    usando cadenas en linea para no depender de sharedStrings."""
    def cell(col0, row, val):
        ref = f"{_col_letter(col0)}{row}"
        txt = escape("" if val is None else str(val))
        return (f'<c r="{ref}" t="inlineStr"><is>'
                f'<t xml:space="preserve">{txt}</t></is></c>')
    rows_xml = []
    # Fila 1: cabecera
    cells = "".join(cell(c, 1, h) for c, h in enumerate(_DATOS_HEADER))
    rows_xml.append(f'<row r="1">{cells}</row>')
    # Filas de datos
    for i, b in enumerate(bloques_data, start=2):
        vals = list(b) + [""] * (len(_DATOS_HEADER) - len(b))
        cells = "".join(cell(c, i, vals[c]) for c in range(len(_DATOS_HEADER)))
        rows_xml.append(f'<row r="{i}">{cells}</row>')
    return "<sheetData>" + "".join(rows_xml) + "</sheetData>"


def _resolver_hoja_datos(zf):
    """Devuelve la ruta del XML de la hoja `_Datos` resolviendo los rels."""
    wb_xml = zf.read("xl/workbook.xml").decode("utf-8", "ignore")
    m = re.search(r'<sheet[^>]*name="_Datos"[^>]*r:id="(rId\d+)"', wb_xml)
    if not m:
        return None
    rid = m.group(1)
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8", "ignore")
    m2 = re.search(rf'<Relationship[^>]*Id="{rid}"[^>]*Target="([^"]+)"', rels)
    if not m2:
        return None
    target = m2.group(1).lstrip("/")
    if not target.startswith("xl/"):
        target = "xl/" + target
    return target


def generar_plantilla_ds(fichas=None, bloques_data=None):
    """Devuelve los bytes del .xlsx oficial V3.

    Para conservar intactas TODAS las celdas de validacion (desplegables x14
    que referencian la hoja `_Listas`), NO se reabre el libro con openpyxl
    (openpyxl elimina esas validaciones al guardar). En su lugar se reescribe
    a nivel ZIP unicamente el XML de la hoja oculta `_Datos` con la lista de
    bloques actual; el resto del archivo se conserva byte a byte.

    Args:
        fichas: ignorado (la plantilla trae siempre las 7 fichas).
        bloques_data: lista de tuplas
            (codigo, microcuenca, provincia, distrito[, zona, este, norte, area]).

    Returns:
        bytes con el contenido del .xlsx.
    """
    if not os.path.exists(PLANTILLA_V3_PATH):
        raise FileNotFoundError(
            f"No se encontro la plantilla V3 oficial: {PLANTILLA_V3_PATH}")

    with open(PLANTILLA_V3_PATH, "rb") as fh:
        raw = fh.read()

    if not bloques_data:
        return raw

    zin = zipfile.ZipFile(io.BytesIO(raw))
    datos_path = _resolver_hoja_datos(zin)
    if not datos_path or datos_path not in zin.namelist():
        return raw  # sin _Datos no hay nada que inyectar; servir tal cual

    sheet_xml = zin.read(datos_path).decode("utf-8", "ignore")
    nuevo_sheetdata = _build_datos_sheetdata(bloques_data)
    if "<sheetData" in sheet_xml:
        sheet_xml = re.sub(r"<sheetData[^>]*>.*?</sheetData>|<sheetData[^>]*/>",
                           nuevo_sheetdata, sheet_xml, count=1, flags=re.S)
    # Recalcular dimension para abarcar las filas escritas
    last_row = len(bloques_data) + 1
    sheet_xml = re.sub(r'<dimension ref="[^"]*"/>',
                       f'<dimension ref="A1:H{last_row}"/>', sheet_xml, count=1)

    buf_out = io.BytesIO()
    with zipfile.ZipFile(buf_out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == datos_path:
                data = sheet_xml.encode("utf-8")
            zout.writestr(item, data)
    buf_out.seek(0)
    return buf_out.getvalue()


# ─── Parseo de la plantilla V3 llenada ─────────────────────────────────────

def _clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s in ("▼", "-"):
        return ""
    return s


def _find_label(ws, substr, max_row=None):
    """Devuelve (row, col) de la primera celda cuyo texto contiene substr."""
    substr = substr.lower()
    mr = max_row or ws.max_row
    for r in range(1, mr + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and substr in v.lower():
                return r, c
    return None, None


def _value_right(ws, row, col, span=8):
    """Primer valor no vacio a la derecha de (row, col) en la misma fila."""
    if row is None:
        return ""
    for c in range(col + 1, min(col + 1 + span, ws.max_column + 1)):
        v = _clean(ws.cell(row=row, column=c).value)
        if v:
            return v
    return ""


def _read_table(ws, row_ini, row_fin, colmap):
    """Lee filas row_ini..row_fin. colmap: {nombre_columna: indice_col_1based}.
    Devuelve lista de dicts (solo filas con algun valor)."""
    out = []
    for r in range(row_ini, row_fin + 1):
        rec = {}
        for name, col in colmap.items():
            rec[name] = _clean(ws.cell(row=r, column=col).value)
        if any(v for v in rec.values()):
            out.append(rec)
    return out


def _parse_header(ws):
    """Campos comunes de cabecera (fecha, responsable, bloque, distrito, mc)."""
    r, c = _find_label(ws, "Fecha")
    fecha = _value_right(ws, r, c)
    r, c = _find_label(ws, "Responsable")
    evaluador = _value_right(ws, r, c)
    r, c = _find_label(ws, "Codigo del Bloque")
    if r is None:
        r, c = _find_label(ws, "Código Bloque")
    if r is None:
        r, c = _find_label(ws, "Codigo Bloque")
    codigo = _value_right(ws, r, c)
    r, c = _find_label(ws, "Distrito")
    distrito = _value_right(ws, r, c)
    r, c = _find_label(ws, "Microcuenca")
    microcuenca = _value_right(ws, r, c)
    return {
        "fecha": fecha, "evaluador": evaluador, "codigo_bloque": codigo,
        "distrito": distrito, "microcuenca": microcuenca,
    }


# Mapas de tablas por ficha (filas y columnas fijas de la plantilla V3)
def _parse_fds01(ws):
    form = {}
    # 1. Datos del entrevistado (fila 12): los rotulos estan en celdas
    # combinadas y el valor cae en la primera celda del bloque siguiente:
    # "Nombres y apellidos del entrevistado" (A12:C12) -> D12,
    # "DNI" (E12:F12) -> G12, "Oficio/ocupacion" (I12:J12) -> K12.
    form["f1_entrevistado"] = _clean(ws.cell(row=12, column=4).value)
    form["f1_entrevistado_dni"] = _clean(ws.cell(row=12, column=7).value)
    form["f1_entrevistado_oficio"] = _clean(ws.cell(row=12, column=11).value)
    # El destino (cols I..M) es una grilla de marcas "X"; se omite en el parseo
    # automatico y queda para completar en el aplicativo.
    # Plantilla V4: la tabla bajo 2 filas (98->100) y se corrieron las columnas
    # "N fam." (6->5) y "Productos principales" (7->6) respecto de V3.
    form["f1_activ"] = _read_table(ws, 100, 107, {
        "Actividad / Rubro": 2, "N fam.": 5, "Productos principales": 6,
        "Ingreso (S/./mes)": 14,
    })
    return form


def _parse_fds02(ws):
    form = {}
    form["f2_actores"] = _read_table(ws, 16, 35, {
        "Nombre del actor / Organizacion": 2, "Tipo": 3,
        "Rol / Funcion frente al proyecto": 4, "Influencia": 5, "Interes": 6,
        "Posicion": 7, "Nivel territorial": 8, "Telefono": 9,
        "Correo / Contacto": 10, "Observaciones / Historial": 11,
    })
    return form


def _parse_fds03(ws):
    return {}


def _parse_fds04(ws):
    form = {}
    form["f4_part"] = _read_table(ws, 35, 54, {
        "Nombres y Apellidos": 2, "DNI": 4, "Institucion / Comunidad": 5,
        "Cargo / Rol": 7, "Telefono": 9, "Sexo": 11, "Edad": 12,
    })
    return form


def _parse_fds05(ws):
    form = {}
    form["f5_conflictos"] = _read_table(ws, 14, 21, {
        "Tipo": 2, "Actores involucrados": 3, "Estado": 6, "Antiguedad": 7,
        "Descripcion / Causa raiz": 8, "Impacto potencial en el proyecto": 11,
    })
    form["f5_oportunidades"] = _read_table(ws, 36, 45, {
        "Oportunidad identificada": 2, "Actores relacionados": 5,
        "Tipo (alianza / plataforma / proy.)": 8, "Potencial": 10,
        "Como aprovecharla": 11,
    })
    return form


def _parse_fds06(ws):
    form = {}
    form["f6_peligros"] = _read_table(ws, 15, 26, {
        "Peligro observado": 2, "¿Ocurre?": 3, "Frecuencia": 5, "Magnitud": 7,
        "Tendencia": 9, "Ultimo evento (año)": 11, "Principales daños observados": 12,
    })
    form["f6_cambios"] = _read_table(ws, 35, 42, {
        "Cambio observado": 2, "¿Se percibe?": 5, "Intensidad": 7,
        "Año aprox. de inicio": 10, "Impacto en la comunidad / territorio": 12,
    })
    return form


def _parse_fds07(ws):
    form = {}
    # Puntos de informacion 3.1..3.10 (columna M=13, filas 27..36)
    puntos = ["f7_info_anin", "f7_info_objetivo", "f7_info_no_minero",
              "f7_info_medidas", "f7_info_temporalidad", "f7_info_voluntaria",
              "f7_info_actualizada", "f7_info_confidencialidad",
              "f7_info_preguntas", "f7_info_material"]
    for idx, key in enumerate(puntos):
        form[key] = _clean(ws.cell(row=27 + idx, column=13).value)
    return form


_PARSERS = {
    "F-DS-01": _parse_fds01, "F-DS-02": _parse_fds02, "F-DS-03": _parse_fds03,
    "F-DS-04": _parse_fds04, "F-DS-05": _parse_fds05, "F-DS-06": _parse_fds06,
    "F-DS-07": _parse_fds07,
}


def parsear_excel_ds(file_bytes, ficha=None):
    """Parsea la plantilla V3 llenada. Devuelve lista de
    {"ficha": str, "datos": {...header..., "form": {...}}}.
    Robusto: nunca lanza por una hoja individual."""
    if hasattr(file_bytes, "read"):
        raw = file_bytes.read()
    else:
        raw = file_bytes
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(io.BytesIO(raw), data_only=True)

    objetivo = [ficha] if ficha else FICHAS_HOJAS
    resultados = []
    for hoja in objetivo:
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        try:
            datos = _parse_header(ws)
            form = _PARSERS[hoja](ws)
            datos["form"] = form
            # Solo incluir si hay algo util
            tiene_datos = any(form.values()) or any(
                datos.get(k) for k in ("fecha", "evaluador", "codigo_bloque"))
            if tiene_datos:
                resultados.append({"ficha": hoja, "datos": datos})
        except Exception:
            continue
    return resultados


# Slots de tablas (deben coincidir con streamlit_app._DS_TABLE_SLOTS)
_TABLE_SLOTS = {
    "f1_activ", "f2_actores", "f4_part", "f4_agenda", "f4_acuerdos",
    "f5_conflictos", "f5_oportunidades", "f6_peligros", "f6_cambios",
}


def mapear_a_session_state(resultado, bloques_map):
    """Convierte un resultado de parseo en el dict de precarga {key: valor}
    consumido por streamlit_app._ds_apply_pending()."""
    ficha = resultado.get("ficha", "")
    datos = resultado.get("datos", {})
    form = datos.get("form", {}) or {}

    pend = {"ds_ficha_sel": ficha}
    if datos.get("evaluador"):
        pend["ds_eval"] = datos["evaluador"]
    # Distrito y microcuenca se auto-resuelven del bloque en el aplicativo;
    # no se importan de la cabecera para evitar valores adyacentes erroneos.

    # Fecha
    fecha_str = str(datos.get("fecha", "") or "").split(" ")[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            pend["ds_fecha"] = datetime.strptime(fecha_str, fmt).date()
            break
        except (ValueError, TypeError):
            continue

    # Resolver bloque por codigo
    codigo = str(datos.get("codigo_bloque", "") or "").strip()
    if codigo and bloques_map:
        for label in bloques_map:
            if codigo and codigo in label:
                pend["ds_bl"] = label
                break

    # Volcar el formulario
    for k, v in form.items():
        if k in _TABLE_SLOTS:
            pend[f"_dsinit_{k}"] = v if isinstance(v, list) else []
        else:
            pend[k] = v
    return pend
